import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def _sanitize_filename(s: str) -> str:
    s = " ".join((s or "").split()).strip()
    s = re.sub(r"[<>:\"/\\\\|?*]+", "_", s)
    s = s.strip(" .")
    return s or "report"


def _load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def _points_by_section(points: dict, quarter_num: int) -> dict[int, str]:
    out: dict[int, str] = {}
    prefix = f"chetvert_{quarter_num}_razdel_"
    for pid, val in (points or {}).items():
        if not isinstance(pid, str) or not pid.startswith(prefix):
            continue
        parts = pid.split("_")
        # chetvert_{q}_razdel_{k}_{row}
        if len(parts) >= 5 and parts[2] == "razdel":
            try:
                sec = int(parts[3])
            except Exception:
                continue
            out[sec] = val
    return out


def _parse_class_liter(class_text: str) -> str:
    """
    From values like '5 «В»' -> '5В'
    """
    s = (class_text or "").replace("«", " ").replace("»", " ").strip()
    m = re.search(r"(\d+)\s*([A-Za-zА-ЯЁӘҒҚҢӨҰҮҺа-яёәғқңөұүһ])?", s)
    if not m:
        return (class_text or "").strip()
    num = m.group(1)
    lit = (m.group(2) or "").upper()
    return f"{num}{lit}".strip()


def _to_number(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("%", "").replace(",", ".")
    try:
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _extract_max_points_from_criteria_html(criteria_html: Path, quarter_num: int) -> dict[int, int]:
    """
    Fallback: parse max points from saved criteria.html:
      id="chetvert_{q}_razdel_{k}_max" value="NN"
    """
    if not criteria_html.exists():
        return {}
    html = criteria_html.read_text(encoding="utf-8", errors="ignore")
    out: dict[int, int] = {}
    pattern = re.compile(rf'id="chetvert_{quarter_num}_razdel_(\d+)_max"[^>]*\svalue="([^"]+)"', re.IGNORECASE)
    for sec_s, val_s in pattern.findall(html):
        try:
            sec = int(sec_s)
            val = int(float(val_s.strip()))
            out[sec] = val
        except Exception:
            continue
    return out


def _clear_rows(ws, start_row: int, end_row: int, cols: list[str]):
    for r in range(start_row, end_row + 1):
        for col in cols:
            ws[f"{col}{r}"].value = None


def _fill_template_page(
    ws,
    *,
    organization_name: str | None,
    class_liter: str | None,
    teacher_fio: str | None,
    page_title: str,
    fio_list: list[str],
    result_list: list,
    max_value,
    mode: str,
):
    """
    Fills a single copied template sheet similarly to the provided extract_to_template.py example.

    mode:
      - 'percent' -> result_list values treated as percent 0..100
      - 'points'  -> percent computed as result/max_value
      - 'grade'   -> result_list values treated as grade (2..5)
    """
    # Header cells (as in example)
    if organization_name:
        ws["B1"].value = organization_name
    if class_liter:
        ws["C3"].value = class_liter
    ws["C4"].value = len(fio_list)
    if teacher_fio:
        ws["C5"].value = teacher_fio
    ws["C6"].value = page_title

    start_row = 8
    max_row = 39  # template is built for up to ~32 students; keep same cleanup window as example

    # Ensure the template column header for max points is present (your note: D7 is the column name).
    if mode == "points":
        ws["D7"].value = "Макс"

    # Fill rows
    count_5 = count_4 = count_3 = count_2 = 0
    # If you want max shown "under D7" regardless of row, prefill D8:D39 with the same max.
    if mode == "points" and max_value not in (None, ""):
        mv = _to_number(max_value)
        mv = int(mv) if mv.is_integer() else mv
        for r in range(start_row, max_row + 1):
            ws[f"D{r}"].value = mv

    for i, fio in enumerate(fio_list):
        r = start_row + i
        if r > max_row:
            break
        ws[f"B{r}"].value = fio

        res = result_list[i] if i < len(result_list) else ""
        ws[f"C{r}"].value = res

        # D - max points already prefilled for points mode; clear for non-points.
        if mode != "points":
            ws[f"D{r}"].value = None

        # Compute percent and grade flags similar to example template logic
        percent = 0.0
        if mode == "percent":
            percent = _to_number(res)
        elif mode == "points":
            rv = _to_number(res)
            mv = _to_number(max_value)
            percent = (rv / mv * 100.0) if mv > 0 else 0.0
        elif mode == "grade":
            g = int(_to_number(res))
            # For grade mode, place flags directly (percent is not meaningful)
            ws[f"E{r}"].value = None
            ws[f"F{r}"].value = 1 if g == 5 else 0
            ws[f"G{r}"].value = 1 if g == 4 else 0
            ws[f"H{r}"].value = 1 if g == 3 else 0
            ws[f"I{r}"].value = 1 if g == 2 else 0
            if g == 5:
                count_5 += 1
                ws[f"L{r}"].value = fio
                ws[f"M{r}"].value = ""
                ws[f"N{r}"].value = ""
            elif g in (3, 4):
                if g == 4:
                    count_4 += 1
                else:
                    count_3 += 1
                ws[f"L{r}"].value = ""
                ws[f"M{r}"].value = fio
                ws[f"N{r}"].value = ""
            elif g == 2:
                count_2 += 1
                ws[f"L{r}"].value = ""
                ws[f"M{r}"].value = ""
                ws[f"N{r}"].value = fio
            else:
                ws[f"L{r}"].value = ""
                ws[f"M{r}"].value = ""
                ws[f"N{r}"].value = ""
            continue

        ws[f"E{r}"].value = round(percent, 2)
        is5 = percent >= 85
        is4 = 65 <= percent < 85
        is3 = 40 <= percent < 65
        is2 = percent < 40

        ws[f"F{r}"].value = 1 if is5 else 0
        ws[f"G{r}"].value = 1 if is4 else 0
        ws[f"H{r}"].value = 1 if is3 else 0
        ws[f"I{r}"].value = 1 if is2 else 0

        if is5:
            count_5 += 1
            ws[f"L{r}"].value = fio
            ws[f"M{r}"].value = ""
            ws[f"N{r}"].value = ""
        elif is4 or is3:
            if is4:
                count_4 += 1
            else:
                count_3 += 1
            ws[f"L{r}"].value = ""
            ws[f"M{r}"].value = fio
            ws[f"N{r}"].value = ""
        else:
            count_2 += 1
            ws[f"L{r}"].value = ""
            ws[f"M{r}"].value = ""
            ws[f"N{r}"].value = fio

    # Clear remaining rows in template range
    last_filled = min(max_row, start_row + len(fio_list) - 1)
    if last_filled < max_row:
        _clear_rows(ws, last_filled + 1, max_row, ["B", "C", "D", "E", "F", "G", "H", "I", "L", "M", "N"])

    # Summary cells (following example positions)
    ws["F41"].value = count_5
    ws["G42"].value = count_4
    ws["H43"].value = count_3
    ws["I44"].value = count_2

    total_students = len(fio_list)
    if total_students > 0:
        quality = ((count_4 + count_5) / total_students) * 100.0
        success_rate = ((count_3 + count_4 + count_5) / total_students) * 100.0
    else:
        quality = 0.0
        success_rate = 0.0
    # In user's example they wrote these to J8 / K8; keep same.
    ws["J8"].value = round(quality, 2)
    ws["K8"].value = round(success_rate, 2)


def _fill_grades_page(
    ws,
    *,
    organization_name: str | None,
    class_liter: str | None,
    teacher_fio: str | None,
    fio_list: list[str],
    nums: list[int],
    grades: list,
    percents: list,
):
    """
    Custom simplified page for "Оценки":
    columns: №, ФИО, Оценка, %, кач-ва, успев
    and remove/clear any "Высокий/..." extra blocks.
    """
    if organization_name:
        ws["B1"].value = organization_name
    if class_liter:
        ws["C3"].value = class_liter
    ws["C4"].value = len(fio_list)
    if teacher_fio:
        ws["C5"].value = teacher_fio
    ws["C6"].value = "Оценки"

    header_row = 7
    start_row = 8
    max_row = 39

    # Header
    ws[f"A{header_row}"].value = "№"
    ws[f"B{header_row}"].value = "ФИО"
    ws[f"C{header_row}"].value = "Оценка"
    ws[f"D{header_row}"].value = "%"
    # Per user: E and F must be empty. Put labels back to J7/K7.
    ws[f"E{header_row}"].value = None
    ws[f"F{header_row}"].value = None
    ws[f"J{header_row}"].value = "кач-ва"
    ws[f"K{header_row}"].value = "успев"

    # Remove/clear everything from L onwards (and keep G-I empty too)
    for col in ["G", "H", "I", "L", "M", "N"]:
        ws[f"{col}{header_row}"].value = None

    count_5 = count_4 = count_3 = count_2 = 0
    qual_flags = []
    succ_flags = []

    for i, fio in enumerate(fio_list):
        r = start_row + i
        if r > max_row:
            break
        num = nums[i] if i < len(nums) else i + 1
        g = grades[i] if i < len(grades) else ""
        p = percents[i] if i < len(percents) else ""

        g_num = int(_to_number(g)) if str(g).strip() != "" else 0
        if g_num == 5:
            count_5 += 1
        elif g_num == 4:
            count_4 += 1
        elif g_num == 3:
            count_3 += 1
        elif g_num == 2:
            count_2 += 1

        ws[f"A{r}"].value = num
        ws[f"B{r}"].value = fio
        ws[f"C{r}"].value = g_num if g_num else g
        ws[f"D{r}"].value = _to_number(p) if str(p).strip() != "" else p
        # E and F must be empty (no per-student flags)
        ws[f"E{r}"].value = None
        ws[f"F{r}"].value = None

        qual = 1 if g_num in (4, 5) else 0
        succ = 1 if g_num in (3, 4, 5) else 0
        qual_flags.append(qual)
        succ_flags.append(succ)

        # Clear everything from L onwards ("Высокий" etc), and keep G-I empty.
        for col in ["G", "H", "I", "L", "M", "N"]:
            ws[f"{col}{r}"].value = None

    # Clear remaining rows
    last_filled = min(max_row, start_row + len(fio_list) - 1)
    if last_filled < max_row:
        _clear_rows(ws, last_filled + 1, max_row, ["A", "B", "C", "D", "E", "F", "G", "H", "I", "L", "M", "N"])

    # Keep summary cells consistent with template example
    ws["F41"].value = count_5
    ws["G42"].value = count_4
    ws["H43"].value = count_3
    ws["I44"].value = count_2

    total = len(fio_list)
    quality = (sum(qual_flags) / total * 100.0) if total else 0.0
    success_rate = (sum(succ_flags) / total * 100.0) if total else 0.0
    ws["J8"].value = round(quality, 2)
    ws["K8"].value = round(success_rate, 2)

    # Also wipe any leftover "table tail" beyond N for the used range (best-effort).
    for r in range(header_row, max_row + 1):
        for col in ["L", "M", "N"]:
            ws[f"{col}{r}"].value = None


def build_report(
    template_path: Path,
    students_path: Path,
    context_path: Path,
    out_dir: Path,
    max_points_path: Path | None = None,
    criteria_html_path: Path | None = None,
    org_name_path: Path | None = None,
) -> Path:
    students = _load_json(students_path)
    ctx = _load_json(context_path)
    max_points = {}
    if max_points_path and max_points_path.exists():
        try:
            mp = _load_json(max_points_path)
            if isinstance(mp, dict):
                # keys might come as strings in JSON
                max_points = {int(k): int(v) for k, v in mp.items()}
        except Exception:
            max_points = {}

    class_name = str(ctx.get("class", "") or "").strip()
    subject = str(ctx.get("subject", "") or "").strip()
    quarter_num = int(students[0]["quarter_num"]) if students else 0
    org_name = ctx.get("org_name")
    if (not org_name) and org_name_path and org_name_path.exists():
        org_name = org_name_path.read_text(encoding="utf-8", errors="ignore").strip() or None
    teacher_fio = ctx.get("profile_name")
    class_liter = _parse_class_liter(class_name)

    # Fallback: if max_points file is missing/empty, parse from criteria.html
    if (not max_points) and quarter_num and criteria_html_path:
        max_points = _extract_max_points_from_criteria_html(criteria_html_path, quarter_num)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = _sanitize_filename(f"{class_name} {subject}".strip())
    out_path = out_dir / f"{out_name}.xlsx"

    wb = load_workbook(template_path)
    template_ws = None
    for name in wb.sheetnames:
        if "шаблон" in name.lower() or "template" in name.lower():
            template_ws = wb[name]
            break
    if template_ws is None:
        template_ws = wb.worksheets[0]

    def mk_sheet(name: str):
        ws = wb.copy_worksheet(template_ws)
        ws.title = name[:31]
        return ws

    # Sort by student number to keep stable ordering
    students_sorted = sorted(students, key=lambda s: int(s.get("num") or 0))
    fio_list = [s.get("fio", "") for s in students_sorted]

    # Decide which pages exist (based on JSON fields / section points)
    # "Формативная оценка" in the table is the numeric value in the 3rd column (we store it as `average`).
    # Max is always 10 (per user).
    has_formative = any((s.get("average") not in (None, "", "0", 0)) for s in students_sorted)
    has_grades = any((s.get("grade") not in (None, "", "0", 0)) for s in students_sorted)

    sec_present: set[int] = set()
    for s in students_sorted:
        sec_present |= set(_points_by_section(s.get("points") or {}, quarter_num).keys())

    # Pages in requested order
    if has_formative:
        ws = mk_sheet("Формативное оценивание")
        formative_vals = [s.get("average") or "" for s in students_sorted]
        _fill_template_page(
            ws,
            organization_name=org_name,
            class_liter=class_liter,
            teacher_fio=teacher_fio,
            page_title="Формативное оценивание",
            fio_list=fio_list,
            result_list=formative_vals,
            max_value=10,
            mode="points",
        )

    for sec in [1, 2, 3]:
        if sec not in sec_present:
            continue
        ws = mk_sheet(f"СОр {sec}")
        vals = []
        for s in students_sorted:
            pts = _points_by_section(s.get("points") or {}, quarter_num)
            vals.append(pts.get(sec, ""))
        _fill_template_page(
            ws,
            organization_name=org_name,
            class_liter=class_liter,
            teacher_fio=teacher_fio,
            page_title=f"СОр {sec}",
            fio_list=fio_list,
            result_list=vals,
            max_value=max_points.get(sec),
            mode="points",
        )

    if 0 in sec_present:
        ws = mk_sheet("СОч")
        vals = []
        for s in students_sorted:
            pts = _points_by_section(s.get("points") or {}, quarter_num)
            vals.append(pts.get(0, ""))
        _fill_template_page(
            ws,
            organization_name=org_name,
            class_liter=class_liter,
            teacher_fio=teacher_fio,
            page_title="СОч",
            fio_list=fio_list,
            result_list=vals,
            max_value=max_points.get(0),
            mode="points",
        )

    if has_grades:
        ws = mk_sheet("Оценки")
        nums = [int(s.get("num") or 0) for s in students_sorted]
        fio_list2 = [s.get("fio", "") for s in students_sorted]
        grade_vals = [s.get("grade") or "" for s in students_sorted]
        percent_vals = [s.get("total_pct") or "" for s in students_sorted]
        _fill_grades_page(
            ws,
            organization_name=org_name,
            class_liter=class_liter,
            teacher_fio=teacher_fio,
            fio_list=fio_list2,
            nums=nums,
            grades=grade_vals,
            percents=percent_vals,
        )

    # Keep original template sheet, rename explicitly
    # Per user: remove template sheet from final output.
    if template_ws.title in wb.sheetnames:
        wb.remove(template_ws)

    wb.save(out_path)
    return out_path


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--template", default="Шаблон.xlsx")
    p.add_argument("--students", default="out/mektep/criteria_students.json")
    p.add_argument("--context", default="out/mektep/criteria_context.json")
    p.add_argument("--maxpoints", default="out/mektep/criteria_max_points.json")
    p.add_argument("--criteriahtml", default="out/mektep/criteria.html")
    p.add_argument("--orgfile", default="out/mektep/org_name.txt")
    p.add_argument("--outdir", default="out/mektep/reports")
    args = p.parse_args()

    out_path = build_report(
        template_path=Path(args.template),
        students_path=Path(args.students),
        context_path=Path(args.context),
        out_dir=Path(args.outdir),
        max_points_path=Path(args.maxpoints) if args.maxpoints else None,
        criteria_html_path=Path(args.criteriahtml) if args.criteriahtml else None,
        org_name_path=Path(args.orgfile) if args.orgfile else None,
    )
    print(f"Saved report: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

