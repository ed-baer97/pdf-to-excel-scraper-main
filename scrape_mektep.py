import argparse
import os
import sys
import getpass
import json
import csv
import re
import time
from pathlib import Path
from urllib.parse import urljoin
from datetime import datetime

from dotenv import load_dotenv
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright
from playwright.sync_api import Page

# Импорт логгера
try:
    from scraper_logger import (
        init_logger,
        get_logger,
        log_stage,
        log_info,
        log_success,
        log_warning,
        log_error,
        log_timing,
        timing_block,
        ScraperLogger,
    )
except ImportError:
    # Fallback если логгер недоступен
    def init_logger(*args, **kwargs):
        """Заглушка инициализации логгера при отсутствии scraper_logger."""
        return None

    def get_logger():
        """Заглушка: возвращает None вместо объекта логгера."""
        return None

    def log_stage(*args, **kwargs):
        """Заглушка: этап скрапинга не логируется."""
        pass

    def log_info(msg):
        """Заглушка: печатает информационное сообщение в консоль."""
        print(f"[INFO] {msg}")

    def log_success(msg):
        """Заглушка: печатает сообщение об успехе в консоль."""
        print(f"[SUCCESS] {msg}")

    def log_warning(msg):
        """Заглушка: печатает предупреждение в консоль."""
        print(f"[WARNING] {msg}")

    def log_error(msg, exc=None):
        """Заглушка: печатает ошибку в консоль."""
        print(f"[ERROR] {msg}")

    def log_timing(label: str, seconds: float) -> None:
        print(f"[TIMING] {label}: {seconds:.2f}s")

    from contextlib import contextmanager

    @contextmanager
    def timing_block(label: str):
        t0 = time.perf_counter()
        try:
            yield
        finally:
            log_timing(label, time.perf_counter() - t0)

    class ScraperLogger:
        STAGE_INIT = "INIT"
        STAGE_BROWSER = "BROWSER"
        STAGE_PAGE_LOAD = "PAGE_LOAD"
        STAGE_LOGIN_FORM = "LOGIN_FORM"
        STAGE_AUTH = "AUTH"
        STAGE_LANGUAGE = "LANGUAGE"
        STAGE_NAVIGATION = "NAVIGATION"
        STAGE_GRADES_TABLE = "GRADES_TABLE"
        STAGE_CRITERIA = "CRITERIA"
        STAGE_STUDENTS = "STUDENTS"
        STAGE_EXCEL_REPORT = "EXCEL_REPORT"
        STAGE_WORD_REPORT = "WORD_REPORT"
        STAGE_COMPLETE = "COMPLETE"
        STAGE_ERROR = "ERROR"


def _timed_sleep(seconds: float, reason: str) -> None:
    """Явный time.sleep с записью фактической длительности в лог [TIMING]."""
    t0 = time.perf_counter()
    time.sleep(seconds)
    log_timing(f"sleep({seconds:g}s): {reason}", time.perf_counter() - t0)


def _debug_artifacts_enabled() -> bool:
    """Управляет сохранением HTML/скриншотов в штатном сценарии."""
    return os.getenv("MEKTEP_DEBUG_ARTIFACTS", "1").strip().lower() not in ("0", "false", "no", "off")


# Fix stdout/stderr for frozen PyInstaller builds (console=False → stdout is None)
if getattr(sys, 'frozen', False):
    import io as _io
    if sys.stdout is None:
        sys.stdout = open(os.devnull, "w", encoding="utf-8")
    if sys.stderr is None:
        sys.stderr = open(os.devnull, "w", encoding="utf-8")
    
    # Playwright driver: в frozen-режиме driver находится внутри _MEIPASS
    _meipass = getattr(sys, '_MEIPASS', None)
    if _meipass:
        _frozen_driver = os.path.join(_meipass, 'playwright', 'driver')
        if os.path.isdir(_frozen_driver):
            os.environ.setdefault('PLAYWRIGHT_DRIVER_PATH', _frozen_driver)

# Fix encoding issues on Windows when printing Unicode characters
if sys.platform == "win32":
    try:
        if sys.stdout is not None and hasattr(sys.stdout, 'encoding') and sys.stdout.encoding != "utf-8":
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if sys.stderr is not None and hasattr(sys.stderr, 'encoding') and sys.stderr.encoding != "utf-8":
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        import io
        if sys.stdout is not None and hasattr(sys.stdout, 'buffer'):
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        if sys.stderr is not None and hasattr(sys.stderr, 'buffer'):
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Alignment = None  # type: ignore
    Font = None  # type: ignore
    get_column_letter = None  # type: ignore


URL = "https://mektep.edu.kz/?school=logout&language=rus"
LOGIN_PANEL_SELECTOR = "#collapseThree"

LANG_MAP = {
    "ru": {"label": "Русский", "query": "language=rus"},
    "kk": {"label": "Қазақша", "query": "language=kaz"},
    "en": {"label": "English", "query": "language=eng"},
}

# Import PERIOD_MAP from webapp.constants if available, otherwise define locally
try:
    from webapp.constants import PERIOD_MAP
except ImportError:
    PERIOD_MAP = {
        "1": "1 четверть",
        "2": "2 четверть (1 полугодие)",
        "3": "3 четверть",
        "4": "4 четверть (2 полугодие)",
        "5": "Учебный год",
    }

def _safe_slug(s: str) -> str:
    """Делает строку безопасной для имён файлов: пробелы, запрещённые символы."""
    s = " ".join((s or "").split()).strip()
    s = re.sub(r"[<>:\"/\\\\|?*]+", "_", s)
    s = s.strip(" .")
    return s or "item"


def _update_progress(percent: int, message: str, total_reports: int | None = None, processed_reports: int = 0):
    """Пишет JSON прогресса в файл из переменной окружения PROGRESS_FILE (для десктопа/UI)."""
    progress_file = os.getenv("PROGRESS_FILE")
    if progress_file:
        try:
            progress_path = Path(progress_file)
            data = {
                "percent": percent,
                "message": message,
                "total_reports": total_reports,
                "processed_reports": processed_reports,
                "finished": False
            }
            progress_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass  # Silently fail if progress file can't be written


def _ensure_dir(p: Path) -> None:
    """Создаёт каталог и родителей, если их ещё нет."""
    p.mkdir(parents=True, exist_ok=True)


def _resolve_template_path(name: str) -> Path:
    """Ищет шаблон Excel/Word по MEKTEP_TEMPLATES_DIR, PyInstaller, cwd и корню проекта."""
    candidates: list[Path] = []
    env_dir = os.getenv("MEKTEP_TEMPLATES_DIR")
    if env_dir:
        candidates.append(Path(env_dir) / name)
    # PyInstaller one-file/one-dir temp bundle
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(Path(meipass) / "templates" / name)
    # Current working directory
    candidates.append(Path(name))
    # Project root (where this file lives)
    base_dir = Path(__file__).resolve().parent
    candidates.append(base_dir / name)
    # Desktop app folders
    desktop_dir = base_dir / "mektep-desktop"
    candidates.append(desktop_dir / "templates" / name)
    candidates.append(desktop_dir / name)
    # Optional templates folder in root
    candidates.append(base_dir / "templates" / name)

    for c in candidates:
        if c.exists():
            return c
    return candidates[0] if candidates else Path(name)

def _try_click(locator, timeout_ms: int = 1500) -> None:
    """Пытается один раз кликнуть по локатору Playwright; ошибки игнорируются."""
    try:
        locator.first.click(timeout=timeout_ms)
    except Exception:
        pass


def _dismiss_announcement_modal(page: Page, timeout_ms: int = 600) -> bool:
    """
    Закрывает модальное окно «Объявление», если оно перекрывает интерфейс.
    Возвращает True, если модалка была найдена (попытка закрытия выполнена).
    """
    try:
        title = page.locator("#jurnalCloseWarningModal h5.modal-title:has-text('Объявление')").first
        title.wait_for(state="visible", timeout=timeout_ms)
    except Exception:
        return False

    try:
        btn = page.locator(
            "#jurnalCloseWarningModal button.btn.btn-danger:has-text('Закрыть'), "
            "#jurnalCloseWarningModal button[data-dismiss='modal']:has-text('Закрыть')"
        ).first
        btn.click(timeout=1500)
    except Exception:
        pass

    try:
        page.locator("#jurnalCloseWarningModal").first.wait_for(state="hidden", timeout=2000)
    except Exception:
        pass

    return True


def _dismiss_blocking_ui(page: Page) -> None:
    """Снимает мешающие модалки: общие кнопки закрытия и окно объявления."""
    # Generic close buttons (Bootstrap/modal patterns)
    _try_click(page.locator("button:has-text('×')"))
    _try_click(page.locator("[aria-label='Close']"))
    # Site-specific announcement modal (may or may not appear)
    _dismiss_announcement_modal(page)


def _click_login_button(page) -> None:
    """Нажимает кнопку «Вход в систему» в шапке по нескольким возможным селекторам."""
    # Prefer stable attributes from the provided HTML:
    # <button ... data-toggle="collapse" href="#collapseThree" aria-controls="collapseThree">Вход в систему</button>
    candidates = [
        page.locator('button[aria-controls="collapseThree"]'),
        page.locator('button[href="#collapseThree"]'),
        page.locator('button[data-toggle="collapse"][href="#collapseThree"]'),
        page.locator('header button.btn.btn-primary:has-text("Вход в систему")'),
        page.locator('button.btn.btn-primary:has-text("Вход в систему")'),
        page.locator('button:has-text("Вход в систему")'),
    ]

    last_err: Exception | None = None
    for loc in candidates:
        try:
            loc.first.wait_for(state="visible", timeout=10000)
            loc.first.scroll_into_view_if_needed(timeout=5000)
            loc.first.click(timeout=10000)
            return
        except Exception as e:
            last_err = e

    raise RuntimeError("Could not find/click the 'Вход в систему' button") from last_err


def _get_current_language(page) -> str | None:
    """Читает подпись текущего языка из выпадающего списка в шапке (Қазақша/Русский/…)."""
    # На некоторых страницах модальные окна могут перекрывать переключатель языка.
    _dismiss_blocking_ui(page)
    btn = page.locator("div.topline .btn-group button.btn.btn-default.dropdown-toggle").first
    try:
        btn.wait_for(state="visible", timeout=7000)
        txt = btn.inner_text().strip()
        # Example: "Қазақша" or "Русский" (may include caret/icon/newlines)
        for v in LANG_MAP.values():
            if v["label"] in txt:
                return v["label"]
        return txt.splitlines()[-1].strip() if txt else None
    except Exception:
        return None


def _ensure_language(page, lang_code: str) -> None:
    """Переключает язык интерфейса на нужный (desktop dropdown или мобильная ссылка)."""
    desired = LANG_MAP[lang_code]["label"]
    current = _get_current_language(page)
    if current and desired in current:
        print(f"Language already set: {desired}")
        return

    print(f"Switching language to: {desired}")

    # Desktop dropdown
    dropdown_btn = page.locator("div.topline .btn-group button.btn.btn-default.dropdown-toggle").first
    try:
        _dismiss_blocking_ui(page)
        dropdown_btn.wait_for(state="visible", timeout=7000)
        dropdown_btn.click(timeout=7000)
        page.locator(f'div.topline .dropdown-menu a.dropdown-item[href*="{LANG_MAP[lang_code]["query"]}"]').first.click(
            timeout=7000
        )
        page.wait_for_load_state("domcontentloaded")
        return
    except Exception:
        pass

    # Mobile fallback (direct link)
    _dismiss_blocking_ui(page)
    page.locator(f'div.mobile_lang a[href*="{LANG_MAP[lang_code]["query"]}"]').first.click(timeout=7000)
    page.wait_for_load_state("domcontentloaded")


def _get_profile_name(page) -> str | None:
    """Возвращает ФИО учителя из блока nav .profile после входа."""
    # Example HTML:
    # <div class="profile"> ... <p>Баер<br>Эдуард</p>
    loc = page.locator("nav .profile p").first
    try:
        loc.wait_for(state="visible", timeout=7000)
        txt = loc.inner_text()
        # Normalize whitespace/newlines.
        name = " ".join(txt.split()).strip()
        return name or None
    except Exception:
        return None


def _get_org_name(page) -> str | None:
    """Возвращает название организации (школы) из шапки сайта."""
    # Example HTML:
    # <div class="orgname">...<strong>Специализированный IT лицей</strong>
    loc = page.locator(".topline .orgname strong").first
    try:
        loc.wait_for(state="visible", timeout=7000)
        txt = loc.inner_text()
        name = " ".join((txt or "").split()).strip()
        return name or None
    except Exception:
        return None


def _choose_period() -> tuple[str, str]:
    """
    Возвращает (код периода, подпись): из MEKTEP_PERIOD или интерактивный ввод 1–5.
    Правило: 2-я четверть соответствует 1 полугодию в подписи PERIOD_MAP.
    Код 5 — «Учебный год» (вкладка #chetvert_5).
    """
    chosen = os.getenv("MEKTEP_PERIOD", "").strip()
    if chosen in PERIOD_MAP:
        return chosen, PERIOD_MAP[chosen]

    print("Какой период будем извлекать?")
    print("  1 - 1 четверть")
    print("  2 - 2 четверть (1 полугодие, если нет 2 четверти)")
    print("  3 - 3 четверть")
    print("  4 - 4 четверть (2 полугодие)")
    print("  5 - Учебный год")
    chosen = input("Выбор (1/2/3/4/5) [2]: ").strip() or "2"
    if chosen not in PERIOD_MAP:
        raise ValueError("Unknown period. Use: 1,2,3,4,5")
    return chosen, PERIOD_MAP[chosen]


def _go_to_grades(page) -> None:
    """Переходит на страницу «Оценки» (/office/?action=semester) по ссылке или через goto."""
    _dismiss_blocking_ui(page)

    # Fast path: if already on grades page and table is visible, skip navigation.
    try:
        if "action=semester" in page.url and page.locator("table.table.table-hover").first.is_visible():
            return
    except Exception:
        pass

    # Fast path: click visible grades nav link.
    link = page.locator('a.nav-link[href="/office/?action=semester"]:visible').first
    try:
        link.wait_for(state="visible", timeout=7000)
        link.click(timeout=7000)
        page.wait_for_selector("table.table.table-hover", timeout=8000)
        page.wait_for_load_state("domcontentloaded")
        _dismiss_blocking_ui(page)
        return
    except Exception:
        # Fallback: direct navigation.
        page.goto("https://mektep.edu.kz/office/?action=semester", wait_until="domcontentloaded")

    page.wait_for_load_state("domcontentloaded")
    _dismiss_blocking_ui(page)

def _extract_grades_table(page) -> list[dict]:
    """
    Собирает строки таблицы успеваемости: класс, предмет, ссылка на критерии (semester2).
    """
    table = page.locator("table.table.table-hover").first
    table.wait_for(state="visible", timeout=15000)

    rows: list[dict] = []
    trs = table.locator("tbody tr")
    count = trs.count()
    for i in range(count):
        tr = trs.nth(i)
        tds = tr.locator("td")
        
        # Check for very small cells (2x2px) - these indicate problematic rows
        # If such a cell exists and there's no button, skip this row (case 1)
        has_small_cell = False
        td_count = tds.count()
        for j in range(td_count):
            try:
                td = tds.nth(j)
                box = td.bounding_box()
                if box:
                    width = box.get("width", 0)
                    height = box.get("height", 0)
                    # Check if cell is very small (2x2px or similar tiny size)
                    if width <= 3 and height <= 3:
                        has_small_cell = True
                        break
            except Exception:
                pass
        
        # Check if row has criteria button
        criteria_a = tr.locator('a[href*="action=semester2"]').first
        has_button = False
        href = ""
        try:
            if criteria_a.count() > 0:
                has_button = True
                href = criteria_a.get_attribute("href", timeout=1000) or ""
        except Exception:
            pass
        
        # Skip rows with small cells but no button (case 1)
        if has_small_cell and not has_button:
            continue
        
        class_name = " ".join((tds.nth(0).inner_text() or "").split())
        # Subject cell can contain extra muted text like "Обновленное содержание".
        # We only want the <strong>...</strong> text (e.g., "Алгебра").
        subject_strong = tds.nth(1).locator("strong").first
        try:
            subject = " ".join(((subject_strong.inner_text() or "").split()))
        except Exception:
            subject = ""
        if not subject:
            # Fallback: use full cell text but drop muted part if present.
            full = " ".join((tds.nth(1).inner_text() or "").split())
            muted = ""
            try:
                muted = " ".join((tds.nth(1).locator("div.text-muted").first.inner_text() or "").split())
            except Exception:
                muted = ""
            if muted and muted in full:
                full = full.replace(muted, "").strip()
            subject = full

        if not href:
            # Skip rows without criteria link (they can't be opened).
            continue
        rows.append(
            {
                "index": i + 1,
                "class": class_name,
                "subject": subject,
                "criteria_href": href,
                "criteria_url": urljoin(page.url, href) if href else "",
            }
        )

    return rows


def _save_grades_table(rows: list[dict], out_dir: Path) -> None:
    """Сохраняет таблицу успеваемости в grades_table.json и grades_table.csv."""
    (out_dir / "grades_table.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    csv_path = out_dir / "grades_table.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["index", "class", "subject", "criteria_url"])
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in ["index", "class", "subject", "criteria_url"]})


def _choose_row(rows: list[dict]) -> dict:
    """Выбирает строку таблицы по MEKTEP_PICK или запрашивает номер у пользователя в консоли."""
    pick = os.getenv("MEKTEP_PICK", "").strip()
    if pick.isdigit():
        idx = int(pick)
        if 1 <= idx <= len(rows):
            return rows[idx - 1]

    print("Доступные строки (класс / предмет):")
    for r in rows:
        print(f'  {r["index"]:>2}. {r["class"]} — {r["subject"]}')
    raw = input(f"Выберите строку (1..{len(rows)}): ").strip()
    if not raw.isdigit():
        raise ValueError("Pick must be a number.")
    idx = int(raw)
    if not (1 <= idx <= len(rows)):
        raise ValueError("Pick out of range.")
    return rows[idx - 1]


def _open_criteria(page, href: str, absolute_url: str | None = None, prefer_goto: bool = False) -> None:
    """Открывает страницу критериев по ссылке (клик или page.goto)."""
    if not href:
        raise ValueError("Empty criteria link.")
    # Batch-fast-path: go directly by URL, then fallback to click.
    _dismiss_blocking_ui(page)
    if prefer_goto:
        try:
            target = absolute_url or urljoin(page.url, href)
            page.goto(target, wait_until="domcontentloaded")
            page.wait_for_load_state("domcontentloaded")
            _dismiss_blocking_ui(page)
            return
        except Exception:
            pass

    # Default path: click if possible; fallback to goto.
    try:
        page.locator(f'a[href="{href}"]').first.click(timeout=7000)
    except Exception:
        target = absolute_url or urljoin(page.url, href)
        page.goto(target, wait_until="domcontentloaded")
    page.wait_for_load_state("domcontentloaded")
    _dismiss_blocking_ui(page)


def _check_criteria_warning(page) -> bool:
    """
    Проверяет предупреждение «необходимо установить данные оценивания» на странице критериев.
    Возвращает True, если страницу нужно пропустить.
    """
    try:
        # Look for alert warning with text about missing evaluation data
        # Try multiple selectors for robustness
        warning_selectors = [
            'div.alert.alert-warning:has-text("Для начала работы необходимо установить данные оценивания!")',
            'div.alert.alert-warning',
        ]
        
        for selector in warning_selectors:
            warning = page.locator(selector)
            if warning.count() > 0:
                # Check if it's visible and contains the warning text
                for i in range(warning.count()):
                    try:
                        warning_elem = warning.nth(i)
                        if warning_elem.is_visible():
                            text = warning_elem.inner_text()
                            if "Для начала работы необходимо установить данные оценивания!" in text:
                                return True
                    except Exception:
                        continue
    except Exception:
        pass
    return False


def _list_criteria_tabs(page) -> list[dict]:
    """Возвращает список вкладок критериев (текст и href якоря #chetvert_N)."""
    tabs = page.locator("ul#pills-tab a[data-toggle='pill']")
    out: list[dict] = []
    n = tabs.count()
    for i in range(n):
        a = tabs.nth(i)
        href = (a.get_attribute("href") or "").strip()
        text = " ".join((a.inner_text() or "").split())
        out.append({"text": text, "href": href})
    return [t for t in out if t["href"].startswith("#")]


def _click_criteria_tab(page, href: str) -> None:
    """Переключает вкладку критериев по href и ждёт видимости соответствующей панели."""
    # href like "#chetvert_2"
    loc = page.locator(f'ul#pills-tab a[data-toggle="pill"][href="{href}"]').first
    try:
        loc.click(timeout=7000)
    except Exception:
        # Fixed header (topline/top_header) may intercept pointer events; use JS click as fallback
        loc.evaluate("el => el.click()")
    # Wait until the corresponding pane is shown (Bootstrap adds 'show' + 'active').
    pane_id = href.lstrip("#")
    pane = page.locator(f"div.tab-content div.tab-pane#{pane_id}").first
    pane.wait_for(state="visible", timeout=15000)
    try:
        page.wait_for_selector(f"div.tab-content div.tab-pane#{pane_id}.show", timeout=15000)
    except Exception:
        # Some pages render panes visible without 'show' toggling consistently.
        pass


def _count_tab_rows(page, href: str) -> int:
    """Считает число строк данных в первой таблице внутри вкладки с данным href."""
    pane_id = href.lstrip("#")
    pane = page.locator(f"div.tab-content div.tab-pane#{pane_id}").first
    # Count data rows in the first table inside the pane.
    table = pane.locator("table").first
    try:
        table.wait_for(state="visible", timeout=3000)
    except Exception:
        return 0
    return table.locator("tbody tr").count()


def _pick_tab_href_for_period(period_code: str, tabs: list[dict]) -> str | None:
    """
    Сопоставляет выбранную четверть с href вкладки на сайте (с запасными правилами для полугодий).
    """
    hrefs = {t["href"] for t in tabs}
    texts = {t["href"]: t["text"] for t in tabs}

    if period_code in {"1", "2", "3"}:
        direct = f"#chetvert_{period_code}"
        if direct in hrefs:
            return direct

    # Period 4 could be either 4th quarter or "2 полугодие" depending on school settings.
    if period_code == "4":
        if "#chetvert_4" in hrefs:
            return "#chetvert_4"
        # Fallback by label (Russian: полугодие, Kazakh: жартыжылдық):
        for href, txt in texts.items():
            if "2 полугод" in txt.lower() or ("2" in txt and "жартыжылдық" in txt):
                return href

    if period_code == "2":
        # Fallback to "1 полугодие" if 2nd quarter is absent.
        for href, txt in texts.items():
            if "1 полугод" in txt.lower() or ("1" in txt and "жартыжылдық" in txt):
                return href

    # Учебный год: вкладка #chetvert_5 или подпись, содержащая "учебн"/"оқу жыл".
    if period_code == "5":
        if "#chetvert_5" in hrefs:
            return "#chetvert_5"
        for href, txt in texts.items():
            low = txt.lower()
            if "учебн" in low or "оқу жыл" in low or "оку жыл" in low:
                return href

    # Fallback by matching quarter number in text.
    desired_label = PERIOD_MAP.get(period_code, "")
    desired_num = desired_label.split()[0] if desired_label else ""
    if desired_num:
        for href, txt in texts.items():
            if desired_num in txt:
                return href

    return tabs[0]["href"] if tabs else None


def _analyze_and_select_criteria_tabs(page, out_dir: Path, period_code: str) -> str | None:
    """Обходит вкладки, пишет criteria_tabs.json и активирует вкладку, соответствующую периоду."""
    tabs = _list_criteria_tabs(page)
    if not tabs:
        (out_dir / "criteria_tabs.json").write_text(json.dumps([], ensure_ascii=False, indent=2), encoding="utf-8")
        return None

    # Check each tab for content (row counts).
    report: list[dict] = []
    for t in tabs:
        href = t["href"]
        try:
            _click_criteria_tab(page, href)
        except Exception:
            report.append({**t, "rows": 0, "click_error": True})
            continue
        report.append({**t, "rows": _count_tab_rows(page, href)})

    (out_dir / "criteria_tabs.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    # Select the desired period tab (with fallback rules).
    desired_href = _pick_tab_href_for_period(period_code, tabs)
    if desired_href:
        print(f"Selecting period tab: {desired_href}")
        _click_criteria_tab(page, desired_href)
        (out_dir / "criteria_selected_tab.txt").write_text(desired_href, encoding="utf-8")
        return desired_href

    return None


def _get_active_criteria_tab_href(page) -> str | None:
    """Возвращает href активной вкладки критериев или None."""
    a = page.locator('ul#pills-tab a[data-toggle="pill"].active').first
    try:
        if a.count() == 0:
            return None
        href = (a.get_attribute("href") or "").strip()
        return href if href.startswith("#") else None
    except Exception:
        return None


def _text_or_none(locator) -> str | None:
    """Безопасно читает inner_text первого элемента локатора или возвращает None."""
    try:
        if locator.count() == 0:
            return None
        txt = locator.first.inner_text()
        txt = " ".join((txt or "").split()).strip()
        return txt or None
    except Exception:
        return None


def _has_quarter_grade_header(page, tab_href: str) -> bool:
    """
    Проверяет, есть ли в панели заголовок «Расчет оценки за N четверть» / «Бағаны есептеу: N тоқсан».
    Если заголовок есть — оценка четвертная, добавлять в сводную ведомость.
    Если нет — структура без четвертной оценки (полугодовая/по разделам), не добавлять.
    """
    pane_id = tab_href.lstrip("#")
    pane = page.locator(f"div#pills-tabContent div.tab-pane#{pane_id}").first
    try:
        pane.wait_for(state="visible", timeout=5000)
    except Exception:
        return False
    tds = pane.locator("td")
    n = tds.count()
    for i in range(n):
        txt = (tds.nth(i).inner_text() or "").strip()
        if "Расчет оценки за" in txt or "Бағаны есептеу:" in txt:
            return True
    return False


def _extract_students_from_criteria_tab(page, tab_href: str) -> list[dict]:
    """Извлекает список учеников с оценками и баллами по секциям для выбранной вкладки критериев."""
    pane_id = tab_href.lstrip("#")
    pane = page.locator(f"div#pills-tabContent div.tab-pane#{pane_id}").first
    pane.wait_for(state="visible", timeout=15000)

    # Determine quarter_num from tab id (#chetvert_N)
    quarter_num = 0
    try:
        quarter_num = int(pane_id.split("_")[1])
    except Exception:
        quarter_num = 0

    # Try to find table by grade cell first (preferred method)
    first_grade = pane.locator('p[id^="ocenka_"]').first
    table = None
    
    try:
        first_grade.wait_for(state="visible", timeout=7000)
        grade_id = (first_grade.get_attribute("id") or "").strip()
        # Example grade id: ocenka_0_chetvert_2  -> row_idx=0, quarter_num=2
        try:
            # ocenka_{row}_chetvert_{q}
            parts = grade_id.split("_")
            if len(parts) >= 4 and parts[0] == "ocenka" and parts[2] == "chetvert":
                quarter_num = int(parts[3])  # Override with actual quarter from grade id
        except Exception:
            pass
        
        table = first_grade.locator("xpath=ancestor::table[1]")
        table.wait_for(state="visible", timeout=15000)
    except Exception:
        # Fallback: try to find table directly by structure
        # Table is usually inside form within the tab pane
        try:
            # Try to find table with student data (has tbody with rows)
            tables = pane.locator("table.table tbody tr")
            if tables.count() > 0:
                # Get the first table that has rows
                table = tables.first.locator("xpath=ancestor::table[1]")
                table.wait_for(state="visible", timeout=5000)
            else:
                # Try alternative: find table inside form
                form_table = pane.locator("form table, table.table")
                if form_table.count() > 0:
                    table = form_table.first
                    table.wait_for(state="visible", timeout=5000)
        except Exception:
            print(f"Warning: Could not find students table in tab {tab_href}")
            return []
    
    if table is None:
        return []

    trs = table.locator("tbody tr, tr")
    n = trs.count()
    if n == 0:
        print(f"Warning: No table rows found in tab {tab_href}")
        return []

    # Batch-read row contents in one browser round-trip to reduce Playwright overhead.
    row_payload = trs.evaluate_all(
        """
        (rows) => rows.map((tr) => {
          const norm = (s) => (s || "").replace(/\\s+/g, " ").trim();
          const cells = Array.from(tr.querySelectorAll("td"), (td) => norm(td.innerText));
          const pNodes = Array.from(tr.querySelectorAll("p[id]"), (p) => ({
            id: p.id || "",
            text: norm(p.innerText),
          }));
          const points = {};
          for (const inp of tr.querySelectorAll('input[id^="chetvert_"][id*="_razdel_"]')) {
            if (inp.id) points[inp.id] = inp.value || "";
          }
          return { cells, pNodes, points };
        })
        """
    )

    out: list[dict] = []
    for i, row in enumerate(row_payload):
        all_cells = row.get("cells") or []
        if len(all_cells) < 2:
            continue

        num_txt = all_cells[0]
        fio_txt = all_cells[1]
        if not str(num_txt).isdigit():
            continue

        p_map: dict[str, str] = {}
        for p in row.get("pNodes") or []:
            pid = (p.get("id") or "").strip()
            if pid:
                p_map[pid] = (p.get("text") or "").strip()

        row_index = i
        grade_val = ""
        for pid, txt in p_map.items():
            marker = f"_chetvert_{quarter_num}"
            if pid.startswith("ocenka_") and marker in pid:
                parts = pid.split("_")
                if len(parts) >= 4 and parts[2] == "chetvert":
                    try:
                        row_index = int(parts[1])
                    except Exception:
                        row_index = i
                grade_val = txt
                break

        average_val = p_map.get(f"average_{quarter_num}_chetvert_{row_index}", "")
        formative_pct_val = p_map.get(f"average_itog_{quarter_num}_chetvert_{row_index}", "")
        sor_pct_val = p_map.get(f"sor_{row_index}_chetvert_{quarter_num}", "")
        soch_pct_val = p_map.get(f"soch_{row_index}_chetvert_{quarter_num}", "")
        total_pct_val = p_map.get(f"summa_{row_index}_chetvert_{quarter_num}", "")

        if not average_val or not any([formative_pct_val, sor_pct_val, soch_pct_val, total_pct_val]):
            for cell_text in all_cells[2:]:
                if not cell_text:
                    continue
                try:
                    float_val = float(str(cell_text).replace(",", "."))
                    if not average_val and 0 <= float_val <= 20:
                        average_val = cell_text
                        continue
                except (ValueError, AttributeError):
                    pass
                if "%" in str(cell_text):
                    if not formative_pct_val:
                        formative_pct_val = cell_text
                    elif not sor_pct_val:
                        sor_pct_val = cell_text
                    elif not soch_pct_val:
                        soch_pct_val = cell_text
                    elif not total_pct_val:
                        total_pct_val = cell_text

        if not grade_val and all_cells:
            for cell_text in all_cells[max(0, len(all_cells) - 3) :]:
                if str(cell_text) in {"1", "2", "3", "4", "5"}:
                    grade_val = str(cell_text)
                    break

        out.append(
            {
                "num": int(num_txt),
                "fio": fio_txt,
                "quarter_num": quarter_num,
                "average": average_val or "",
                "formative_pct": formative_pct_val or "",
                "sor_pct": sor_pct_val or "",
                "soch_pct": soch_pct_val or "",
                "total_pct": total_pct_val or "",
                "grade": grade_val or "",
                "raw_cells": [str(c) for c in all_cells],
                "points": row.get("points") or {},
            }
        )

    return out


def _extract_quarter_max_points(page: Page, tab_href: str) -> dict[int, int]:
    """
    Читает максимальные баллы по секциям (СОч/СОр) из полей chetvert_*_razdel_*_max.
    """
    pane_id = tab_href.lstrip("#")
    pane = page.locator(f"div#pills-tabContent div.tab-pane#{pane_id}").first

    # Prefer quarter number from actual input IDs (more reliable than pane id naming).
    inputs_any = pane.locator('input[id^="chetvert_"][id*="_razdel_"][id$="_max"]')
    if inputs_any.count() == 0:
        return {}

    # Infer quarter_num from first input id: chetvert_{q}_razdel_{k}_max
    first_id = (inputs_any.first.get_attribute("id") or "").strip()
    quarter_num = None
    try:
        parts = first_id.split("_")
        if len(parts) >= 5 and parts[0] == "chetvert":
            quarter_num = int(parts[1])
    except Exception:
        quarter_num = None
    if quarter_num is None:
        return {}

    inputs = pane.locator(f'input[id^="chetvert_{quarter_num}_razdel_"][id$="_max"]')
    n = inputs.count()
    out: dict[int, int] = {}
    for i in range(n):
        inp = inputs.nth(i)
        inp_id = (inp.get_attribute("id") or "").strip()
        val = (inp.input_value() or "").strip()
        # chetvert_2_razdel_1_max -> section=1
        parts = inp_id.split("_")
        if len(parts) >= 5 and parts[0] == "chetvert" and parts[2] == "razdel":
            try:
                section = int(parts[3])
                out[section] = int(float(val)) if val else out.get(section, 0)
            except Exception:
                continue
    return out


def _parse_points_by_section(points: dict[str, str], quarter_num: int) -> dict[int, str]:
    """
    Группирует значения input по номеру секции razdel из id chetvert_q_razdel_k_...
    """
    out: dict[int, str] = {}
    prefix = f"chetvert_{quarter_num}_razdel_"
    for k, v in points.items():
        if not k.startswith(prefix):
            continue
        parts = k.split("_")
        # ["chetvert", q, "razdel", k, rowIndex]
        if len(parts) >= 5 and parts[2] == "razdel":
            try:
                section = int(parts[3])
            except Exception:
                continue
            out[section] = v
    return out


def _export_students_xlsx(
    out_dir: Path,
    students: list[dict],
    ctx: dict,
    max_points: dict[int, int],
) -> None:
    """Формирует criteria_students.xlsx (лист students + context) через openpyxl."""
    if Workbook is None:
        print("openpyxl not installed; skipping XLSX export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    meta = wb.create_sheet("context")

    # Determine quarter number from first row.
    quarter_num = int(students[0].get("quarter_num", 0) or 0) if students else 0

    # Determine which sections exist in points across all students.
    sections: set[int] = set()
    for s in students:
        pts = s.get("points") or {}
        if isinstance(pts, dict):
            for sid in _parse_points_by_section(pts, quarter_num).keys():
                sections.add(sid)
    section_list = sorted(sections)

    def section_label(sec: int) -> str:
        """Возвращает заголовок колонки для секции: СОч (0) или СОр N."""
        if sec == 0:
            return "СОч"
        return f"СОр {sec}"

    headers = [
        "№",
        "ФИО",
        "Формативная (среднее)",
        *[section_label(sec) for sec in section_list],
        "% ФО",
        "% СОр",
        "% СОч",
        "Итог %",
        "Оценка",
    ]

    ws.append(headers)
    # Row 2: max points for section columns (instead of embedding it into header text)
    max_row = [""] * len(headers)
    max_row[1] = "Макс."
    sec_start_idx = 3  # 0-based; after №, ФИО, average
    for i, sec in enumerate(section_list):
        mp = max_points.get(sec)
        if mp is not None:
            max_row[sec_start_idx + i] = mp
    ws.append(max_row)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Style header
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for s in students:
        pts = s.get("points") or {}
        sec_points = _parse_points_by_section(pts, quarter_num) if isinstance(pts, dict) else {}
        row = [
            s.get("num", ""),
            s.get("fio", ""),
            s.get("average", ""),
            *[sec_points.get(sec, "") for sec in section_list],
            s.get("formative_pct", ""),
            s.get("sor_pct", ""),
            s.get("soch_pct", ""),
            s.get("total_pct", ""),
            s.get("grade", ""),
        ]
        ws.append(row)

    # Basic column widths
    widths = {
        1: 5,
        2: 28,
        3: 18,
    }
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 14)

    # Context sheet
    meta.append(["generated_at", datetime.now().isoformat(timespec="seconds")])
    for k in [
        "org_name",
        "profile_name",
        "class",
        "subject",
        "period_code",
        "period_label",
        "selected_tab",
        "criteria_url",
    ]:
        meta.append([k, ctx.get(k, "")])

    xlsx_path = out_dir / "criteria_students.xlsx"
    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")


def _auto_install_chromium() -> bool:
    """Автоматическая установка Chromium через Playwright CLI.
    
    Возвращает True если установка прошла успешно.
    """
    import subprocess as _sp
    log_info("Браузер Chromium не найден. Запускаю автоустановку...")
    
    try:
        # В frozen-режиме sys.executable — это .exe, поэтому используем
        # playwright CLI напрямую через его внутренний driver
        from playwright._impl._driver import compute_driver_executable
        driver = str(compute_driver_executable())
        log_info(f"Playwright driver: {driver}")
        
        result = _sp.run(
            [driver, "install", "chromium"],
            capture_output=True, text=True, timeout=600,
            env={**os.environ, "PLAYWRIGHT_BROWSERS_PATH": "0"},  # 0 = стандартная папка
        )
        if result.returncode == 0:
            log_success("Chromium успешно установлен")
            return True
        else:
            log_error(f"Ошибка установки Chromium (код {result.returncode}): {result.stderr[:500]}")
            return False
    except Exception as e:
        log_error(f"Не удалось установить Chromium автоматически: {e}")
        return False


def _launch_browser(p, headless: bool, slow_mo_ms: int):
    """Запуск браузера с fallback-цепочкой:
    
    1. Playwright-управляемый Chromium (ms-playwright)
    2. Системный Microsoft Edge
    3. Системный Google Chrome
    4. Автоустановка Chromium → повторная попытка
    """
    # ── Попытка 1: Playwright Chromium ──
    try:
        log_info("Попытка запуска: Playwright Chromium...")
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        log_success("Браузер запущен: Playwright Chromium")
        return browser
    except Exception as e1:
        log_warning(f"Playwright Chromium недоступен: {e1}")

    # ── Попытка 2: Microsoft Edge ──
    try:
        log_info("Попытка запуска: Microsoft Edge...")
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms, channel="msedge")
        log_success("Браузер запущен: Microsoft Edge")
        return browser
    except Exception as e2:
        log_warning(f"Microsoft Edge недоступен: {e2}")

    # ── Попытка 3: Google Chrome ──
    try:
        log_info("Попытка запуска: Google Chrome...")
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms, channel="chrome")
        log_success("Браузер запущен: Google Chrome")
        return browser
    except Exception as e3:
        log_warning(f"Google Chrome недоступен: {e3}")

    # ── Попытка 4: автоустановка Chromium и повтор ──
    log_info("Ни один браузер не найден. Запускаю автоустановку Chromium...")
    if _auto_install_chromium():
        try:
            browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
            log_success("Браузер запущен: Playwright Chromium (после автоустановки)")
            return browser
        except Exception as e4:
            log_error(f"Не удалось запустить Chromium после установки: {e4}")

    raise RuntimeError(
        "Не удалось запустить браузер. Убедитесь, что установлен "
        "Microsoft Edge, Google Chrome или выполните команду: playwright install chromium"
    )


def run(headless: bool, out_dir: Path, slow_mo_ms: int) -> int:
    """Основной сценарий скрапинга Mektep: вход, оценки, критерии, отчёты; код выхода для десктопа."""
    _ensure_dir(out_dir)
    
    # Инициализация логгера
    logger = init_logger(out_dir, os.getenv("PROGRESS_FILE"))
    log_stage(ScraperLogger.STAGE_INIT, "Инициализация скрапера", 1)
    log_info(f"Директория вывода: {out_dir}")
    log_info(f"Режим: {'headless' if headless else 'с отображением браузера'}")

    login = os.getenv("MEKTEP_LOGIN") or ""
    password = os.getenv("MEKTEP_PASSWORD") or ""
    
    if login:
        log_info(f"Логин: {login[:3]}***")
    else:
        log_warning("Логин не указан в переменных окружения")

    # Проверка ИИН до запуска браузера (MEKTEP_EXPECTED_IIN из сервера/админки)
    expected_iin = (os.getenv("MEKTEP_EXPECTED_IIN") or "").strip()
    if expected_iin:
        try:
            from iin_utils import normalize_kz_iin as _norm_iin
        except ImportError:
            def _norm_iin(s):
                d = re.sub(r"\D", "", s or "")
                return d if len(d) == 12 else None

        ld = _norm_iin(login)
        if not ld or ld != expected_iin:
            log_error(
                "Логин mektep.edu.kz не совпадает с ИИН, указанным администратором в системе."
            )
            if logger:
                logger.finish(success=False)
            return 6

    with sync_playwright() as p:
        log_stage(ScraperLogger.STAGE_BROWSER, "Запуск браузера Chromium", 2)
        with timing_block("запуск браузера (_launch_browser)"):
            browser = _launch_browser(p, headless, slow_mo_ms)
        context = browser.new_context(locale="ru-RU")
        page = context.new_page()
        log_success("Браузер запущен успешно")

        log_stage(ScraperLogger.STAGE_PAGE_LOAD, f"Загрузка страницы: {URL}", 3)
        try:
            with timing_block("page.goto главная mektep (wait_until=load, timeout=60s)"):
                page.goto(URL, wait_until="load", timeout=60000)
            log_success("Страница загружена")
        except Exception as e:
            log_error("Ошибка загрузки страницы", e)
            log_info("Возможные причины:")
            log_info("  - Проблемы с интернет-соединением")
            log_info("  - Сайт mektep.edu.kz недоступен или перегружен")
            log_info("  - Превышено время ожидания ответа сервера")
            if logger:
                logger.finish(success=False)
            raise
        page.wait_for_load_state("domcontentloaded")

        # Best-effort dismiss of any popups/modals that can block clicks.
        _dismiss_blocking_ui(page)

        log_stage(ScraperLogger.STAGE_LOGIN_FORM, "Открытие формы входа", 4)
        # Debug artifacts before clicking (useful if selector fails due to unexpected content).
        if _debug_artifacts_enabled():
            with timing_block("артефакты before_click (url, HTML, full_page screenshot)"):
                (out_dir / "before_click.url.txt").write_text(page.url, encoding="utf-8")
                (out_dir / "before_click.html").write_text(page.content(), encoding="utf-8")
                page.screenshot(path=str(out_dir / "before_click.png"), full_page=True)
            log_info("Сохранён скриншот: before_click.png")

        log_info("Нажатие кнопки 'Вход в систему'...")
        with timing_block("_click_login_button"):
            _click_login_button(page)
        log_success("Кнопка нажата")

        # Wait for the collapsed section to expand and reveal the login form.
        log_info("Ожидание раскрытия формы входа...")
        try:
            with timing_block("wait_for_selector форма входа (#collapseThree.show, до 15s)"):
                page.wait_for_selector(f"{LOGIN_PANEL_SELECTOR}.show", timeout=15000)
            log_success("Форма входа открыта")
        except Exception as e:
            log_error("Не удалось открыть форму входа", e)
            if logger:
                logger.finish(success=False)
            return 2
        panel = page.locator(LOGIN_PANEL_SELECTOR)

        log_info("Поиск полей ввода логина и пароля...")
        login_input = panel.locator(
            'input[name="usr_login"]:visible, input[name*="usr_login" i]:visible, input[name*="login" i]:visible, input[name*="iin" i]:visible, input[type="text"]:visible'
        ).first
        password_input = panel.locator(
            'input[name="usr_password"]:visible, input#password:visible, input[name*="pass" i]:visible, input[type="password"]:visible'
        ).first

        try:
            login_input.wait_for(state="visible", timeout=15000)
            password_input.wait_for(state="visible", timeout=15000)
            log_success("Поля ввода найдены")
        except Exception as e:
            log_error("Не найдены поля логина/пароля", e)
            if logger:
                logger.finish(success=False)
            return 2

        # Check credentials before attempting login
        all_mode = os.getenv("MEKTEP_ALL", "") == "1"
        headless_mode = headless
        
        # In headless/batch mode, don't ask for credentials interactively
        if not login or not password:
            if headless_mode or all_mode:
                log_error("Логин или пароль не указаны")
                if logger:
                    logger.finish(success=False)
                return 2
            # Only ask interactively if not in headless mode
            if not login:
                login = input("Логин или ИИН: ").strip()
            if not password:
                password = getpass.getpass("Пароль: ")
            if not login or not password:
                log_error("Логин или пароль пустые")
                if logger:
                    logger.finish(success=False)
                return 2
        
        # Single login attempt - if it fails, exit immediately
        log_stage(ScraperLogger.STAGE_AUTH, "Авторизация на сайте", 5)
        log_info(f"Ввод логина: {login[:3]}***")
        try:
            with timing_block("ввод логина/пароля (type delay=20) и submit формы"):
                login_input.click()
                login_input.press("Control+A")
                login_input.type(login, delay=20)
                log_info("Логин введён")

                password_input.click()
                password_input.press("Control+A")
                password_input.type(password, delay=20)
                log_info("Пароль введён")

                log_info("Отправка формы авторизации...")
                panel.locator("form button[type='submit'], form input[type='submit']").first.click(timeout=10000)
                page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception as e:
            log_error("Ошибка при вводе данных или отправке формы", e)
            if logger:
                logger.finish(success=False)
            return 4

        # Wait for login to process
        log_info("Ожидание обработки авторизации (3 сек)...")
        _timed_sleep(3, "после отправки формы логина (фиксированная пауза)")
        
        # Scenario 2 & 3 Combined: Check for role/school selection dialog
        # If there are multiple "Войти как учитель" buttons, it means multiple schools
        log_info("Проверка наличия окна выбора роли/школы...")
        _t_school = time.perf_counter()
        try:
            # Prefer condition-based wait; keep short fallback if dialog is not rendered yet.
            try:
                page.wait_for_selector('button[name="account_choice"][value="true"]', timeout=1200)
            except Exception:
                _timed_sleep(0.5, "fallback: окно выбора роли/школы еще не отрисовано")

            # Look for all role/school selection buttons by attributes (works for any language)
            # This selector finds buttons with name="account_choice" and value="true"
            # These buttons can have text "Войти как учитель" (RU) or "Мұғалім ретінде кіру" (KZ)
            # We filter by text containing "учитель" or "Мұғалім" to exclude other buttons like "Exit"
            all_buttons = page.locator('button[name="account_choice"][value="true"]')
            
            # Filter buttons that contain teacher-related text
            school_buttons_list = []
            for i in range(all_buttons.count()):
                btn = all_buttons.nth(i)
                btn_text = btn.inner_text().strip().lower()
                if 'учитель' in btn_text or 'мұғалім' in btn_text:
                    school_buttons_list.append(btn)
            
            button_count = len(school_buttons_list)
            log_info(f"Найдено кнопок выбора роли/школы (учитель): {button_count}")
            
            # Debug: show button texts if found
            if button_count > 0:
                try:
                    first_button_text = school_buttons_list[0].inner_text().strip()
                    log_info(f"Текст первой кнопки: '{first_button_text}'")
                except:
                    pass
            elif button_count == 0:
                # Дополнительная отладка: если не нашли, сохраняем файлы
                log_warning("Кнопки выбора не найдены, сохраняем отладочные файлы...")
                try:
                    page.screenshot(path=str(out_dir / "school_selection_not_found.png"))
                    (out_dir / "school_selection_not_found.html").write_text(page.content(), encoding="utf-8")
                    log_info("Сохранены файлы отладки: school_selection_not_found.png/html")
                except:
                    pass
            
            if button_count == 1:
                # Only one "Войти как учитель" button - simple role selection
                log_info("Обнаружено окно выбора роли (учитель/родитель)")
                school_buttons_list[0].click()
                log_success("Выбрана роль: Учитель")
                _timed_sleep(2, "после выбора роли «учитель»")
            elif button_count > 1:
                # Multiple "Войти как учитель" buttons - teacher works at multiple schools
                log_info(f"Обнаружено окно выбора школы. Доступно школ: {button_count}")
                
                # Get school names
                school_list = []
                for i in range(button_count):
                    try:
                        btn = school_buttons_list[i]
                        # Try multiple ways to find school name
                        school_name = None
                        
                        # Method 1: Look for <small> in parent form
                        try:
                            parent_form = btn.locator('xpath=ancestor::form')
                            school_name_elem = parent_form.locator('p small, small').first
                            school_name = school_name_elem.inner_text().strip()
                        except:
                            pass
                        
                        # Method 2: Look for <small> as next sibling
                        if not school_name:
                            try:
                                school_name_elem = btn.locator('xpath=following-sibling::*[1]//small, xpath=following-sibling::small')
                                if school_name_elem.count() > 0:
                                    school_name = school_name_elem.first.inner_text().strip()
                            except:
                                pass
                        
                        # Method 3: Default naming
                        if not school_name:
                            school_name = f"Школа {i+1}"
                    except Exception as e:
                        log_warning(f"Ошибка получения названия школы {i}: {e}")
                        school_name = f"Школа {i+1}"
                    
                    school_list.append(school_name)
                    log_info(f"  {i}: {school_name}")
                
                # Check if we have environment variable first (for backwards compatibility)
                chosen_school_idx_str = os.getenv("MEKTEP_SCHOOL_INDEX", "").strip()
                chosen_school_idx = None
                
                if chosen_school_idx_str and chosen_school_idx_str.isdigit():
                    # Use environment variable
                    chosen_school_idx = int(chosen_school_idx_str)
                    if 0 <= chosen_school_idx < button_count:
                        log_success(f"Используется предустановленный выбор школы #{chosen_school_idx}")
                    else:
                        log_warning(f"Неверный индекс школы {chosen_school_idx}, требуется выбор пользователя")
                        chosen_school_idx = None
                
                if chosen_school_idx is None:
                    # No environment variable - request user selection via progress file
                    log_info("Ожидание выбора школы пользователем...")
                    _update_progress(5, f"schools_selection_needed|{json.dumps(school_list, ensure_ascii=False)}")
                    _timed_sleep(2, "UI: сообщение о выборе школы")
                    _update_progress(5, "Ожидание выбора школы...")  # Очищаем специальное сообщение
                    
                    # Wait for user selection (with timeout)
                    school_choice_file = out_dir / "school_choice.txt"
                    if school_choice_file.exists():
                        school_choice_file.unlink()  # Remove old choice if exists
                    
                    timeout = 60  # 60 seconds timeout
                    start_time = time.time()
                    
                    while time.time() - start_time < timeout:
                        if school_choice_file.exists():
                            try:
                                file_content = school_choice_file.read_text(encoding="utf-8").strip()
                                chosen_school_idx = int(file_content)
                                log_info(f"Прочитан файл выбора: индекс {chosen_school_idx}")
                                if 0 <= chosen_school_idx < button_count:
                                    log_success(f"Получен выбор пользователя: школа #{chosen_school_idx}")
                                    school_choice_file.unlink()  # Удаляем файл после использования
                                    break
                                else:
                                    log_warning(f"Неверный индекс в файле выбора: {chosen_school_idx}")
                                    chosen_school_idx = None
                                    school_choice_file.unlink()  # Удаляем некорректный файл
                            except Exception as e:
                                log_warning(f"Ошибка чтения файла выбора: {e}")
                                chosen_school_idx = None
                            
                            if chosen_school_idx is not None:
                                break
                        
                        _timed_sleep(1.0, "poll school_choice.txt (интервал итерации)")
                    
                    # If timeout or no valid selection, use first school
                    if chosen_school_idx is None:
                        chosen_school_idx = 0
                        log_warning(f"Выбор не получен за {timeout} секунд, используется первая школа автоматически")
                
                # Click selected school button
                school_buttons_list[chosen_school_idx].click()
                log_success(f"Выбрана школа: {school_list[chosen_school_idx]}")
                try:
                    page.locator("nav .profile p, .topline .profile, .user-profile").first.wait_for(
                        state="visible", timeout=2000
                    )
                except Exception:
                    _timed_sleep(0.5, "fallback: после выбора школы (обработка UI)")
            else:
                # button_count == 0, no role/school selection needed
                log_info("Окно выбора роли/школы не обнаружено (стандартный вход)")

        except Exception as e:
            # No role/school selection dialog - that's fine, continue normally
            log_info(f"Окно выбора роли/школы не обнаружено: {e}")
        finally:
            log_timing(
                "выбор роли/школа: поиск кнопок, клики, ожидания",
                time.perf_counter() - _t_school,
            )

        # Check if login was successful
        log_info("Проверка результата авторизации...")
        try:
            with timing_block("wait_for профиль после входа (до 10s)"):
                page.locator("nav .profile p, .topline .profile, .user-profile").first.wait_for(
                    state="visible", timeout=10000
                )
            log_success("Авторизация успешна!")
        except Exception:
            # Login failed - check if we're still on login page
            page_content = page.content().lower()
            if "неверн" in page_content or "ошибк" in page_content or LOGIN_PANEL_SELECTOR in page.content():
                log_error("Неверный логин или пароль")
                if logger:
                    logger.finish(success=False)
                return 4
            else:
                # Unexpected state, but might be success - try one more check
                log_warning("Неопределённое состояние, повторная проверка...")
                _timed_sleep(2, "перед повторной проверкой профиля")
                try:
                    page.locator("nav .profile p, .topline .profile").first.wait_for(state="visible", timeout=5000)
                    log_success("Авторизация успешна (повторная проверка)")
                except Exception:
                    log_error("Неверный логин или пароль (повторная проверка)")
                    if logger:
                        logger.finish(success=False)
                    return 4

        # ============================================================
        # Определяем целевой язык отчётов
        # ============================================================
        chosen = os.getenv("MEKTEP_LANG", "").strip().lower()
        if chosen not in LANG_MAP:
            if headless or all_mode:
                chosen = "ru"
            else:
                chosen = input("Язык данных (ru/kk/en) [ru]: ").strip().lower() or "ru"
        if chosen not in LANG_MAP:
            log_error(f"Неизвестный язык: {chosen}")
            if logger:
                logger.finish(success=False)
            return 2

        # ============================================================
        # Читаем org_name СТРОГО НА РУССКОЙ СТРАНИЦЕ.
        # После выбора школы язык может быть любым (казахский/русский).
        # Принудительно переключаем на русский, читаем название,
        # затем переключаем на целевой язык для скрапинга.
        # ============================================================
        log_stage(ScraperLogger.STAGE_LANGUAGE, "Настройка языка интерфейса", 6)
        current_lang = _get_current_language(page)
        log_info(f"Текущий язык после авторизации: {current_lang or 'не определён'}")

        # Принудительно переключаем на русский для чтения org_name
        with timing_block("_ensure_language(ru) и чтение org_name_ru"):
            log_info("Переключение на русский для чтения названия организации...")
            _ensure_language(page, "ru")
            org_name_ru = _get_org_name(page)
        if org_name_ru:
            (out_dir / "org_name_ru.txt").write_text(org_name_ru, encoding="utf-8")
            log_info(f"Организация (рус): {org_name_ru}")
        else:
            log_warning("Название организации (рус) не найдено")

        # ===== Проверка организации (защита от передачи аккаунта) =====
        # MEKTEP_EXPECTED_SCHOOL передаётся из scraper_runner.py / десктоп-приложения
        # и содержит название школы, к которой привязан аккаунт в БД.
        expected_school = os.getenv("MEKTEP_EXPECTED_SCHOOL", "").strip()
        if expected_school and org_name_ru:
            a = " ".join(org_name_ru.lower().split())
            b = " ".join(expected_school.lower().split())
            if a != b and a not in b and b not in a:
                log_error(
                    f"Организация «{org_name_ru}» не совпадает с вашей школой «{expected_school}». "
                    f"Создание отчётов для других школ запрещено."
                )
                _update_progress(0, f"Организация «{org_name_ru}» не совпадает с «{expected_school}».")
                context.close()
                browser.close()
                if logger:
                    logger.finish(success=False)
                return 5  # Код ошибки: несовпадение организации
        elif expected_school and not org_name_ru:
            log_error("Не удалось прочитать название организации — скрапинг отклонён.")
            _update_progress(0, "Не удалось прочитать название организации с mektep.edu.kz.")
            context.close()
            browser.close()
            if logger:
                logger.finish(success=False)
            return 5  # Код ошибки: не удалось прочитать организацию

        # Save profile (teacher) name
        profile_name = _get_profile_name(page)
        if profile_name:
            (out_dir / "profile_name.txt").write_text(profile_name, encoding="utf-8")
            log_info(f"Профиль: {profile_name}")
        else:
            log_warning("Имя профиля не найдено")

        # Переключаем на целевой язык отчётов
        with timing_block("язык отчётов, org_name, период (_choose_period)"):
            log_info(f"Установка языка отчётов: {LANG_MAP[chosen]['label']}")
            _ensure_language(page, chosen)
            log_success(f"Язык установлен: {LANG_MAP[chosen]['label']}")

            # Save org name on target language (for report filenames/content)
            org_name = _get_org_name(page)
            if org_name:
                (out_dir / "org_name.txt").write_text(org_name, encoding="utf-8")
                log_info(f"Организация: {org_name}")
            else:
                org_name = org_name_ru
                if org_name:
                    (out_dir / "org_name.txt").write_text(org_name, encoding="utf-8")
                log_warning("Название организации на текущем языке не найдено, использовано русское")

            # Choose quarter/period for future extraction steps.
            try:
                period_code, period_label = _choose_period()
            except ValueError as e:
                log_error(str(e))
                if logger:
                    logger.finish(success=False)
                return 2
            (out_dir / "period.txt").write_text(period_label, encoding="utf-8")
            log_info(f"Выбранный период: {period_label}")

        # Go to "Оценки" section.
        log_stage(ScraperLogger.STAGE_NAVIGATION, "Переход в раздел 'Оценки'", 7)
        _update_progress(10, "Авторизация завершена, переход к оценкам...")
        with timing_block("_go_to_grades (первый переход)"):
            _go_to_grades(page)
        log_success("Раздел 'Оценки' открыт")

        # Artifacts after navigation to grades.
        if _debug_artifacts_enabled():
            with timing_block("артефакты раздела «Оценки» (url, HTML, full_page screenshot)"):
                (out_dir / "grades.url.txt").write_text(page.url, encoding="utf-8")
                (out_dir / "grades.html").write_text(page.content(), encoding="utf-8")
                page.screenshot(path=str(out_dir / "grades.png"), full_page=True)
            log_info("Сохранён скриншот: grades.png")

        def process_one(selected: dict, batch_subdir: Path) -> None:
            """Строит отчёт по одной паре класс/предмет: критерии, JSON учеников, Excel/Word в batch_subdir."""
            _ensure_dir(batch_subdir)
            legacy = os.getenv("MEKTEP_LEGACY_FILES", "") == "1"
            (batch_subdir / "selected_row.json").write_text(json.dumps(selected, ensure_ascii=False, indent=2), encoding="utf-8")
            # Single metadata file instead of multiple txt files
            meta = {
                "org_name": org_name,
                "profile_name": profile_name,
                "period_code": period_code,
                "period_label": period_label,
                "class": selected.get("class"),
                "subject": selected.get("subject"),
                "criteria_href": selected.get("criteria_href"),
            }
            (batch_subdir / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
            if legacy:
                (batch_subdir / "period.txt").write_text(period_label, encoding="utf-8")
                if org_name:
                    (batch_subdir / "org_name.txt").write_text(org_name, encoding="utf-8")
                if profile_name:
                    (batch_subdir / "profile_name.txt").write_text(profile_name, encoding="utf-8")
                (batch_subdir / "subject.txt").write_text(str(selected.get("subject", "")).strip(), encoding="utf-8")

            class_name = selected.get("class", "Unknown")
            subject_name = selected.get("subject", "Unknown")
            log_stage(ScraperLogger.STAGE_CRITERIA, f"Открытие критериев: {class_name} - {subject_name}", None)
            log_info(f'[{class_name} - {subject_name}] Открытие критериев...')
            criteria_href = selected.get("criteria_href", "")
            criteria_url = selected.get("criteria_url", "")
            if all_mode:
                with timing_block(f"goto criteria (batch) [{class_name} — {subject_name}]"):
                    _open_criteria(page, criteria_href, criteria_url, prefer_goto=True)
            else:
                with timing_block(f"_open_criteria [{class_name} — {subject_name}]"):
                    _open_criteria(page, criteria_href)

            # Prefer lightweight readiness check over fixed delay.
            try:
                page.wait_for_selector(
                    "div#pills-tabContent div.tab-pane.show table, div#pills-tabContent div.tab-pane.active table",
                    timeout=1200,
                )
            except Exception:
                _timed_sleep(0.3, "fallback: рендер страницы критериев перед проверкой предупреждений")

            # Check for warning about missing evaluation data (case 2)
            if _check_criteria_warning(page):
                log_warning(f'[{class_name} - {subject_name}] Обнаружено предупреждение: "Для начала работы необходимо установить данные оценивания!" - пропуск страницы.')
                return
            
            log_success(f'[{class_name} - {subject_name}] Критерии открыты')

            log_info(f'[{class_name} - {subject_name}] Выбор вкладки периода...')
            with timing_block(f"_analyze_and_select_criteria_tabs [{class_name} — {subject_name}]"):
                selected_tab = _analyze_and_select_criteria_tabs(page, batch_subdir, period_code) or _get_active_criteria_tab_href(page)
            if not selected_tab:
                log_error(f'[{class_name} - {subject_name}] Не удалось определить вкладку критериев, пропуск.')
                return
            log_info(f'[{class_name} - {subject_name}] Вкладка выбрана: {selected_tab}')

            with timing_block(f"_has_quarter_grade_header [{class_name} — {subject_name}]"):
                has_quarter_grade_header = _has_quarter_grade_header(page, selected_tab)
            log_info(f'[{class_name} - {subject_name}] Заголовок четвертной оценки: {"да" if has_quarter_grade_header else "нет"}')

            log_stage(ScraperLogger.STAGE_STUDENTS, f"Извлечение учащихся: {class_name}", None)
            log_info(f'[{class_name} - {subject_name}] Извлечение данных учащихся...')
            with timing_block(f"_extract_students_from_criteria_tab [{class_name} — {subject_name}]"):
                students = _extract_students_from_criteria_tab(page, selected_tab)
            students_count = len(students)
            log_success(f'[{class_name} - {subject_name}] Найдено учащихся: {students_count}')
            (batch_subdir / "criteria_students.json").write_text(json.dumps(students, ensure_ascii=False, indent=2), encoding="utf-8")
            with (batch_subdir / "criteria_students.csv").open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(
                    f,
                    fieldnames=["num", "fio", "quarter_num", "average", "formative_pct", "sor_pct", "soch_pct", "total_pct", "grade"],
                )
                w.writeheader()
                for s in students:
                    w.writerow({k: s.get(k, "") for k in w.fieldnames})

            ctx = {
                "org_name": org_name,
                "profile_name": profile_name,
                "class": selected.get("class"),
                "subject": selected.get("subject"),
                "period_code": period_code,
                "period_label": period_label,
                "selected_tab": selected_tab,
                "criteria_url": page.url,
                "has_quarter_grade_header": has_quarter_grade_header,
            }
            (batch_subdir / "criteria_context.json").write_text(json.dumps(ctx, ensure_ascii=False, indent=2), encoding="utf-8")

            with timing_block(f"_extract_quarter_max_points [{class_name} — {subject_name}]"):
                max_points = _extract_quarter_max_points(page, selected_tab)
            (batch_subdir / "criteria_max_points.json").write_text(json.dumps(max_points, ensure_ascii=False, indent=2), encoding="utf-8")

            if students_count == 0:
                log_warning(f'[{class_name} - {subject_name}] Нет учащихся: генерация Excel/Word пропущена')
                return

            # Build Excel report
            try:
                from build_report import build_report
                template_path = _resolve_template_path("Шаблон.xlsx")
                if template_path.exists():
                    log_stage(ScraperLogger.STAGE_EXCEL_REPORT, f"Создание Excel: {class_name} - {subject_name}", None)
                    log_info(f'[{class_name} - {subject_name}] Создание Excel отчета...')
                    with timing_block(f"build_report (Excel) [{class_name} — {subject_name}]"):
                        report_path = build_report(
                            template_path=template_path,
                            students_path=batch_subdir / "criteria_students.json",
                            context_path=batch_subdir / "criteria_context.json",
                            out_dir=out_dir / "reports",
                            max_points_path=batch_subdir / "criteria_max_points.json",
                            criteria_html_path=batch_subdir / "criteria.html",
                            org_name_path=batch_subdir / "org_name.txt",
                        )
                    log_success(f'[{class_name} - {subject_name}] Excel отчет создан: {report_path.name}')
                    if logger:
                        logger.report_created(class_name, subject_name, "Excel")

                    # Build Word report
                    try:
                        from build_word_report import build_word_report
                        # Select template based on language (ru/kk)
                        if chosen == "kk":
                            tpl = _resolve_template_path("Шаблон_каз.docx")
                            if not tpl.exists():
                                tpl = _resolve_template_path("Шаблон.docx")  # Fallback to Russian
                                log_warning(f'Казахский шаблон не найден, используется русский')
                        else:
                            tpl = _resolve_template_path("Шаблон.docx")
                        
                        if tpl.exists():
                            log_stage(ScraperLogger.STAGE_WORD_REPORT, f"Создание Word: {class_name} - {subject_name}", None)
                            log_info(f'[{class_name} - {subject_name}] Создание Word отчета ({tpl.name})...')
                            with timing_block(f"build_word_report [{class_name} — {subject_name}]"):
                                word_path = build_word_report(
                                    template_docx=tpl,
                                    report_xlsx=report_path,
                                    out_dir=out_dir / "reports",
                                    context_json=batch_subdir / "criteria_context.json",
                                    lang=chosen,
                                )
                            log_success(f'[{class_name} - {subject_name}] Word отчет создан: {word_path.name}')
                        else:
                            log_warning(f'[{class_name} - {subject_name}] Шаблон Word не найден: {tpl}')
                    except Exception as e:
                        import traceback
                        log_error(f'[{class_name} - {subject_name}] ОШИБКА создания Word отчета', e)
                        if os.getenv("DEBUG", "").lower() == "1":
                            print(traceback.format_exc())
                        # Continue even if Word report fails - Excel is still saved
                else:
                    log_error(f'[{class_name} - {subject_name}] Шаблон Excel не найден: {template_path}')
            except Exception as e:
                import traceback
                log_error(f'[{class_name} - {subject_name}] ОШИБКА создания отчетов', e)
                if os.getenv("DEBUG", "").lower() == "1":
                    print(traceback.format_exc())

        # Extract table rows and either process one (interactive) or all (batch).
        log_stage(ScraperLogger.STAGE_GRADES_TABLE, "Извлечение таблицы оценок", 8)
        log_info("Извлечение данных таблицы оценок...")
        with timing_block("_extract_grades_table"):
            rows = _extract_grades_table(page)
        if not rows:
            log_error("Таблица оценок пуста или не найдена")
            if logger:
                logger.finish(success=False)
            return 3
        log_success(f"Найдено классов/предметов: {len(rows)}")
        _save_grades_table(rows, out_dir)

        all_mode = os.getenv("MEKTEP_ALL", "") == "1"
        if all_mode:
            batch_root = out_dir / "batch"
            limit = int(os.getenv("MEKTEP_LIMIT", "0") or "0")
            rows_to_process = rows
            if limit > 0:
                rows_to_process = rows[:limit]
                log_info(f"Применён лимит: {limit} отчетов")
            total_reports = len(rows_to_process)
            
            # Установка общего количества отчетов в логгер
            if logger:
                logger.set_total_reports(total_reports)
            
            log_info("=" * 60)
            log_info(f"ПАКЕТНЫЙ РЕЖИМ: Создание отчетов для {total_reports} классов/предметов")
            log_info("=" * 60)
            _update_progress(10, f"Начало обработки {total_reports} отчетов...", total_reports, 0)
            
            for idx, r in enumerate(rows_to_process, 1):
                class_name = r.get("class", "Unknown")
                subject_name = r.get("subject", "Unknown")
                log_info("")
                log_info(f"[{idx}/{total_reports}] === Обработка: {class_name} - {subject_name} ===")
                log_info("-" * 60)
                sub = batch_root / _safe_slug(f'{class_name} {subject_name}')
                with timing_block(f"process_one [{idx}/{total_reports}] {class_name} — {subject_name}"):
                    process_one(r, sub)

                # Calculate progress: 10% (auth) to 90% (reports processing)
                progress_percent = min(90, 10 + int((idx / total_reports) * 80))
                _update_progress(
                    progress_percent,
                    f"Обработано отчетов: {idx} из {total_reports}",
                    total_reports,
                    idx,
                )
            log_info("")
            log_info("=" * 60)
            log_success(f"ПАКЕТНЫЙ РЕЖИМ ЗАВЕРШЕН: Создано отчетов")
            log_info("=" * 60)
            _update_progress(90, f"Обработка завершена: {total_reports} отчетов", total_reports, total_reports)
        else:
            try:
                selected = _choose_row(rows)
            except ValueError as e:
                print(str(e))
                return 2
            process_one(selected, out_dir)

        if _debug_artifacts_enabled():
            with timing_block("финальные артефакты criteria + after_login (HTML, screenshots)"):
                (out_dir / "criteria.url.txt").write_text(page.url, encoding="utf-8")
                (out_dir / "criteria.html").write_text(page.content(), encoding="utf-8")
                page.screenshot(path=str(out_dir / "criteria.png"), full_page=True)

                html_path = out_dir / "after_login.html"
                png_path = out_dir / "after_login.png"
                url_path = out_dir / "after_login.url.txt"

                html_path.write_text(page.content(), encoding="utf-8")
                page.screenshot(path=str(png_path), full_page=True)
                url_path.write_text(page.url, encoding="utf-8")

            log_info(f"Сохранён: {html_path}")
            log_info(f"Сохранён: {png_path}")
            log_info(f"Сохранён: {url_path}")

        log_stage(ScraperLogger.STAGE_COMPLETE, "Закрытие браузера", 95)
        context.close()
        browser.close()
        log_success("Браузер закрыт")

    # Финализация логгера
    if logger:
        logger.finish(success=True)
    
    return 0


def main() -> int:
    """Разбирает аргументы CLI, выставляет MEKTEP_* в окружении и вызывает run()."""
    parser = argparse.ArgumentParser(description="Login automation for mektep.edu.kz")
    parser.add_argument("--headless", type=int, default=1, help="1=headless, 0=show browser")
    parser.add_argument("--out", type=str, default="out/mektep", help="Output directory")
    parser.add_argument("--slowmo", type=int, default=0, help="Slow motion (ms) for debugging")
    parser.add_argument("--lang", type=str, default="", help="Language: ru / kk / en (if empty, asks after login)")
    parser.add_argument("--period", type=str, default="", help="Quarter: 1/2/3/4 (if empty, asks after login)")
    parser.add_argument("--pick", type=str, default="", help="Row index in grades table (if empty, asks on run)")
    parser.add_argument("--all", type=int, default=0, help="1 = generate reports for ALL rows in grades table")
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Limit how many rows to process in --all mode (0 = no limit). Useful for quotas.",
    )
    parser.add_argument("--legacy-files", type=int, default=0, help="1 = also write separate *.txt helper files (subject/period/org/teacher)")
    parser.add_argument(
        "--dotenv",
        type=str,
        default="",
        help="Optional path to env file (not .env in this workspace). Example: env.local",
    )
    args = parser.parse_args()

    if args.dotenv:
        load_dotenv(args.dotenv)
    else:
        # Safe default: try env.example name if user copied it to env.local etc.
        load_dotenv()

    if args.lang:
        os.environ["MEKTEP_LANG"] = args.lang
    if args.period:
        os.environ["MEKTEP_PERIOD"] = args.period
    if args.pick:
        os.environ["MEKTEP_PICK"] = args.pick
    if args.all:
        os.environ["MEKTEP_ALL"] = "1"
    if args.limit:
        os.environ["MEKTEP_LIMIT"] = str(args.limit)
    if args.legacy_files:
        os.environ["MEKTEP_LEGACY_FILES"] = "1"

    return run(headless=bool(args.headless), out_dir=Path(args.out), slow_mo_ms=args.slowmo)


if __name__ == "__main__":
    raise SystemExit(main())

