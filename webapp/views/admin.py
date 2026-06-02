import json
import secrets
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from flask import (
    Blueprint,
    abort,
    current_app,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    session,
    url_for,
    flash,
    send_file,
)
from flask_login import current_user
from sqlalchemy import func
from openpyxl import Workbook, load_workbook

from ..extensions import db
from ..models import Role, User, GradeReport, Class, School, ReportFile, TeacherSubject, TeacherClass, SubjectNameAlias
from ..security import decrypt_password, encrypt_password
from ..constants import kazakh_sort_key, normalize_subject_name
from ..services.admin_common import apply_analytics_filters, redirect_back
from ..services.report_teacher import get_report_teacher_name
from ..services.admin_dashboard import (
    YEAR_UI_PERIOD,
    aggregate_class_metrics,
    aggregate_year_metrics,
    chart_series_from_class_totals,
    class_accordion_group,
    class_name_sort_key,
    get_period_reports,
    get_quarter_reports,
    parse_class_grade,
    parse_ui_period_number,
    student_class_summary_category,
    teacher_accordion_group,
    ui_period_display_name,
)
from ..services.grade_reports.analytics import (
    build_analytics_maps,
    sort_analytics_subject_keys,
)
from ..services.grade_reports.context import load_school_period_context
from ..services.grade_reports.overview import (
    build_grades_overview,
    sort_grades_overview_classes,
)
from ..services.grade_reports.payload import (
    report_analytics_payload,
    report_grades_payload,
)
from ..services.grade_reports.excel import (
    build_analytics_workbook,
    build_class_metrics_charts_workbook,
    build_class_teacher_workbook,
    build_grades_class_workbook,
)
from ..services.criteria_grades import (
    build_criteria_period_zip,
    build_criteria_subject_summary,
    build_criteria_table,
    build_final_table,
    build_simple_grades_table,
    collect_classes_with_criteria,
    find_criteria_subject_entry,
    list_criteria_subject_entries,
    criteria_from_grades_payload,
    criteria_period_path_slug,
    final_from_grades_payload,
    has_criteria_data,
    has_final_data,
    is_final_period,
    is_year_period,
    parse_grades_json,
    report_has_criteria_block,
    report_has_final_block,
    safe_path_segment,
)
from ..services.auth_guards import admin_or_superadmin_required as admin_required
from ..services.subject_aliases import ensure_default_aliases, restore_default_aliases
from ..services.year_grades import (
    build_year_student_subjects,
    math_round_percent,
    students_data_from_year_map,
)
from ..translator import gettext as translate_gettext

from iin_utils import normalize_kz_iin

bp = Blueprint("admin", __name__, url_prefix="/admin")


def _iin_taken_by_other_teacher(school_id: int, iin_norm: str, exclude_id: int | None = None) -> bool:
    q = User.query.filter_by(role=Role.TEACHER.value, school_id=school_id, iin=iin_norm)
    if exclude_id is not None:
        q = q.filter(User.id != exclude_id)
    return q.first() is not None


def _redirect_back(fallback_url: str):
    """Backward-compatible wrapper around shared redirect helper."""
    return redirect_back(fallback_url)


def _management_list_context(school_id: int) -> dict:
    """Teachers/classes lists and accordion buckets for the management page."""
    teachers = User.query.filter_by(
        role=Role.TEACHER.value, school_id=school_id
    ).all()
    classes = Class.query.filter_by(school_id=school_id).all()
    teachers.sort(key=lambda t: kazakh_sort_key(t.full_name or t.username))
    classes.sort(key=lambda c: kazakh_sort_key(c.name))
    teachers_by_accordion = {
        "1-4": [],
        "5-9": [],
        "10-11": [],
        "no_leadership": [],
    }
    for t in teachers:
        group = teacher_accordion_group(t, classes)
        teachers_by_accordion[group].append(t)
    classes_by_accordion = {
        "1-4": [],
        "5-9": [],
        "10-11": [],
    }
    for cls in classes:
        group = class_accordion_group(cls.name)
        classes_by_accordion[group].append(cls)
    ensure_default_aliases(school_id)
    subject_aliases = (
        SubjectNameAlias.query.filter_by(school_id=school_id)
        .order_by(SubjectNameAlias.alias_name)
        .all()
    )
    return {
        "teachers": teachers,
        "classes": classes,
        "teachers_by_accordion": teachers_by_accordion,
        "classes_by_accordion": classes_by_accordion,
        "subject_aliases": subject_aliases,
    }


@bp.get("/")
@admin_required
def dashboard():
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    classes = Class.query.filter_by(school_id=current_user.school_id).all()
    active_class_names = {c.name for c in classes}
    school_metrics = aggregate_class_metrics(current_user.school_id, period_number, active_class_names)
    year_metrics = aggregate_year_metrics(current_user.school_id, active_class_names)
    teachers_count = User.query.filter_by(
        role=Role.TEACHER.value, school_id=current_user.school_id
    ).count()
    classes_count = Class.query.filter_by(school_id=current_user.school_id).count()
    return render_template(
        "admin/dashboard.html",
        teachers_count=teachers_count,
        classes_count=classes_count,
        period_number=period_number,
        school_metrics=school_metrics,
        year_metrics=year_metrics,
    )


@bp.get("/management")
@admin_required
def management():
    """Учителя и классы: отдельная страница."""
    return render_template(
        "admin/management.html",
        **_management_list_context(current_user.school_id),
    )


@bp.post("/subject-aliases")
@admin_required
def create_subject_alias():
    """Добавить пару в словарь предметов школы."""
    alias_name = (request.form.get("alias_name") or "").strip()
    canonical_name = (request.form.get("canonical_name") or "").strip()
    school_id = current_user.school_id

    if not alias_name or not canonical_name:
        flash("Укажите оба названия: вариант и каноническое (русское).", "danger")
        return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")

    existing = SubjectNameAlias.query.filter_by(
        school_id=school_id, alias_name=alias_name
    ).first()
    if existing:
        existing.canonical_name = canonical_name
        flash(f'Запись «{alias_name}» обновлена.', "success")
    else:
        db.session.add(
            SubjectNameAlias(
                school_id=school_id,
                alias_name=alias_name,
                canonical_name=canonical_name,
            )
        )
        flash(f'Добавлено: «{alias_name}» → «{canonical_name}».', "success")
    db.session.commit()
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.post("/subject-aliases/<int:alias_id>/edit")
@admin_required
def update_subject_alias(alias_id: int):
    """Обновить запись словаря предметов школы."""
    row = SubjectNameAlias.query.filter_by(
        id=alias_id, school_id=current_user.school_id
    ).first_or_404()
    alias_name = (request.form.get("alias_name") or "").strip()
    canonical_name = (request.form.get("canonical_name") or "").strip()
    if not alias_name or not canonical_name:
        flash("Укажите оба названия: вариант и каноническое (русское).", "danger")
        return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")

    conflict = SubjectNameAlias.query.filter(
        SubjectNameAlias.school_id == current_user.school_id,
        SubjectNameAlias.alias_name == alias_name,
        SubjectNameAlias.id != alias_id,
    ).first()
    if conflict:
        flash(f'Вариант «{alias_name}» уже есть в словаре.', "danger")
        return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")

    row.alias_name = alias_name
    row.canonical_name = canonical_name
    db.session.commit()
    flash(f'Запись «{alias_name}» обновлена.', "success")
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.post("/subject-aliases/<int:alias_id>/delete")
@admin_required
def delete_subject_alias(alias_id: int):
    """Удалить запись из словаря предметов."""
    row = SubjectNameAlias.query.filter_by(
        id=alias_id, school_id=current_user.school_id
    ).first_or_404()
    alias_label = row.alias_name
    db.session.delete(row)
    db.session.commit()
    flash(f'Запись «{alias_label}» удалена из словаря.', "success")
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.post("/subject-aliases/restore-defaults")
@admin_required
def restore_subject_alias_defaults():
    """Добавить недостающие стандартные пары словаря."""
    restore_default_aliases(current_user.school_id)
    flash("Стандартные записи словаря добавлены (существующие не изменены).", "success")
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.get("/class-metrics-charts")
@admin_required
def class_metrics_charts():
    """Статистика качества и успеваемости по классам в виде диаграмм."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    active_class_names = {
        row.name
        for row in Class.query.filter_by(school_id=current_user.school_id).with_entities(Class.name).all()
    }
    agg = aggregate_class_metrics(current_user.school_id, period_number, active_class_names)
    labels, quality_values, success_values = chart_series_from_class_totals(agg["class_totals"])

    lang = session.get("language", "ru")

    def tr_key(key: str) -> str:
        return translate_gettext(key, lang)

    metrics_i18n = {
        "indicator": tr_key("metrics_col_indicator"),
        "total": tr_key("metrics_col_total"),
        "row_quality": tr_key("metrics_row_quality"),
        "row_success": tr_key("metrics_row_success"),
        "chart_quality": tr_key("metrics_chart_quality"),
        "chart_success": tr_key("metrics_chart_success"),
        "class_word": tr_key("metrics_class_word"),
        "parallel_empty": tr_key("metrics_parallel_empty"),
    }

    resp = make_response(
        render_template(
            "admin/class_metrics_charts.html",
            period_number=period_number,
            labels=labels,
            quality_values=quality_values,
            success_values=success_values,
            metrics_i18n=metrics_i18n,
        )
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@bp.get("/class-metrics-charts/download-excel")
@admin_required
def download_class_metrics_charts_excel_legacy():
    """Раньше вид экспорта передавался как ?scope= — перенаправляем на URL с сегментом пути (надёжнее для кэша и url_for)."""
    sk = (request.args.get("scope") or "overall").strip().lower()
    if sk not in ("overall", "parallel"):
        sk = "overall"
    pn = parse_ui_period_number(request.args.get("period_number", 2))
    return redirect(
        url_for("admin.download_class_metrics_charts_excel", export_kind=sk, period_number=pn),
        code=302,
    )


@bp.get("/class-metrics-charts/download-excel/<export_kind>")
@admin_required
def download_class_metrics_charts_excel(export_kind: str):
    """Книга Excel: лист «Таблицы» и гистограммы (overall / parallel)."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    export_kind = (export_kind or "").strip().lower()
    if export_kind not in ("overall", "parallel"):
        abort(404)
    scope = export_kind
    lang = session.get("language", "ru")

    def tr(key: str) -> str:
        return translate_gettext(key, lang)

    active_class_names = {
        row.name
        for row in Class.query.filter_by(school_id=current_user.school_id).with_entities(Class.name).all()
    }
    agg = aggregate_class_metrics(current_user.school_id, period_number, active_class_names)
    output, filename_ascii, filename_local = build_class_metrics_charts_workbook(
        scope=scope,
        period_number=period_number,
        agg=agg,
        tr=tr,
    )
    resp = send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename_ascii,
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{quote(filename_local)}'
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@bp.post("/teachers/create")
@admin_required
def create_teacher():
    username = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    iin_raw = request.form.get("iin", "").strip()
    if not username:
        flash("Логин учителя обязателен.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")
    if User.query.filter_by(username=username).first():
        flash("Такой логин уже существует.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")

    iin_norm = normalize_kz_iin(iin_raw) if iin_raw else None
    if not iin_norm:
        flash("Укажите корректный ИИН (ЖСН): 12 цифр — тот же номер, что для входа на mektep.edu.kz.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")
    if _iin_taken_by_other_teacher(current_user.school_id, iin_norm):
        flash("Этот ИИН уже привязан к другому учителю в школе.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")

    pw = secrets.token_urlsafe(8)
    u = User(
        username=username,
        full_name=full_name or username,
        iin=iin_norm,
        role=Role.TEACHER.value,
        school_id=current_user.school_id,
        is_active=True,
    )
    # Assign per-school sequential number for filesystem paths (teacher_1, teacher_2, ...)
    max_seq = (
        db.session.query(func.max(User.fs_teacher_seq))
        .filter(User.school_id == current_user.school_id, User.role == Role.TEACHER.value)
        .scalar()
    )
    u.fs_teacher_seq = int(max_seq or 0) + 1
    u.set_password(pw)
    u.password_enc = encrypt_password(pw, current_app.config.get("PASSWORD_ENC_KEY", ""))
    db.session.add(u)
    db.session.commit()
    flash(f"Учитель создан. Пароль: {pw}", "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.post("/teachers/import")
@admin_required
def import_teachers():

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Выберите Excel-файл для импорта.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    if not file.filename.lower().endswith(".xlsx"):
        flash("Поддерживается только формат .xlsx.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Не удалось прочитать Excel-файл.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        flash("Файл пустой или не содержит заголовков.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    headers = [str(v).strip().lower() if v is not None else "" for v in header_cells]
    header_map = {name: idx for idx, name in enumerate(headers)}

    required_headers = ("фио", "логин", "пароль")
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        flash(f"Не найдены обязательные столбцы: {', '.join(missing)}.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    fio_idx = header_map["фио"]
    login_idx = header_map["логин"]
    password_idx = header_map["пароль"]
    iin_idx = header_map.get("иин")
    if iin_idx is None:
        iin_idx = header_map.get("жсн")
    if iin_idx is None:
        flash("В Excel нужен столбец «ИИН» или «ЖСН» (12 цифр).", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    max_seq = (
        db.session.query(func.max(User.fs_teacher_seq))
        .filter(User.school_id == current_user.school_id, User.role == Role.TEACHER.value)
        .scalar()
    )
    next_seq = int(max_seq or 0) + 1

    created = 0
    skipped = 0
    seen_usernames = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name_raw = row[fio_idx] if fio_idx < len(row) else None
        username_raw = row[login_idx] if login_idx < len(row) else None
        password_raw = row[password_idx] if password_idx < len(row) else None
        iin_raw = None
        if iin_idx is not None and iin_idx < len(row):
            iin_raw = row[iin_idx]

        full_name = str(full_name_raw).strip() if full_name_raw is not None else ""
        username = str(username_raw).strip() if username_raw is not None else ""
        password = str(password_raw).strip() if password_raw is not None else ""
        iin_norm = normalize_kz_iin(str(iin_raw).strip() if iin_raw is not None else "") if iin_raw is not None else None

        # Пропускаем полностью пустые строки.
        if not full_name and not username and not password:
            continue

        if not username or not password:
            skipped += 1
            continue

        if not iin_norm:
            skipped += 1
            continue

        if _iin_taken_by_other_teacher(current_user.school_id, iin_norm):
            skipped += 1
            continue

        username_key = username.lower()
        if username_key in seen_usernames:
            skipped += 1
            continue
        if User.query.filter_by(username=username).first():
            skipped += 1
            continue

        seen_usernames.add(username_key)

        u = User(
            username=username,
            full_name=full_name or username,
            iin=iin_norm,
            role=Role.TEACHER.value,
            school_id=current_user.school_id,
            is_active=True,
            fs_teacher_seq=next_seq,
        )
        next_seq += 1
        u.set_password(password)
        u.password_enc = encrypt_password(password, current_app.config.get("PASSWORD_ENC_KEY", ""))
        db.session.add(u)
        created += 1

    db.session.commit()

    category = "success" if created else "warning"
    flash(f"Импорт завершён: добавлено {created}, пропущено {skipped}.", category)
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.get("/teachers/import/template")
@admin_required
def download_teachers_import_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    ws.append(["ФИО", "ИИН", "логин", "пароль"])
    ws.append(["Иванов Иван Иванович", "850101300123", "ivanov_i_i", "TempPass123"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Шаблон_импорта_учителей.xlsx",
    )


@bp.get("/teachers/<int:user_id>/password")
@admin_required
def get_teacher_password(user_id: int):
    """AJAX endpoint: return password as JSON."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        return jsonify({"error": "Not found"}), 404
    pw = decrypt_password(u.password_enc, current_app.config.get("PASSWORD_ENC_KEY", ""))
    return jsonify({"username": u.username, "password": pw or "Недоступен"})


@bp.post("/teachers/<int:user_id>/password")
@admin_required
def update_teacher_password(user_id: int):
    """Update teacher password."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        flash("Пользователь не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    new_password = request.form.get("new_password", "").strip()
    if not new_password or len(new_password) < 4:
        flash("Пароль должен быть не менее 4 символов.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    u.set_password(new_password)
    u.password_enc = encrypt_password(new_password, current_app.config.get("PASSWORD_ENC_KEY", ""))
    db.session.commit()
    flash(f"Пароль для {u.username} обновлен.", "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.post("/teachers/<int:user_id>/edit")
@admin_required
def edit_teacher(user_id: int):
    """Редактирование ФИО и ИИН учителя."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        flash("Учитель не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    full_name = request.form.get("full_name", "").strip()
    if not full_name:
        flash("ФИО не может быть пустым.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    iin_raw = request.form.get("iin", "").strip()
    iin_norm = normalize_kz_iin(iin_raw) if iin_raw else None
    if not iin_norm:
        flash("Укажите корректный ИИН (ЖСН): 12 цифр.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    if _iin_taken_by_other_teacher(current_user.school_id, iin_norm, exclude_id=u.id):
        flash("Этот ИИН уже привязан к другому учителю.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    u.full_name = full_name
    u.iin = iin_norm
    db.session.commit()
    flash(f'Данные учителя обновлены: «{full_name}».', "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.post("/teachers/<int:user_id>/delete")
@admin_required
def delete_teacher(user_id: int):
    """Удаление учителя и всех его данных."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        flash("Учитель не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    
    teacher_name = u.full_name or u.username
    
    # Удаляем связанные данные
    GradeReport.query.filter_by(teacher_id=u.id).delete()
    ReportFile.query.filter_by(teacher_id=u.id).delete()
    
    # Удаляем связи учитель-класс и учитель-предмет
    teacher_subjects = TeacherSubject.query.filter_by(teacher_id=u.id).all()
    for ts in teacher_subjects:
        TeacherClass.query.filter_by(teacher_subject_id=ts.id).delete()
    TeacherSubject.query.filter_by(teacher_id=u.id).delete()
    
    # Снимаем классное руководство
    Class.query.filter_by(class_teacher_id=u.id).update({"class_teacher_id": None})
    
    db.session.delete(u)
    db.session.commit()
    flash(f'Учитель "{teacher_name}" удалён.', "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


# ==============================================================================
# Class CRUD Routes
# ==============================================================================

@bp.post("/classes/create")
@admin_required
def create_class():
    """Создание класса"""
    name = request.form.get("name", "").strip()
    class_teacher_id = request.form.get("class_teacher_id")
    if not name:
        flash("Название класса обязательно.", "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    # Проверяем дубликат
    existing = Class.query.filter_by(school_id=current_user.school_id, name=name).first()
    if existing:
        flash(f'Класс "{name}" уже существует.', "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    cls = Class(name=name, school_id=current_user.school_id)
    if class_teacher_id:
        cls.class_teacher_id = int(class_teacher_id)
    db.session.add(cls)
    db.session.commit()
    flash(f'Класс "{name}" создан.', "success")
    return _redirect_back(url_for("admin.management") + "#classes-tab")


@bp.post("/classes/<int:class_id>/edit")
@admin_required
def edit_class(class_id: int):
    """Редактирование класса"""
    cls = db.session.get(Class, class_id)
    if not cls or cls.school_id != current_user.school_id:
        flash("Класс не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    name = request.form.get("name", "").strip()
    if name:
        # Проверяем, нет ли другого класса с таким именем
        dup = Class.query.filter_by(school_id=current_user.school_id, name=name).first()
        if dup and dup.id != class_id:
            flash(f'Класс "{name}" уже существует.', "danger")
            return _redirect_back(url_for("admin.management") + "#classes-tab")
        cls.name = name
    class_teacher_id = request.form.get("class_teacher_id")
    cls.class_teacher_id = int(class_teacher_id) if class_teacher_id else None
    db.session.commit()
    flash(f'Класс "{cls.name}" обновлён.', "success")
    return _redirect_back(url_for("admin.management") + "#classes-tab")


@bp.post("/classes/<int:class_id>/delete")
@admin_required
def delete_class(class_id: int):
    """Удаление класса из списка школы вместе с оценками (GradeReport) и записями файлов отчётов по этому классу."""
    cls = db.session.get(Class, class_id)
    if not cls or cls.school_id != current_user.school_id:
        flash("Класс не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    name = cls.name
    school_id = cls.school_id

    report_files = ReportFile.query.filter_by(school_id=school_id, class_name=name).all()
    for rf in report_files:
        for path_str in (rf.excel_path, rf.word_path):
            if path_str:
                try:
                    Path(path_str).unlink(missing_ok=True)
                except OSError:
                    pass
        db.session.delete(rf)

    grades_n = GradeReport.query.filter_by(school_id=school_id, class_name=name).delete(
        synchronize_session=False
    )

    # Явно удаляем связи учитель-предмет-класс, чтобы ORM не пытался
    # проставлять class_id = NULL (поле NOT NULL в teacher_classes).
    TeacherClass.query.filter_by(class_id=class_id).delete(synchronize_session=False)
    db.session.delete(cls)
    db.session.commit()
    flash(
        f'Класс «{name}» удалён. Удалено записей оценок: {grades_n}, файлов отчётов: {len(report_files)}.',
        "success",
    )
    return _redirect_back(url_for("admin.management") + "#classes-tab")


# ==============================================================================
# Grades Overview Routes
# ==============================================================================

@bp.get("/grades")
@admin_required
def grades_overview():
    """Обзор оценок: список классов со сводкой"""
    
    # Параметры фильтрации (только четверти)
    period_number = parse_ui_period_number(request.args.get("period_number", 2))

    ctx = load_school_period_context(current_user.school_id, period_number)
    classes_data = build_grades_overview(ctx)
    sorted_classes, classes_by_accordion = sort_grades_overview_classes(classes_data)

    return render_template(
        "admin/grades_overview.html",
        classes=sorted_classes,
        classes_by_accordion=classes_by_accordion,
        period_number=period_number
    )


@bp.get("/grades/class/<class_name>")
@admin_required
def grades_class(class_name: str):
    """Сводная таблица оценок класса: ученик × предмет"""
    
    # Параметры (только четверти)
    period_number = parse_ui_period_number(request.args.get("period_number", 2))

    if not Class.query.filter_by(school_id=current_user.school_id, name=class_name).first():
        flash(
            "Этого класса нет в списке школы (возможно, он удалён). Данные в отчётах остаются в базе, но страница недоступна.",
            "warning",
        )
        return redirect(url_for("admin.grades_overview", period_number=period_number))

    school_id = current_user.school_id

    if period_number == YEAR_UI_PERIOD:
        year_map = build_year_student_subjects(
            school_id, class_name, get_quarter_reports
        )
        students_data = students_data_from_year_map(year_map)
        subjects = {subj for subjs in students_data.values() for subj in subjs}
    else:
        reports = get_period_reports(school_id, period_number, class_name=class_name)
        subjects = set()
        students_data = {}

        for report in reports:
            subj = normalize_subject_name(report.subject_name, school_id)
            subjects.add(subj)

            if report.grades_json:
                try:
                    grades_data = report_grades_payload(report)
                    for student in grades_data.get("students", []):
                        name = student.get("name")
                        if not name:
                            continue
                        if name not in students_data:
                            students_data[name] = {}
                        existing = students_data[name].get(subj)
                        new_grade = {
                            "percent": student.get("percent"),
                            "grade": student.get("grade"),
                        }
                        if existing is None or existing.get("grade") is None:
                            students_data[name][subj] = new_grade
                        elif (
                            new_grade.get("grade") is not None
                            and new_grade["grade"] > existing.get("grade", 0)
                        ):
                            students_data[name][subj] = new_grade
                except json.JSONDecodeError:
                    pass

    # Формируем списки для шаблона
    subjects_list = sorted(subjects, key=kazakh_sort_key)
    students_list = []
    
    for name in sorted(students_data.keys(), key=kazakh_sort_key):
        grades = students_data[name]
        
        # Подсчёт 5, 4, 3 по строке (ученику)
        row_count_5 = sum(1 for g in grades.values() if g.get("grade") == 5)
        row_count_4 = sum(1 for g in grades.values() if g.get("grade") == 4)
        row_count_3 = sum(1 for g in grades.values() if g.get("grade") == 3)
        row_count_2 = sum(1 for g in grades.values() if g.get("grade") == 2)
        
        students_list.append({
            "name": name,
            "grades": grades,
            "count_5": row_count_5,
            "count_4": row_count_4,
            "count_3": row_count_3,
            "count_2": row_count_2,
        })
    
    # Подсчёт по столбцам (предметам): кол-во 5,4,3 + качество + успеваемость
    subject_stats = {}
    for subj in subjects_list:
        s5 = s4 = s3 = s2 = 0
        total_in_subj = 0
        for student in students_list:
            gi = student["grades"].get(subj, {})
            g = gi.get("grade")
            if g is not None:
                total_in_subj += 1
                if g == 5:
                    s5 += 1
                elif g == 4:
                    s4 += 1
                elif g == 3:
                    s3 += 1
                else:
                    s2 += 1
        if period_number == YEAR_UI_PERIOD:
            quality = math_round_percent(s5 + s4, total_in_subj) if total_in_subj else 0
            success = math_round_percent(s5 + s4 + s3, total_in_subj) if total_in_subj else 0
        else:
            quality = round((s5 + s4) / total_in_subj * 100, 1) if total_in_subj else 0
            success = round((s5 + s4 + s3) / total_in_subj * 100, 1) if total_in_subj else 0
        subject_stats[subj] = {
            "count_5": s5, "count_4": s4, "count_3": s3, "count_2": s2,
            "total": total_in_subj,
            "quality_percent": quality,
            "success_percent": success
        }
    
    # Карточки сверху: по ученикам (отличники / хорошисты / троишники с двойками), не по ячейкам таблицы
    total_students = len(students_data)
    grades_count = {"5": 0, "4": 0, "3": 0, "2": 0}
    for student in students_list:
        cat = student_class_summary_category(student["grades"])
        if cat == "excellent":
            grades_count["5"] += 1
        elif cat == "good":
            grades_count["4"] += 1
        elif cat == "troishnik":
            grades_count["3"] += 1
        elif cat == "failing":
            grades_count["2"] += 1

    # Метрики карточек считаем строго из распределения 5/4/3/2,
    # чтобы "Качество" всегда совпадало с карточкой "Распределение".
    quality_percent = 0
    success_percent = 0
    distribution_total = sum(grades_count.values())
    if distribution_total > 0:
        quality_percent = round(
            (grades_count["5"] + grades_count["4"]) / distribution_total * 100, 1
        )
        success_percent = round(
            (grades_count["5"] + grades_count["4"] + grades_count["3"])
            / distribution_total
            * 100,
            1,
        )

    return render_template(
        "admin/grades_class.html",
        class_name=class_name,
        subjects=subjects_list,
        students=students_list,
        subject_stats=subject_stats,
        period_number=period_number,
        summary={
            "total_students": total_students,
            "quality_percent": quality_percent,
            "success_percent": success_percent,
            "grades_count": grades_count
        }
    )


# ==============================================================================
# Criteria assessment routes
# ==============================================================================


@bp.get("/criteria")
@admin_required
def criteria_overview():
    """Критериальное оценивание: период → список классов."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))

    ctx = load_school_period_context(current_user.school_id, period_number)
    classes_data = collect_classes_with_criteria(
        ctx.reports,
        ctx.active_class_names,
        current_user.school_id,
        period_number,
    )

    sorted_classes = sorted(
        classes_data.values(), key=lambda x: kazakh_sort_key(x["class_name"])
    )
    classes_by_accordion = {"1-4": [], "5-9": [], "10-11": []}
    for cls in sorted_classes:
        group = class_accordion_group(cls["class_name"])
        classes_by_accordion[group].append(cls)

    return render_template(
        "admin/criteria_overview.html",
        classes=sorted_classes,
        classes_by_accordion=classes_by_accordion,
        period_number=period_number,
    )


@bp.get("/criteria/download-excel")
@admin_required
def download_criteria_period_excel():
    """ZIP: все классы за период — {орг}/{период}/{класс}/предметы.xlsx."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))

    school = School.query.get(current_user.school_id)
    if not school:
        flash("Школа не найдена.", "danger")
        return redirect(url_for("admin.criteria_overview", period_number=period_number))

    ctx = load_school_period_context(current_user.school_id, period_number)

    zip_buf = build_criteria_period_zip(
        school.name,
        period_number,
        ctx.reports,
        ctx.active_class_names,
        current_user.school_id,
    )
    if not zip_buf:
        flash("Нет данных для выгрузки за выбранный период.", "warning")
        return redirect(url_for("admin.criteria_overview", period_number=period_number))

    org_slug = safe_path_segment(school.name)
    period_slug = criteria_period_path_slug(period_number)
    zip_local = f"{org_slug}/{period_slug}/критериальное_оценивание.zip"
    zip_ascii = f"criteria_{period_number}.zip"

    resp = send_file(
        zip_buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=zip_ascii,
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="{zip_ascii}"; filename*=UTF-8\'\'{quote(zip_local)}'
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@bp.get("/criteria/class/<class_name>")
@admin_required
def criteria_class(class_name: str):
    """Критериальное оценивание: предметы класса."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))

    if not Class.query.filter_by(school_id=current_user.school_id, name=class_name).first():
        flash(
            "Этого класса нет в списке школы (возможно, он удалён).",
            "warning",
        )
        return redirect(url_for("admin.criteria_overview", period_number=period_number))

    school_id = current_user.school_id
    reports = get_period_reports(school_id, period_number, class_name=class_name)

    subjects = list_criteria_subject_entries(
        reports, school_id, period_number, class_name=class_name
    )
    subjects = [
        {
            "name": e["display_name"],
            "teacher": e.get("teacher") or "",
            "report_id": e.get("report_id"),
            "has_criteria": e.get("has_criteria"),
            "has_final": e.get("has_final"),
        }
        for e in subjects
    ]

    return render_template(
        "admin/criteria_class.html",
        class_name=class_name,
        subjects=subjects,
        period_number=period_number,
        year_period=is_year_period(period_number),
        final_period=is_final_period(period_number),
    )


@bp.get("/criteria/class/<class_name>/subject/<path:subject_name>")
@admin_required
def criteria_subject(class_name: str, subject_name: str):
    """Критериальное оценивание: таблица учеников по предмету."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))

    if not Class.query.filter_by(school_id=current_user.school_id, name=class_name).first():
        flash("Этого класса нет в списке школы.", "warning")
        return redirect(url_for("admin.criteria_overview", period_number=period_number))

    school_id = current_user.school_id
    reports = get_period_reports(school_id, period_number, class_name=class_name)

    report_id_arg = request.args.get("report_id", type=int)
    entry = find_criteria_subject_entry(
        reports,
        school_id,
        period_number,
        class_name,
        display_name=subject_name,
        report_id=report_id_arg,
    )

    if not entry or not entry.get("report_id"):
        flash("Нет данных по этому предмету за выбранный период.", "warning")
        return redirect(
            url_for(
                "admin.criteria_class",
                class_name=class_name,
                period_number=period_number,
            )
        )

    report = GradeReport.query.filter_by(
        id=entry["report_id"], school_id=school_id, class_name=class_name
    ).first()
    if not report:
        flash("Отчёт не найден.", "warning")
        return redirect(
            url_for(
                "admin.criteria_class",
                class_name=class_name,
                period_number=period_number,
            )
        )

    display_name = entry["display_name"]
    teacher_name = entry.get("teacher") or get_report_teacher_name(report)
    payload = entry.get("payload") or report_grades_payload(report)
    criteria = criteria_from_grades_payload(payload) if payload else None
    final_block = final_from_grades_payload(payload) if payload else None

    table = None
    show_reupload_hint = False
    if is_final_period(period_number) and final_block and has_final_data(payload):
        table = build_final_table(final_block)
    elif criteria and has_criteria_data(payload):
        table = build_criteria_table(criteria)
    elif is_year_period(period_number) and payload:
        table = build_simple_grades_table(payload)
        show_reupload_hint = False
    else:
        show_reupload_hint = True

    summary = build_criteria_subject_summary(payload) if payload else build_criteria_subject_summary(None)

    return render_template(
        "admin/criteria_subject.html",
        class_name=class_name,
        subject_name=display_name,
        period_number=period_number,
        report_id=entry["report_id"],
        teacher_name=teacher_name,
        table=table,
        show_reupload_hint=show_reupload_hint,
        year_period=is_year_period(period_number),
        final_period=is_final_period(period_number),
        summary=summary,
    )


@bp.post("/grades/class/<class_name>/subjects/delete")
@admin_required
def delete_subject_from_class(class_name: str):
    """Удаление предмета из отчетов указанного класса за выбранную четверть."""

    subject_name = (request.form.get("subject_name") or "").strip()
    period_raw = request.form.get("period_number", "2")
    period_number = parse_ui_period_number(period_raw)

    if not subject_name:
        flash("Не указан предмет для удаления.", "danger")
        return _redirect_back(url_for("admin.grades_class", class_name=class_name, period_number=period_number))

    target_subject = normalize_subject_name(subject_name, current_user.school_id)

    # Удаляем GradeReport по текущему классу и предмету ТОЛЬКО за выбранную четверть.
    # Для 2 и 4 четвертей дополнительно удаляем соответствующее полугодие.
    reports = GradeReport.query.filter_by(
        school_id=current_user.school_id,
        class_name=class_name,
    ).all()

    if period_number == YEAR_UI_PERIOD:
        allowed_periods = {
            ("quarter", 1),
            ("quarter", 2),
            ("quarter", 3),
            ("quarter", 4),
            ("semester", 1),
            ("semester", 2),
        }
    else:
        allowed_periods = {("quarter", period_number)}
        if period_number == 2:
            allowed_periods.add(("semester", 1))
        elif period_number == 4:
            allowed_periods.add(("semester", 2))

    reports_to_delete = [
        r for r in reports
        if normalize_subject_name(r.subject_name, current_user.school_id) == target_subject
        and (r.period_type, r.period_number) in allowed_periods
    ]

    # ReportFile: код четверти 1..4; при удалении «учебного года» — все четверти.
    report_files = ReportFile.query.filter_by(
        school_id=current_user.school_id,
        class_name=class_name,
    ).all()
    if period_number == YEAR_UI_PERIOD:
        period_codes = {"1", "2", "3", "4"}
    else:
        period_codes = {str(period_number)}
    files_to_delete = [
        rf for rf in report_files
        if normalize_subject_name(rf.subject, current_user.school_id) == target_subject
        and str(rf.period_code) in period_codes
    ]

    if not reports_to_delete and not files_to_delete:
        flash(f'Связанные отчёты для предмета "{target_subject}" не найдены.', "warning")
        return _redirect_back(url_for("admin.grades_class", class_name=class_name, period_number=period_number))

    for r in reports_to_delete:
        db.session.delete(r)
    for rf in files_to_delete:
        db.session.delete(rf)
    db.session.commit()

    flash(
        f'Предмет "{target_subject}" удалён: отчётов оценок — {len(reports_to_delete)}, файлов отчётов — {len(files_to_delete)}.',
        "success",
    )
    return _redirect_back(url_for("admin.grades_class", class_name=class_name, period_number=period_number))


@bp.get("/analytics")
@admin_required
def analytics_home():
    """
    Аналитика: 3 вкладки — СОР / СОЧ / Оценки.
    По каждому предмету — карточка с таблицей по классам.
    Структура копирует reference проект.
    """
    
    # Параметры (только четверти)
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    segment = request.args.get("segment")  # '1-4' или '5-11' или None

    ctx = load_school_period_context(current_user.school_id, period_number)
    subjects_data_sor, subjects_data_soch, subjects_data_grades = build_analytics_maps(
        ctx, segment=segment
    )
    subjects_data_sor, subjects_data_soch, subjects_data_grades = sort_analytics_subject_keys(
        subjects_data_sor, subjects_data_soch, subjects_data_grades
    )

    return render_template(
        "admin/analytics_home.html",
        subjects_data_sor=subjects_data_sor,
        subjects_data_soch=subjects_data_soch,
        subjects_data_grades=subjects_data_grades,
        period_number=period_number,
        segment=segment
    )


def _apply_analytics_filters(subjects_data_sor, subjects_data_soch, subjects_data_grades,
                             filter_subject, filter_class, filter_teacher):
    """Backward-compatible wrapper around shared analytics filters."""
    return apply_analytics_filters(
        subjects_data_sor,
        subjects_data_soch,
        subjects_data_grades,
        filter_subject,
        filter_class,
        filter_teacher,
    )


@bp.get("/analytics/download-excel")
@admin_required
def download_analytics_excel():
    """Скачать аналитику СОР/СОЧ/Оценки в Excel (с учётом фильтров subject/class/teacher)"""
    
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    filter_subject = request.args.get("subject", "").strip() or None
    filter_class = request.args.get("class", "").strip() or None
    filter_teacher = request.args.get("teacher", "").strip() or None
    lang = session.get("language", "ru")
    period_name = ui_period_display_name(period_number, lambda k: translate_gettext(k, lang))
    
    ctx = load_school_period_context(current_user.school_id, period_number)
    subjects_data_sor, subjects_data_soch, subjects_data_grades = build_analytics_maps(ctx)

    # Применяем фильтры (subject, class, teacher)
    if filter_subject or filter_class or filter_teacher:
        subjects_data_sor, subjects_data_soch, subjects_data_grades = _apply_analytics_filters(
            subjects_data_sor, subjects_data_soch, subjects_data_grades,
            filter_subject, filter_class, filter_teacher
        )
    
    output = build_analytics_workbook(
        period_name,
        subjects_data_sor,
        subjects_data_soch,
        subjects_data_grades,
    )
    filename = f"Аналитика_СОР_СОЧ_{period_name.replace(' ', '_')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@bp.get("/class-teacher-report")
@admin_required
def class_teacher_report():
    """
    Отчёт классного руководителя (структура из reference проекта).
    
    6 вкладок:
    - на 5 (отличники): Класс | № | ФИО | Классный руководитель
    - на 4 (хорошисты): Класс | № | ФИО | Классный руководитель
    - С одной 4: Класс | № | ФИО | Предмет | Учитель | Классный руководитель
    - на 3 (троечники): Класс | ФИО | Предмет 1..5 | Классный руководитель
    - С одной 3: Класс | № | ФИО | Предмет | Учитель | Классный руководитель
    - Неуспевающие: Класс | № | ФИО | Классный руководитель
    
    Данные сгруппированы по классам.
    """
    
    # Параметры (только четверти)
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    segment = request.args.get("segment")  # '1-4' или '5-11' или None
    
    # Получаем все отчёты (включая полугодовые для 2/4)
    all_reports = get_period_reports(current_user.school_id, period_number)
    active_class_names = {
        row.name
        for row in Class.query.filter_by(school_id=current_user.school_id).with_entities(Class.name).all()
    }
    all_reports = [r for r in all_reports if r.class_name in active_class_names]
    all_class_names = {r.class_name for r in all_reports}
    def _parse_grade_from_name(name: str):
        grade_str = ""
        for ch in str(name):
            if ch.isdigit():
                grade_str += ch
            else:
                break
        return int(grade_str) if grade_str else None

    class_names = []
    for cls_name in all_class_names:
        grade_num = _parse_grade_from_name(cls_name)
        if segment == "1-4":
            if grade_num and 1 <= grade_num <= 4:
                class_names.append((grade_num, cls_name))
        elif segment == "5-11":
            if grade_num and 5 <= grade_num <= 11:
                class_names.append((grade_num, cls_name))
        else:
            class_names.append((grade_num if grade_num is not None else 999, cls_name))

    class_names = [name for _, name in sorted(class_names, key=lambda x: (x[0], kazakh_sort_key(x[1])))]
    
    # Собираем данные по каждому классу
    categories_data = {
        "excellent": [],     # на 5
        "good": [],          # на 4
        "one_4": [],         # С одной 4
        "satisfactory": [],  # на 3
        "one_3": [],         # С одной 3
        "poor": []           # Неуспевающие
    }
    
    for cls_name in class_names:
        # Получаем классного руководителя
        cls_obj = Class.query.filter_by(school_id=current_user.school_id, name=cls_name).first()
        class_teacher_name = ""
        if cls_obj and cls_obj.class_teacher:
            class_teacher_name = cls_obj.class_teacher.full_name or cls_obj.class_teacher.username
        
        # Отчёты для класса из уже загруженных
        reports = [r for r in all_reports if r.class_name == cls_name]
        
        # Собираем оценки: name -> {subject_name: grade}
        # И учителей: subject_name -> teacher_name
        students_grades = {}   # name -> {subject_name: grade}
        subject_teachers = {}  # subject_name -> teacher_name
        
        for report in reports:
            subj = normalize_subject_name(report.subject_name, current_user.school_id)
            teacher_name = get_report_teacher_name(report)
            subject_teachers[subj] = teacher_name
            
            if report.grades_json:
                try:
                    grades_data = report_grades_payload(report)
                    for student in grades_data.get("students", []):
                        name = student.get("name")
                        grade = student.get("grade")
                        if name and grade is not None:
                            if name not in students_grades:
                                students_grades[name] = {}
                            prev = students_grades[name].get(subj)
                            if prev is None or grade > prev:
                                students_grades[name][subj] = grade
                except json.JSONDecodeError:
                    pass
        
        # Категоризируем
        excellent_students = []
        good_students = []
        one_4_students = []
        satisfactory_students = []
        troechniki_detailed = []
        one_3_students = []
        poor_students = []
        
        for name, subj_grades in sorted(students_grades.items(), key=lambda item: kazakh_sort_key(item[0])):
            grades_list = list(subj_grades.values())
            if not grades_list:
                continue
            
            count_5 = grades_list.count(5)
            count_4 = grades_list.count(4)
            count_3 = grades_list.count(3)
            count_2 = sum(1 for g in grades_list if g <= 2)
            
            if count_2 > 0:
                # Для неуспевающих: предметы с двойками
                failing_subjects = [
                    {"subject": s, "teacher": subject_teachers.get(s, "")}
                    for s, g in subj_grades.items() if g <= 2
                ]
                for fs in failing_subjects:
                    poor_students.append({
                        "student": name,
                        "subject": fs["subject"],
                        "teacher": fs["teacher"]
                    })
            elif all(g >= 5 for g in grades_list):
                excellent_students.append(name)
            elif count_4 == 1 and count_3 == 0:
                # С одной 4: найдём предмет
                subj_with_4 = next((s for s, g in subj_grades.items() if g == 4), "")
                one_4_students.append({
                    "student": name,
                    "subject": subj_with_4,
                    "teacher": subject_teachers.get(subj_with_4, "")
                })
            elif count_3 == 0:
                good_students.append(name)
            elif count_3 == 1:
                # С одной 3: найдём предмет
                subj_with_3 = next((s for s, g in subj_grades.items() if g == 3), "")
                one_3_students.append({
                    "student": name,
                    "subject": subj_with_3,
                    "teacher": subject_teachers.get(subj_with_3, "")
                })
            else:
                satisfactory_students.append(name)
                # Подробности: предметы с тройками
                subjects_with_3 = [
                    {"subject_name": s, "grade": g}
                    for s, g in subj_grades.items() if g == 3
                ]
                # Разделяем на первые 4 и остальные (для колонок)
                troechniki_detailed.append({
                    "student": name,
                    "subjects_1_4": subjects_with_3[:4],
                    "subjects_5": subjects_with_3[4:]
                })
        
        # Добавляем блок класса в каждую непустую категорию
        class_block = lambda students: {
            "class_name": cls_name,
            "class_teacher": class_teacher_name,
            "students": students
        }
        
        if excellent_students:
            categories_data["excellent"].append(class_block(excellent_students))
        if good_students:
            categories_data["good"].append(class_block(good_students))
        if one_4_students:
            categories_data["one_4"].append(class_block(one_4_students))
        if satisfactory_students:
            block = class_block(satisfactory_students)
            block["troechniki_detailed"] = troechniki_detailed
            categories_data["satisfactory"].append(block)
        if one_3_students:
            categories_data["one_3"].append(class_block(one_3_students))
        if poor_students:
            categories_data["poor"].append(class_block(poor_students))
    
    return render_template(
        "admin/class_teacher_report.html",
        categories_data=categories_data,
        period_number=period_number,
        segment=segment
    )


# ==============================================================================
# Excel Export Routes
# ==============================================================================

@bp.get("/grades/class/<class_name>/download-excel")
@admin_required
def download_grades_class_excel(class_name: str):
    """Скачать сводную таблицу оценок класса в Excel"""
    
    # Параметры (только четверти)
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    
    # Получаем все отчёты для этого класса (включая полугодовые для 2/4)
    reports = get_period_reports(current_user.school_id, period_number, class_name=class_name)
    
    # Собираем данные
    subjects = set()
    students_data = {}  # name -> {subject -> {percent, grade}}
    
    for report in reports:
        subj = normalize_subject_name(report.subject_name, current_user.school_id)
        subjects.add(subj)
        
        if report.grades_json:
            try:
                grades_data = report_grades_payload(report)
                students_list = grades_data.get("students", [])
                
                for student in students_list:
                    name = student.get("name")
                    if not name:
                        continue
                    
                    if name not in students_data:
                        students_data[name] = {}
                    
                    existing = students_data[name].get(subj)
                    new_grade = {"percent": student.get("percent"), "grade": student.get("grade")}
                    if existing is None or existing.get("grade") is None:
                        students_data[name][subj] = new_grade
                    elif new_grade.get("grade") is not None and new_grade["grade"] > existing.get("grade", 0):
                        students_data[name][subj] = new_grade
            except json.JSONDecodeError:
                pass
    
    # Формируем списки
    subjects_list = sorted(subjects, key=kazakh_sort_key)
    students_list = []
    
    for name in sorted(students_data.keys(), key=kazakh_sort_key):
        grades = students_data[name]
        row_count_5 = sum(1 for g in grades.values() if g.get("grade") == 5)
        row_count_4 = sum(1 for g in grades.values() if g.get("grade") == 4)
        row_count_3 = sum(1 for g in grades.values() if g.get("grade") == 3)
        row_count_2 = sum(1 for g in grades.values() if g.get("grade") == 2)
        
        students_list.append({
            "name": name,
            "grades": grades,
            "count_5": row_count_5,
            "count_4": row_count_4,
            "count_3": row_count_3,
            "count_2": row_count_2,
        })
    
    # Статистика по предметам (столбцам)
    subject_stats = {}
    for subj in subjects_list:
        s5 = s4 = s3 = s2 = 0
        total_in_subj = 0
        for student in students_list:
            gi = student["grades"].get(subj, {})
            g = gi.get("grade")
            if g is not None:
                total_in_subj += 1
                if g == 5: s5 += 1
                elif g == 4: s4 += 1
                elif g == 3: s3 += 1
                else: s2 += 1
        quality = round((s5 + s4) / total_in_subj * 100, 1) if total_in_subj else 0
        success = round((s5 + s4 + s3) / total_in_subj * 100, 1) if total_in_subj else 0
        subject_stats[subj] = {"count_5": s5, "count_4": s4, "count_3": s3, "count_2": s2,
                                "total": total_in_subj, "quality_percent": quality, "success_percent": success}
    
    lang = session.get("language", "ru")
    period_name = ui_period_display_name(period_number, lambda k: translate_gettext(k, lang))
    output, filename = build_grades_class_workbook(
        class_name,
        period_name,
        period_number,
        subjects_list,
        students_list,
        subject_stats,
    )
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@bp.get("/class-teacher-report/download-excel")
@admin_required
def download_class_teacher_report_excel():
    """Скачать отчёт классного руководителя в Excel — все классы, по категориям"""
    
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    segment = request.args.get("segment")
    class_filter = (request.args.get("class") or "").strip().lower()
    class_teacher_filter = (request.args.get("class_teacher") or "").strip().lower()
    student_filter = (request.args.get("student") or "").strip().lower()
    lang = session.get("language", "ru")
    period_name = ui_period_display_name(period_number, lambda k: translate_gettext(k, lang))
    
    # --- Собираем данные (повторяем логику class_teacher_report) ---
    all_reports = get_period_reports(current_user.school_id, period_number)
    active_class_names = {
        row.name
        for row in Class.query.filter_by(school_id=current_user.school_id).with_entities(Class.name).all()
    }
    all_reports = [r for r in all_reports if r.class_name in active_class_names]
    all_class_names = {r.class_name for r in all_reports}

    def _parse_grade_from_name(name: str):
        grade_str = ""
        for ch in str(name):
            if ch.isdigit():
                grade_str += ch
            else:
                break
        return int(grade_str) if grade_str else None

    class_names = []
    for cls_name in all_class_names:
        grade_num = _parse_grade_from_name(cls_name)
        if segment == "1-4":
            if grade_num and 1 <= grade_num <= 4:
                class_names.append((grade_num, cls_name))
        elif segment == "5-11":
            if grade_num and 5 <= grade_num <= 11:
                class_names.append((grade_num, cls_name))
        else:
            class_names.append((grade_num if grade_num is not None else 999, cls_name))

    class_names = [name for _, name in sorted(class_names, key=lambda x: (x[0], kazakh_sort_key(x[1])))]
    
    categories_data = {
        "excellent": [], "good": [], "one_4": [],
        "satisfactory": [], "one_3": [], "poor": []
    }
    
    for cls_name in class_names:
        cls_obj = Class.query.filter_by(school_id=current_user.school_id, name=cls_name).first()
        class_teacher_name = ""
        if cls_obj and cls_obj.class_teacher:
            class_teacher_name = cls_obj.class_teacher.full_name or cls_obj.class_teacher.username

        if class_filter and cls_name.strip().lower() != class_filter:
            continue
        if class_teacher_filter and class_teacher_name.strip().lower() != class_teacher_filter:
            continue
        
        reports = [r for r in all_reports if r.class_name == cls_name]
        
        students_grades = {}
        subject_teachers = {}
        for report in reports:
            subj = normalize_subject_name(report.subject_name, current_user.school_id)
            t_name = get_report_teacher_name(report)
            subject_teachers[subj] = t_name
            if report.grades_json:
                try:
                    gd = report_grades_payload(report)
                    for st in gd.get("students", []):
                        nm = st.get("name")
                        gr = st.get("grade")
                        if nm and gr is not None:
                            students_grades.setdefault(nm, {})
                            prev = students_grades[nm].get(subj)
                            if prev is None or gr > prev:
                                students_grades[nm][subj] = gr
                except json.JSONDecodeError:
                    pass
        
        excellent_s, good_s, one4_s, satisf_s, troech_d, one3_s, poor_s = [], [], [], [], [], [], []
        for name, sg in sorted(students_grades.items(), key=lambda item: kazakh_sort_key(item[0])):
            gl = list(sg.values())
            if not gl:
                continue
            c5, c4, c3, c2 = gl.count(5), gl.count(4), gl.count(3), sum(1 for g in gl if g<=2)
            if c2 > 0:
                for s_name, g_val in sg.items():
                    if g_val <= 2:
                        poor_s.append({"student": name, "subject": s_name, "teacher": subject_teachers.get(s_name, "")})
            elif all(g>=5 for g in gl):
                excellent_s.append(name)
            elif c4==1 and c3==0:
                subj4 = next((s for s,g in sg.items() if g==4), "")
                one4_s.append({"student": name, "subject": subj4, "teacher": subject_teachers.get(subj4,"")})
            elif c3==0:
                good_s.append(name)
            elif c3==1:
                subj3 = next((s for s,g in sg.items() if g==3), "")
                one3_s.append({"student": name, "subject": subj3, "teacher": subject_teachers.get(subj3,"")})
            else:
                satisf_s.append(name)
                subjs3 = [{"subject_name": s, "grade": g} for s,g in sg.items() if g==3]
                troech_d.append({"student": name, "subjects_1_4": subjs3[:4], "subjects_5": subjs3[4:]})
        
        if student_filter:
            excellent_s = [s for s in excellent_s if s.strip().lower() == student_filter]
            good_s = [s for s in good_s if s.strip().lower() == student_filter]
            one4_s = [s for s in one4_s if (s.get("student") or "").strip().lower() == student_filter]
            satisf_s = [s for s in satisf_s if s.strip().lower() == student_filter]
            troech_d = [s for s in troech_d if (s.get("student") or "").strip().lower() == student_filter]
            one3_s = [s for s in one3_s if (s.get("student") or "").strip().lower() == student_filter]
            poor_s = [s for s in poor_s if (s.get("student") or "").strip().lower() == student_filter]

        def _block(students):
            return {"class_name": cls_name, "class_teacher": class_teacher_name, "students": students}
        if excellent_s: categories_data["excellent"].append(_block(excellent_s))
        if good_s: categories_data["good"].append(_block(good_s))
        if one4_s: categories_data["one_4"].append(_block(one4_s))
        if satisf_s:
            b = _block(satisf_s); b["troechniki_detailed"] = troech_d; categories_data["satisfactory"].append(b)
        if one3_s: categories_data["one_3"].append(_block(one3_s))
        if poor_s: categories_data["poor"].append(_block(poor_s))
    
    output, filename = build_class_teacher_workbook(categories_data, period_name)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


