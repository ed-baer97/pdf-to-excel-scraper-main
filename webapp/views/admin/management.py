import json
import secrets
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from flask import (
    Blueprint,
    abort,
    current_app,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    session,
    url_for,
    flash,
    send_file,
)
from flask_login import current_user
from sqlalchemy import func
from openpyxl import Workbook, load_workbook

from ...extensions import db
from ...models import Role, User, GradeReport, Class, School, ReportFile, TeacherSubject, TeacherClass, SubjectNameAlias
from ...security import decrypt_password, encrypt_password
from ...constants import kazakh_sort_key, normalize_subject_name
from ...services.admin_common import apply_analytics_filters, redirect_back
from ...services.report_teacher import get_report_teacher_name
from ...services.admin_dashboard import (
    YEAR_UI_PERIOD,
    aggregate_class_metrics,
    aggregate_year_metrics,
    chart_series_from_class_totals,
    class_accordion_group,
    class_name_sort_key,
    get_period_reports,
    get_quarter_reports,
    parse_class_grade,
    parse_ui_period_number,
    student_class_summary_category,
    teacher_accordion_group,
    ui_period_display_name,
)
from ...services.grade_reports.analytics import (
    build_analytics_maps,
    sort_analytics_subject_keys,
)
from ...services.grade_reports.context import load_school_period_context
from ...services.grade_reports.overview import (
    build_grades_overview,
    sort_grades_overview_classes,
)
from ...services.grade_reports.payload import (
    report_analytics_payload,
    report_grades_payload,
)
from ...services.grade_reports.excel import (
    build_analytics_workbook,
    build_class_metrics_charts_workbook,
    build_class_teacher_workbook,
    build_grades_class_workbook,
)
from ...services.criteria_grades import (
    build_criteria_period_zip,
    build_criteria_subject_summary,
    build_criteria_table,
    build_final_table,
    build_simple_grades_table,
    collect_classes_with_criteria,
    find_criteria_subject_entry,
    list_criteria_subject_entries,
    criteria_from_grades_payload,
    criteria_period_path_slug,
    final_from_grades_payload,
    has_criteria_data,
    has_final_data,
    is_final_period,
    is_year_period,
    parse_grades_json,
    report_has_criteria_block,
    report_has_final_block,
    safe_path_segment,
)
from ...services.auth_guards import admin_or_superadmin_required as admin_required
from ...services.subject_aliases import ensure_default_aliases, restore_default_aliases
from ...services.year_grades import (
    build_year_student_subjects,
    math_round_percent,
    students_data_from_year_map,
)
from ...translator import gettext as translate_gettext

from iin_utils import normalize_kz_iin

from . import bp

def _iin_taken_by_other_teacher(school_id: int, iin_norm: str, exclude_id: int | None = None) -> bool:
    q = User.query.filter_by(role=Role.TEACHER.value, school_id=school_id, iin=iin_norm)
    if exclude_id is not None:
        q = q.filter(User.id != exclude_id)
    return q.first() is not None


def _redirect_back(fallback_url: str):
    """Backward-compatible wrapper around shared redirect helper."""
    return redirect_back(fallback_url)


def _management_list_context(school_id: int) -> dict:
    """Teachers/classes lists and accordion buckets for the management page."""
    teachers = User.query.filter_by(
        role=Role.TEACHER.value, school_id=school_id
    ).all()
    classes = Class.query.filter_by(school_id=school_id).all()
    teachers.sort(key=lambda t: kazakh_sort_key(t.full_name or t.username))
    classes.sort(key=lambda c: kazakh_sort_key(c.name))
    teachers_by_accordion = {
        "1-4": [],
        "5-9": [],
        "10-11": [],
        "no_leadership": [],
    }
    for t in teachers:
        group = teacher_accordion_group(t, classes)
        teachers_by_accordion[group].append(t)
    classes_by_accordion = {
        "1-4": [],
        "5-9": [],
        "10-11": [],
    }
    for cls in classes:
        group = class_accordion_group(cls.name)
        classes_by_accordion[group].append(cls)
    ensure_default_aliases(school_id)
    subject_aliases = (
        SubjectNameAlias.query.filter_by(school_id=school_id)
        .order_by(SubjectNameAlias.alias_name)
        .all()
    )
    return {
        "teachers": teachers,
        "classes": classes,
        "teachers_by_accordion": teachers_by_accordion,
        "classes_by_accordion": classes_by_accordion,
        "subject_aliases": subject_aliases,
    }

@bp.get("/")
@admin_required
def dashboard():
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    classes = Class.query.filter_by(school_id=current_user.school_id).all()
    active_class_names = {c.name for c in classes}
    school_metrics = aggregate_class_metrics(current_user.school_id, period_number, active_class_names)
    year_metrics = aggregate_year_metrics(current_user.school_id, active_class_names)
    teachers_count = User.query.filter_by(
        role=Role.TEACHER.value, school_id=current_user.school_id
    ).count()
    classes_count = Class.query.filter_by(school_id=current_user.school_id).count()
    return render_template(
        "admin/dashboard.html",
        teachers_count=teachers_count,
        classes_count=classes_count,
        period_number=period_number,
        school_metrics=school_metrics,
        year_metrics=year_metrics,
    )


@bp.get("/management")
@admin_required
def management():
    """Учителя и классы: отдельная страница."""
    return render_template(
        "admin/management.html",
        **_management_list_context(current_user.school_id),
    )


@bp.post("/subject-aliases")
@admin_required
def create_subject_alias():
    """Добавить пару в словарь предметов школы."""
    alias_name = (request.form.get("alias_name") or "").strip()
    canonical_name = (request.form.get("canonical_name") or "").strip()
    school_id = current_user.school_id

    if not alias_name or not canonical_name:
        flash("Укажите оба названия: вариант и каноническое (русское).", "danger")
        return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")

    existing = SubjectNameAlias.query.filter_by(
        school_id=school_id, alias_name=alias_name
    ).first()
    if existing:
        existing.canonical_name = canonical_name
        flash(f'Запись «{alias_name}» обновлена.', "success")
    else:
        db.session.add(
            SubjectNameAlias(
                school_id=school_id,
                alias_name=alias_name,
                canonical_name=canonical_name,
            )
        )
        flash(f'Добавлено: «{alias_name}» → «{canonical_name}».', "success")
    db.session.commit()
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.post("/subject-aliases/<int:alias_id>/edit")
@admin_required
def update_subject_alias(alias_id: int):
    """Обновить запись словаря предметов школы."""
    row = SubjectNameAlias.query.filter_by(
        id=alias_id, school_id=current_user.school_id
    ).first_or_404()
    alias_name = (request.form.get("alias_name") or "").strip()
    canonical_name = (request.form.get("canonical_name") or "").strip()
    if not alias_name or not canonical_name:
        flash("Укажите оба названия: вариант и каноническое (русское).", "danger")
        return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")

    conflict = SubjectNameAlias.query.filter(
        SubjectNameAlias.school_id == current_user.school_id,
        SubjectNameAlias.alias_name == alias_name,
        SubjectNameAlias.id != alias_id,
    ).first()
    if conflict:
        flash(f'Вариант «{alias_name}» уже есть в словаре.', "danger")
        return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")

    row.alias_name = alias_name
    row.canonical_name = canonical_name
    db.session.commit()
    flash(f'Запись «{alias_name}» обновлена.', "success")
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.post("/subject-aliases/<int:alias_id>/delete")
@admin_required
def delete_subject_alias(alias_id: int):
    """Удалить запись из словаря предметов."""
    row = SubjectNameAlias.query.filter_by(
        id=alias_id, school_id=current_user.school_id
    ).first_or_404()
    alias_label = row.alias_name
    db.session.delete(row)
    db.session.commit()
    flash(f'Запись «{alias_label}» удалена из словаря.', "success")
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.post("/subject-aliases/restore-defaults")
@admin_required
def restore_subject_alias_defaults():
    """Добавить недостающие стандартные пары словаря."""
    restore_default_aliases(current_user.school_id)
    flash("Стандартные записи словаря добавлены (существующие не изменены).", "success")
    return _redirect_back(url_for("admin.management") + "#subjects-dict-tab")


@bp.get("/class-metrics-charts")
@admin_required
def class_metrics_charts():
    """Статистика качества и успеваемости по классам в виде диаграмм."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    active_class_names = {
        row.name
        for row in Class.query.filter_by(school_id=current_user.school_id).with_entities(Class.name).all()
    }
    agg = aggregate_class_metrics(current_user.school_id, period_number, active_class_names)
    labels, quality_values, success_values = chart_series_from_class_totals(agg["class_totals"])

    lang = session.get("language", "ru")

    def tr_key(key: str) -> str:
        return translate_gettext(key, lang)

    metrics_i18n = {
        "indicator": tr_key("metrics_col_indicator"),
        "total": tr_key("metrics_col_total"),
        "row_quality": tr_key("metrics_row_quality"),
        "row_success": tr_key("metrics_row_success"),
        "chart_quality": tr_key("metrics_chart_quality"),
        "chart_success": tr_key("metrics_chart_success"),
        "class_word": tr_key("metrics_class_word"),
        "parallel_empty": tr_key("metrics_parallel_empty"),
    }

    resp = make_response(
        render_template(
            "admin/class_metrics_charts.html",
            period_number=period_number,
            labels=labels,
            quality_values=quality_values,
            success_values=success_values,
            metrics_i18n=metrics_i18n,
        )
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@bp.get("/class-metrics-charts/download-excel")
@admin_required
def download_class_metrics_charts_excel_legacy():
    """Раньше вид экспорта передавался как ?scope= — перенаправляем на URL с сегментом пути (надёжнее для кэша и url_for)."""
    sk = (request.args.get("scope") or "overall").strip().lower()
    if sk not in ("overall", "parallel"):
        sk = "overall"
    pn = parse_ui_period_number(request.args.get("period_number", 2))
    return redirect(
        url_for("admin.download_class_metrics_charts_excel", export_kind=sk, period_number=pn),
        code=302,
    )


@bp.get("/class-metrics-charts/download-excel/<export_kind>")
@admin_required
def download_class_metrics_charts_excel(export_kind: str):
    """Книга Excel: лист «Таблицы» и гистограммы (overall / parallel)."""
    period_number = parse_ui_period_number(request.args.get("period_number", 2))
    export_kind = (export_kind or "").strip().lower()
    if export_kind not in ("overall", "parallel"):
        abort(404)
    scope = export_kind
    lang = session.get("language", "ru")

    def tr(key: str) -> str:
        return translate_gettext(key, lang)

    active_class_names = {
        row.name
        for row in Class.query.filter_by(school_id=current_user.school_id).with_entities(Class.name).all()
    }
    agg = aggregate_class_metrics(current_user.school_id, period_number, active_class_names)
    output, filename_ascii, filename_local = build_class_metrics_charts_workbook(
        scope=scope,
        period_number=period_number,
        agg=agg,
        tr=tr,
    )
    resp = send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename_ascii,
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{quote(filename_local)}'
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@bp.post("/teachers/create")
@admin_required
def create_teacher():
    username = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    iin_raw = request.form.get("iin", "").strip()
    if not username:
        flash("Логин учителя обязателен.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")
    if User.query.filter_by(username=username).first():
        flash("Такой логин уже существует.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")

    iin_norm = normalize_kz_iin(iin_raw) if iin_raw else None
    if not iin_norm:
        flash("Укажите корректный ИИН (ЖСН): 12 цифр — тот же номер, что для входа на mektep.edu.kz.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")
    if _iin_taken_by_other_teacher(current_user.school_id, iin_norm):
        flash("Этот ИИН уже привязан к другому учителю в школе.", "danger")
        return redirect_back(url_for("admin.management") + "#teachers-tab")

    pw = secrets.token_urlsafe(8)
    u = User(
        username=username,
        full_name=full_name or username,
        iin=iin_norm,
        role=Role.TEACHER.value,
        school_id=current_user.school_id,
        is_active=True,
    )
    # Assign per-school sequential number for filesystem paths (teacher_1, teacher_2, ...)
    max_seq = (
        db.session.query(func.max(User.fs_teacher_seq))
        .filter(User.school_id == current_user.school_id, User.role == Role.TEACHER.value)
        .scalar()
    )
    u.fs_teacher_seq = int(max_seq or 0) + 1
    u.set_password(pw)
    u.password_enc = encrypt_password(pw, current_app.config.get("PASSWORD_ENC_KEY", ""))
    db.session.add(u)
    db.session.commit()
    flash(f"Учитель создан. Пароль: {pw}", "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.post("/teachers/import")
@admin_required
def import_teachers():

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Выберите Excel-файл для импорта.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    if not file.filename.lower().endswith(".xlsx"):
        flash("Поддерживается только формат .xlsx.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Не удалось прочитать Excel-файл.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        flash("Файл пустой или не содержит заголовков.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    headers = [str(v).strip().lower() if v is not None else "" for v in header_cells]
    header_map = {name: idx for idx, name in enumerate(headers)}

    required_headers = ("фио", "логин", "пароль")
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        flash(f"Не найдены обязательные столбцы: {', '.join(missing)}.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    fio_idx = header_map["фио"]
    login_idx = header_map["логин"]
    password_idx = header_map["пароль"]
    iin_idx = header_map.get("иин")
    if iin_idx is None:
        iin_idx = header_map.get("жсн")
    if iin_idx is None:
        flash("В Excel нужен столбец «ИИН» или «ЖСН» (12 цифр).", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")

    max_seq = (
        db.session.query(func.max(User.fs_teacher_seq))
        .filter(User.school_id == current_user.school_id, User.role == Role.TEACHER.value)
        .scalar()
    )
    next_seq = int(max_seq or 0) + 1

    created = 0
    skipped = 0
    seen_usernames = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name_raw = row[fio_idx] if fio_idx < len(row) else None
        username_raw = row[login_idx] if login_idx < len(row) else None
        password_raw = row[password_idx] if password_idx < len(row) else None
        iin_raw = None
        if iin_idx is not None and iin_idx < len(row):
            iin_raw = row[iin_idx]

        full_name = str(full_name_raw).strip() if full_name_raw is not None else ""
        username = str(username_raw).strip() if username_raw is not None else ""
        password = str(password_raw).strip() if password_raw is not None else ""
        iin_norm = normalize_kz_iin(str(iin_raw).strip() if iin_raw is not None else "") if iin_raw is not None else None

        # Пропускаем полностью пустые строки.
        if not full_name and not username and not password:
            continue

        if not username or not password:
            skipped += 1
            continue

        if not iin_norm:
            skipped += 1
            continue

        if _iin_taken_by_other_teacher(current_user.school_id, iin_norm):
            skipped += 1
            continue

        username_key = username.lower()
        if username_key in seen_usernames:
            skipped += 1
            continue
        if User.query.filter_by(username=username).first():
            skipped += 1
            continue

        seen_usernames.add(username_key)

        u = User(
            username=username,
            full_name=full_name or username,
            iin=iin_norm,
            role=Role.TEACHER.value,
            school_id=current_user.school_id,
            is_active=True,
            fs_teacher_seq=next_seq,
        )
        next_seq += 1
        u.set_password(password)
        u.password_enc = encrypt_password(password, current_app.config.get("PASSWORD_ENC_KEY", ""))
        db.session.add(u)
        created += 1

    db.session.commit()

    category = "success" if created else "warning"
    flash(f"Импорт завершён: добавлено {created}, пропущено {skipped}.", category)
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.get("/teachers/import/template")
@admin_required
def download_teachers_import_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    ws.append(["ФИО", "ИИН", "логин", "пароль"])
    ws.append(["Иванов Иван Иванович", "850101300123", "ivanov_i_i", "TempPass123"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Шаблон_импорта_учителей.xlsx",
    )


@bp.get("/teachers/<int:user_id>/password")
@admin_required
def get_teacher_password(user_id: int):
    """AJAX endpoint: return password as JSON."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        return jsonify({"error": "Not found"}), 404
    pw = decrypt_password(u.password_enc, current_app.config.get("PASSWORD_ENC_KEY", ""))
    return jsonify({"username": u.username, "password": pw or "Недоступен"})


@bp.post("/teachers/<int:user_id>/password")
@admin_required
def update_teacher_password(user_id: int):
    """Update teacher password."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        flash("Пользователь не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    new_password = request.form.get("new_password", "").strip()
    if not new_password or len(new_password) < 4:
        flash("Пароль должен быть не менее 4 символов.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    u.set_password(new_password)
    u.password_enc = encrypt_password(new_password, current_app.config.get("PASSWORD_ENC_KEY", ""))
    db.session.commit()
    flash(f"Пароль для {u.username} обновлен.", "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.post("/teachers/<int:user_id>/edit")
@admin_required
def edit_teacher(user_id: int):
    """Редактирование ФИО и ИИН учителя."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        flash("Учитель не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    full_name = request.form.get("full_name", "").strip()
    if not full_name:
        flash("ФИО не может быть пустым.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    iin_raw = request.form.get("iin", "").strip()
    iin_norm = normalize_kz_iin(iin_raw) if iin_raw else None
    if not iin_norm:
        flash("Укажите корректный ИИН (ЖСН): 12 цифр.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    if _iin_taken_by_other_teacher(current_user.school_id, iin_norm, exclude_id=u.id):
        flash("Этот ИИН уже привязан к другому учителю.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    u.full_name = full_name
    u.iin = iin_norm
    db.session.commit()
    flash(f'Данные учителя обновлены: «{full_name}».', "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


@bp.post("/teachers/<int:user_id>/delete")
@admin_required
def delete_teacher(user_id: int):
    """Удаление учителя и всех его данных."""
    u = db.session.get(User, user_id)
    if not u or u.role != Role.TEACHER.value or u.school_id != current_user.school_id:
        flash("Учитель не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#teachers-tab")
    
    teacher_name = u.full_name or u.username
    
    # Удаляем связанные данные
    GradeReport.query.filter_by(teacher_id=u.id).delete()
    ReportFile.query.filter_by(teacher_id=u.id).delete()
    
    # Удаляем связи учитель-класс и учитель-предмет
    teacher_subjects = TeacherSubject.query.filter_by(teacher_id=u.id).all()
    for ts in teacher_subjects:
        TeacherClass.query.filter_by(teacher_subject_id=ts.id).delete()
    TeacherSubject.query.filter_by(teacher_id=u.id).delete()
    
    # Снимаем классное руководство
    Class.query.filter_by(class_teacher_id=u.id).update({"class_teacher_id": None})
    
    db.session.delete(u)
    db.session.commit()
    flash(f'Учитель "{teacher_name}" удалён.', "success")
    return _redirect_back(url_for("admin.management") + "#teachers-tab")


# ==============================================================================
# Class CRUD Routes
# ==============================================================================

@bp.post("/classes/create")
@admin_required
def create_class():
    """Создание класса"""
    name = request.form.get("name", "").strip()
    class_teacher_id = request.form.get("class_teacher_id")
    if not name:
        flash("Название класса обязательно.", "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    # Проверяем дубликат
    existing = Class.query.filter_by(school_id=current_user.school_id, name=name).first()
    if existing:
        flash(f'Класс "{name}" уже существует.', "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    cls = Class(name=name, school_id=current_user.school_id)
    if class_teacher_id:
        cls.class_teacher_id = int(class_teacher_id)
    db.session.add(cls)
    db.session.commit()
    flash(f'Класс "{name}" создан.', "success")
    return _redirect_back(url_for("admin.management") + "#classes-tab")


@bp.post("/classes/<int:class_id>/edit")
@admin_required
def edit_class(class_id: int):
    """Редактирование класса"""
    cls = db.session.get(Class, class_id)
    if not cls or cls.school_id != current_user.school_id:
        flash("Класс не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    name = request.form.get("name", "").strip()
    if name:
        # Проверяем, нет ли другого класса с таким именем
        dup = Class.query.filter_by(school_id=current_user.school_id, name=name).first()
        if dup and dup.id != class_id:
            flash(f'Класс "{name}" уже существует.', "danger")
            return _redirect_back(url_for("admin.management") + "#classes-tab")
        cls.name = name
    class_teacher_id = request.form.get("class_teacher_id")
    cls.class_teacher_id = int(class_teacher_id) if class_teacher_id else None
    db.session.commit()
    flash(f'Класс "{cls.name}" обновлён.', "success")
    return _redirect_back(url_for("admin.management") + "#classes-tab")


@bp.post("/classes/<int:class_id>/delete")
@admin_required
def delete_class(class_id: int):
    """Удаление класса из списка школы вместе с оценками (GradeReport) и записями файлов отчётов по этому классу."""
    cls = db.session.get(Class, class_id)
    if not cls or cls.school_id != current_user.school_id:
        flash("Класс не найден.", "danger")
        return _redirect_back(url_for("admin.management") + "#classes-tab")
    name = cls.name
    school_id = cls.school_id

    report_files = ReportFile.query.filter_by(school_id=school_id, class_name=name).all()
    for rf in report_files:
        for path_str in (rf.excel_path, rf.word_path):
            if path_str:
                try:
                    Path(path_str).unlink(missing_ok=True)
                except OSError:
                    pass
        db.session.delete(rf)

    grades_n = GradeReport.query.filter_by(school_id=school_id, class_name=name).delete(
        synchronize_session=False
    )

    # Явно удаляем связи учитель-предмет-класс, чтобы ORM не пытался
    # проставлять class_id = NULL (поле NOT NULL в teacher_classes).
    TeacherClass.query.filter_by(class_id=class_id).delete(synchronize_session=False)
    db.session.delete(cls)
    db.session.commit()
    flash(
        f'Класс «{name}» удалён. Удалено записей оценок: {grades_n}, файлов отчётов: {len(report_files)}.',
        "success",
    )
    return _redirect_back(url_for("admin.management") + "#classes-tab")


# ==============================================================================
# Grades Overview Routes
# ==============================================================================

