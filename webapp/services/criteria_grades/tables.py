"""Критериальное оценивание: разбор grades_json.criteria и построение таблиц."""

from __future__ import annotations

import json
import re
import zipfile
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..constants import kazakh_sort_key, normalize_subject_name
from .grade_reports.payload import parse_grades_json
from .report_teacher import get_report_teacher_name
from .year_grades import YEAR_UI_PERIOD

FINAL_UI_PERIOD = 6


def parse_points_by_section(points: dict[str, str], quarter_num: int) -> dict[int, str]:
    """Группирует значения points по номеру секции razdel (как в scrape_mektep)."""
    out: dict[int, str] = {}
    prefix = f"chetvert_{quarter_num}_razdel_"
    for k, v in (points or {}).items():
        if not isinstance(k, str) or not k.startswith(prefix):
            continue
        parts = k.split("_")
        if len(parts) >= 5 and parts[2] == "razdel":
            try:
                section = int(parts[3])
            except (TypeError, ValueError):
                continue
            out[section] = v
    return out


def criteria_from_grades_payload(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    """Извлекает блок criteria из распарсенного grades_json."""
    if not isinstance(payload, dict):
        return None
    criteria = payload.get("criteria")
    return criteria if isinstance(criteria, dict) else None


def final_from_grades_payload(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    """Извлекает блок final (четвертные/годовые оценки) из grades_json."""
    if not isinstance(payload, dict):
        return None
    final = payload.get("final")
    return final if isinstance(final, dict) else None


def has_final_data(payload: dict[str, Any] | None) -> bool:
    """Есть ли таблица итога (quarter_final scrape)."""
    final = final_from_grades_payload(payload)
    if not final:
        return False
    students = final.get("students")
    return isinstance(students, list) and len(students) > 0


def has_criteria_data(payload: dict[str, Any] | None) -> bool:
    """Есть ли детализированные критериальные строки учеников."""
    criteria = criteria_from_grades_payload(payload)
    if not criteria:
        return False
    students = criteria.get("students")
    return isinstance(students, list) and len(students) > 0


def ordered_criteria_sections(
    criteria: dict[str, Any],
) -> list[tuple[int, str]]:
    """
    Секции для заголовков: СОр 1..N, затем СОЧ (0).
    Возвращает [(section_id, label), ...].
    """
    quarter_num = int(criteria.get("quarter_num") or 0)
    max_points = criteria.get("max_points") or {}
    if not isinstance(max_points, dict):
        max_points = {}

    sections: set[int] = set()
    for sid in criteria.get("sections") or []:
        try:
            sections.add(int(sid))
        except (TypeError, ValueError):
            pass

    for s in criteria.get("students") or []:
        if not isinstance(s, dict):
            continue
        pts = s.get("points") or {}
        if isinstance(pts, dict):
            sections.update(parse_points_by_section(pts, quarter_num).keys())

    for k in max_points:
        try:
            sections.add(int(k))
        except (TypeError, ValueError):
            pass

    sor_secs = sorted(s for s in sections if s > 0)
    result: list[tuple[int, str]] = [(sec, f"СОр {sec}") for sec in sor_secs]
    if 0 in sections:
        result.append((0, "СОЧ"))
    return result


def section_label(section_id: int) -> str:
    if section_id == 0:
        return "СОЧ"
    return f"СОр {section_id}"


def _section_max_points(max_points: dict, section_id: int) -> int | None:
    """Максимальный балл секции из criteria.max_points."""
    if not isinstance(max_points, dict):
        return None
    raw = max_points.get(str(section_id), max_points.get(section_id))
    if raw is None or raw == "":
        return None
    try:
        val = int(raw)
        return val if val > 0 else None
    except (TypeError, ValueError):
        return None


def format_score_with_max(raw: str | int | float | None, section_id: int, max_points: dict) -> str:
    """Балл в формате «17/20», если известен максимум; иначе как есть."""
    if raw is None or raw == "":
        return ""
    text = str(raw).strip()
    if not text or "%" in text:
        return text
    mp = _section_max_points(max_points, section_id)
    if mp is None:
        return text
    return f"{text}/{mp}"


def build_criteria_table(criteria: dict[str, Any]) -> dict[str, Any]:
    """
    Строит заголовки и строки для шаблона.
    Колонки: №, ФИО, ФО, [СОр N...], СОЧ (если sec 0), Сумма, Оценка.
    """
    quarter_num = int(criteria.get("quarter_num") or 0)
    sections = ordered_criteria_sections(criteria)
    max_points = criteria.get("max_points") or {}

    headers = ["№", "ФИО", "ФО"]
    for sec_id, label in sections:
        if sec_id > 0:
            mp = _section_max_points(max_points, sec_id)
            headers.append(f"{label} (макс. {mp})" if mp else label)
    has_soch_col = any(sid == 0 for sid, _ in sections)
    if has_soch_col:
        mp_soch = _section_max_points(max_points, 0)
        headers.append(f"СОЧ (макс. {mp_soch})" if mp_soch else "СОЧ")
    headers.extend(["Сумма", "Оценка"])

    rows: list[dict[str, Any]] = []
    students = criteria.get("students") or []
    sorted_students = sorted(
        [s for s in students if isinstance(s, dict)],
        key=lambda s: int(s.get("num") or 0),
    )

    for s in sorted_students:
        fio = (s.get("fio") or s.get("name") or "").strip()
        if not fio:
            continue
        pts = s.get("points") or {}
        sec_points = (
            parse_points_by_section(pts, quarter_num) if isinstance(pts, dict) else {}
        )

        cells: list[str] = [
            str(s.get("num") or ""),
            fio,
            str(s.get("average") or ""),
        ]
        for sec_id, _label in sections:
            if sec_id > 0:
                cells.append(
                    format_score_with_max(sec_points.get(sec_id, ""), sec_id, max_points)
                )
        if has_soch_col:
            soch_val = sec_points.get(0, "")
            if soch_val:
                cells.append(format_score_with_max(soch_val, 0, max_points))
            else:
                cells.append(str(s.get("soch_pct") or ""))
        cells.append(str(s.get("total_pct") or ""))
        cells.append(str(s.get("grade") or ""))

        rows.append({"cells": cells})

    return {"headers": headers, "rows": rows}


def grade_distribution(payload: dict[str, Any] | None) -> dict[str, int]:
    """Распределение итоговых оценок 5/4/3/2 по payload.students."""
    counts = {"5": 0, "4": 0, "3": 0, "2": 0}
    if not isinstance(payload, dict):
        return counts
    for s in payload.get("students") or []:
        if not isinstance(s, dict):
            continue
        raw = s.get("grade")
        if raw in (None, "", "0", 0):
            continue
        try:
            g = int(float(str(raw).strip()))
        except (TypeError, ValueError):
            continue
        if g >= 5:
            counts["5"] += 1
        elif g >= 4:
            counts["4"] += 1
        elif g >= 3:
            counts["3"] += 1
        else:
            counts["2"] += 1
    return counts


def build_criteria_subject_summary(payload: dict[str, Any] | None) -> dict[str, Any]:
    """Сводка для шапки страницы предмета: распределение, качество, успеваемость."""
    grades_count = grade_distribution(payload)
    total = 0
    if isinstance(payload, dict):
        total = int(payload.get("total_students") or 0)
        if not total:
            total = len(payload.get("students") or [])
    with_grade = sum(grades_count.values())
    quality = payload.get("quality_percent") if isinstance(payload, dict) else 0
    success = payload.get("success_percent") if isinstance(payload, dict) else 0
    if isinstance(payload, dict) and with_grade and (quality in (None, 0) and success in (None, 0)):
        quality = round((grades_count["5"] + grades_count["4"]) / with_grade * 100, 1)
        success = round(
            (grades_count["5"] + grades_count["4"] + grades_count["3"]) / with_grade * 100, 1
        )
    return {
        "grades_count": grades_count,
        "total_students": total,
        "with_grade": with_grade,
        "quality_percent": quality or 0,
        "success_percent": success or 0,
    }


def build_final_table(final: dict[str, Any]) -> dict[str, Any]:
    """Таблица итога: №, ФИО + динамические колонки (четверти, экзамен, итог)."""
    columns = final.get("columns") or []
    headers = ["№", "ФИО"] + [
        str(c.get("label") or c.get("key") or "") for c in columns if isinstance(c, dict)
    ]
    rows: list[dict[str, Any]] = []
    students = final.get("students") or []
    sorted_students = sorted(
        [s for s in students if isinstance(s, dict)],
        key=lambda s: int(s.get("num") or 0),
    )
    for s in sorted_students:
        fio = (s.get("fio") or "").strip()
        if not fio:
            continue
        cells = [str(s.get("num") or ""), fio]
        for col in columns:
            if not isinstance(col, dict):
                continue
            key = col.get("key", "")
            cells.append(str(s.get(key, "")))
        rows.append({"cells": cells})
    return {"headers": headers, "rows": rows}


def build_simple_grades_table(payload: dict[str, Any]) -> dict[str, Any]:
    """Упрощённая таблица для учебного года: №, ФИО, итоговая оценка (без %)."""
    headers = ["№", "ФИО", "Оценка"]
    rows: list[dict[str, Any]] = []
    for i, s in enumerate(payload.get("students") or [], start=1):
        if not isinstance(s, dict):
            continue
        name = (s.get("name") or "").strip()
        if not name:
            continue
        rows.append(
            {
                "cells": [
                    str(i),
                    name,
                    str(s.get("grade") if s.get("grade") is not None else ""),
                ]
            }
        )
    return {"headers": headers, "rows": rows}


def report_has_criteria_block(report: Any) -> bool:
    """Проверяет GradeReport на наличие criteria в grades_json."""
    payload = parse_grades_json(getattr(report, "grades_json", None))
    return has_criteria_data(payload)


def report_has_final_block(report: Any) -> bool:
    """Проверяет GradeReport на наличие final в grades_json."""
    payload = parse_grades_json(getattr(report, "grades_json", None))
    return has_final_data(payload)


def report_eligible_for_criteria_period(
    report: Any,
    period_number: int,
) -> tuple[bool, dict[str, Any] | None]:
    """Подходит ли отчёт для раздела критериального оценивания за период."""
    payload = parse_grades_json(getattr(report, "grades_json", None))
    if is_final_period(period_number):
        return (has_final_data(payload), payload)
    if is_year_period(period_number):
        return (bool(payload), payload)
    return (has_criteria_data(payload), payload)


def list_criteria_subject_entries(
    reports: list,
    school_id: int,
    period_number: int,
    *,
    class_name: str | None = None,
) -> list[dict[str, Any]]:
    """
    Записи предметов для критериального оценивания без слияния отчётов разных учителей.

    Если несколько отчётов с одним normalize_subject_name в классе — display_name:
    «Математика 1», «Математика 2», …
    """
    from collections import defaultdict

    eligible: list[tuple[Any, dict[str, Any]]] = []
    for report in reports:
        if class_name is not None and report.class_name != class_name:
            continue
        ok, payload = report_eligible_for_criteria_period(report, period_number)
        if not ok or payload is None:
            continue
        eligible.append((report, payload))

    groups: dict[tuple[str, str], list[tuple[Any, dict[str, Any]]]] = defaultdict(list)
    for report, payload in eligible:
        base = normalize_subject_name(report.subject_name, school_id)
        groups[(report.class_name, base)].append((report, payload))

    entries: list[dict[str, Any]] = []
    for (cls, base) in sorted(groups.keys(), key=lambda k: (kazakh_sort_key(k[0]), kazakh_sort_key(k[1]))):
        group = groups[(cls, base)]
        group.sort(
            key=lambda item: (
                kazakh_sort_key(get_report_teacher_name(item[0])),
                getattr(item[0], "id", 0) or 0,
            )
        )
        for idx, (report, payload) in enumerate(group, start=1):
            display_name = base if len(group) == 1 else f"{base} {idx}"
            entries.append(
                {
                    "report_id": getattr(report, "id", None),
                    "class_name": cls,
                    "base_name": base,
                    "display_name": display_name,
                    "teacher": get_report_teacher_name(report),
                    "payload": payload,
                    "raw_subject_name": report.subject_name,
                    "has_criteria": has_criteria_data(payload),
                    "has_final": has_final_data(payload),
                }
            )
    return entries


def find_criteria_subject_entry(
    reports: list,
    school_id: int,
    period_number: int,
    class_name: str,
    *,
    display_name: str | None = None,
    report_id: int | None = None,
) -> dict[str, Any] | None:
    """Находит запись предмета по report_id или отображаемому имени."""
    entries = list_criteria_subject_entries(
        reports, school_id, period_number, class_name=class_name
    )
    if report_id is not None:
        for entry in entries:
            if entry.get("report_id") == report_id:
                return entry
        return None
    if display_name:
        for entry in entries:
            if entry.get("display_name") == display_name:
                return entry
        base = normalize_subject_name(display_name, school_id)
        base_matches = [e for e in entries if e.get("base_name") == base]
        if len(base_matches) == 1:
            return base_matches[0]
    return None


def collect_classes_with_criteria(
    reports: list,
    active_class_names: set[str],
    school_id: int,
    period_number: int,
) -> dict[str, dict[str, Any]]:
    """Группирует отчёты по классам; предметы — с нумерацией при нескольких учителях."""
    classes_data: dict[str, dict[str, Any]] = {}
    entries = list_criteria_subject_entries(reports, school_id, period_number)
    for entry in entries:
        class_name = entry["class_name"]
        if class_name not in active_class_names:
            continue
        if class_name not in classes_data:
            classes_data[class_name] = {
                "class_name": class_name,
                "subjects": [],
                "students_count": 0,
            }
        name = entry["display_name"]
        if name not in classes_data[class_name]["subjects"]:
            classes_data[class_name]["subjects"].append(name)
        payload = entry.get("payload") or {}
        total = payload.get("total_students")
        if not total and payload.get("students"):
            total = len(payload.get("students") or [])
        if total:
            classes_data[class_name]["students_count"] = max(
                classes_data[class_name]["students_count"],
                int(total),
            )
    return classes_data


def is_final_period(period_number: int) -> bool:
    return period_number == FINAL_UI_PERIOD


def is_final_period_placeholder(period_number: int) -> bool:
    """Устарело: итог больше не заглушка; оставлено для совместимости."""
    return False


def is_year_period(period_number: int) -> bool:
    return period_number == YEAR_UI_PERIOD


def safe_path_segment(name: str, *, max_len: int = 80) -> str:
    """Безопасный сегмент пути для ZIP / имени файла."""
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", (name or "").strip())
    s = re.sub(r"_+", "_", s).strip("._ ")
    if not s:
        return "unknown"
    return s[:max_len]


def criteria_period_path_slug(period_number: int) -> str:
    if is_year_period(period_number):
        return "учебный_год"
    if is_final_period(period_number):
        return "итог"
    return f"{period_number}_четверть"


def table_for_period_payload(
    period_number: int, payload: dict[str, Any] | None
) -> dict[str, Any] | None:
    """Таблица для Excel: критерии / год / итог — как на странице предмета."""
    if not payload:
        return None
    if is_final_period(period_number):
        final_block = final_from_grades_payload(payload)
        if final_block and has_final_data(payload):
            table = build_final_table(final_block)
            return table if table.get("rows") else None
    elif is_year_period(period_number):
        table = build_simple_grades_table(payload)
        return table if table.get("rows") else None
    criteria = criteria_from_grades_payload(payload)
    if criteria and has_criteria_data(payload):
        table = build_criteria_table(criteria)
        return table if table.get("rows") else None
    return None


def _excel_sheet_title(subject_name: str, used: set[str]) -> str:
    """Имя листа Excel (≤31 символ, уникальное)."""
    base = re.sub(r"[\[\]:*?/\\]", "_", (subject_name or "Предмет").strip())[:31]
    if not base:
        base = "Предмет"
    title = base
    n = 2
    while title in used:
        suffix = f" ({n})"
        title = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        n += 1
    used.add(title)
    return title


# Симметричная сетка шапки: A — подписи, B..M — 4 блока по 3 столбца (12 col).
_META_LABEL_COL = 1
_META_FIRST_COL = 2
_META_BLOCK_COLS = 3
_META_BLOCKS = 4
_META_LAST_COL = _META_FIRST_COL + _META_BLOCKS * _META_BLOCK_COLS - 1  # 13 (M)


def _criteria_workbook_styles() -> dict[str, Any]:
    thin = Side(style="thin", color="CED4DA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "border": border,
        "label_font": Font(bold=True, size=10, color="343A40"),
        "label_fill": PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"),
        "value_font": Font(size=10, color="212529"),
        "value_bold": Font(bold=True, size=11, color="0D6EFD"),
        "sub_label_font": Font(bold=True, size=9, color="495057"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "meta_value_fill": PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid"),
        "table_header_font": Font(bold=True, color="FFFFFF", size=10),
        "table_header_fill": PatternFill(
            start_color="0D6EFD", end_color="0D6EFD", fill_type="solid"
        ),
        "stripe_fill": PatternFill(start_color="F4F8FF", end_color="F4F8FF", fill_type="solid"),
        "grade_styles": {
            "5": {
                "fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"),
                "font": Font(bold=True, size=11, color="198754"),
            },
            "4": {
                "fill": PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="0A58CA"),
            },
            "3": {
                "fill": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
                "font": Font(bold=True, size=11, color="B45309"),
            },
            "2": {
                "fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
                "font": Font(bold=True, size=11, color="B02A37"),
            },
        },
        "metric_styles": {
            "count": {
                "fill": PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="0D6EFD"),
            },
            "quality": {
                "fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"),
                "font": Font(bold=True, size=11, color="198754"),
            },
            "success": {
                "fill": PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="0A58CA"),
            },
            "total": {
                "fill": PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="495057"),
            },
        },
    }


def _block_bounds(block_index: int) -> tuple[int, int]:
    """Индекс блока 0..3 → (first_col, last_col)."""
    c1 = _META_FIRST_COL + block_index * _META_BLOCK_COLS
    return c1, c1 + _META_BLOCK_COLS - 1


def _paint_range(
    ws,
    row: int,
    col_start: int,
    col_end: int,
    *,
    styles: dict,
    fill=None,
    font=None,
) -> None:
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = styles["border"]
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font


def _merge_block(
    ws,
    row_start: int,
    row_end: int,
    block_index: int,
    text: str,
    *,
    styles: dict,
    fill=None,
    font=None,
    align=None,
) -> None:
    c1, c2 = _block_bounds(block_index)
    ws.merge_cells(start_row=row_start, start_column=c1, end_row=row_end, end_column=c2)
    cell = ws.cell(row=row_start, column=c1, value=text)
    cell.font = font or styles["value_font"]
    cell.fill = fill or styles["meta_value_fill"]
    cell.alignment = align or styles["center"]
    cell.border = styles["border"]
    _paint_range(ws, row_start, c1, c2, styles=styles, fill=fill, font=font)
    if row_end > row_start:
        _paint_range(ws, row_end, c1, c2, styles=styles, fill=fill, font=font)


def _merge_row_value(
    ws,
    row: int,
    text: str,
    *,
    styles: dict,
) -> None:
    ws.merge_cells(
        start_row=row,
        start_column=_META_FIRST_COL,
        end_row=row,
        end_column=_META_LAST_COL,
    )
    cell = ws.cell(row=row, column=_META_FIRST_COL, value=text)
    cell.font = styles["value_font"]
    cell.fill = styles["meta_value_fill"]
    cell.alignment = styles["left"]
    cell.border = styles["border"]
    _paint_range(
        ws, row, _META_FIRST_COL, _META_LAST_COL, styles=styles, fill=styles["meta_value_fill"]
    )


def _merge_side_label(ws, row_start: int, row_end: int, text: str, *, styles: dict) -> None:
    ws.merge_cells(
        start_row=row_start,
        start_column=_META_LABEL_COL,
        end_row=row_end,
        end_column=_META_LABEL_COL,
    )
    cell = ws.cell(row=row_start, column=_META_LABEL_COL, value=text)
    cell.font = styles["label_font"]
    cell.fill = styles["label_fill"]
    cell.alignment = styles["center"]
    cell.border = styles["border"]


def _append_subject_sheet_header(
    ws,
    class_name: str,
    subject_name: str,
    teacher_name: str,
    summary: dict[str, Any],
) -> int:
    """
    Симметричная шапка: 4 блока × 3 столбца (B–M), подписи в A.
    Возвращает номер строки заголовка таблицы учеников.
    """
    styles = _criteria_workbook_styles()
    gc = summary.get("grades_count") or {}

    for row, label, value in (
        (1, "Класс", class_name),
        (2, "Предмет", subject_name),
        (3, "ФИО учителя", teacher_name or "—"),
    ):
        c = ws.cell(row=row, column=_META_LABEL_COL, value=label)
        c.font = styles["label_font"]
        c.fill = styles["label_fill"]
        c.alignment = styles["center"]
        c.border = styles["border"]
        _merge_row_value(ws, row, value, styles=styles)
        ws.row_dimensions[row].height = 22

    _merge_side_label(ws, 4, 5, "Распределение", styles=styles)
    for i, grade in enumerate(("5", "4", "3", "2")):
        gs = styles["grade_styles"][grade]
        _merge_block(
            ws,
            4,
            5,
            i,
            f"{grade}\n{gc.get(grade, 0)}",
            styles=styles,
            fill=gs["fill"],
            font=gs["font"],
        )

    _merge_side_label(ws, 6, 7, "Показатели", styles=styles)
    total_students = summary.get("total_students") or summary.get("with_grade", 0)
    metrics = (
        ("С оценкой", f"{summary.get('with_grade', 0)} уч.", "count"),
        ("Качество", f"{summary.get('quality_percent', 0)}%", "quality"),
        ("Успеваемость", f"{summary.get('success_percent', 0)}%", "success"),
        ("Всего в классе", f"{total_students} уч.", "total"),
    )
    for i, (title, val, key) in enumerate(metrics):
        ms = styles["metric_styles"][key]
        _merge_block(
            ws,
            6,
            6,
            i,
            title,
            styles=styles,
            fill=styles["label_fill"],
            font=styles["sub_label_font"],
        )
        _merge_block(
            ws,
            7,
            7,
            i,
            val,
            styles=styles,
            fill=ms["fill"],
            font=ms["font"],
        )

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 26

    return 9


def _apply_table_block(ws, table: dict[str, Any], start_row: int) -> None:
    """Заголовок и строки таблицы учеников со стилем."""
    styles = _criteria_workbook_styles()
    headers = list(table.get("headers") or [])
    if not headers:
        return

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = styles["table_header_font"]
        cell.fill = styles["table_header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    row_idx = start_row + 1
    stripe = False
    for row in table.get("rows") or []:
        if not isinstance(row, dict) or not row.get("cells"):
            continue
        stripe = not stripe
        row_fill = styles["stripe_fill"] if stripe else None
        for col, val in enumerate(row["cells"], start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = styles["value_font"]
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col == 1 else styles["left"]
            if row_fill is not None:
                cell.fill = row_fill
        row_idx += 1


def _autosize_worksheet_columns(ws, max_col: int, *, label_width: float = 16) -> None:
    ws.column_dimensions["A"].width = label_width
    for col in range(2, max_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)) + 2, 36))
        ws.column_dimensions[letter].width = max_len


def build_subjects_workbook(
    class_name: str,
    subject_sheets: list[dict[str, Any]],
) -> BytesIO:
    """Один xlsx: лист на предмет (шапка + таблица учеников)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    for item in subject_sheets:
        subj_name = str(item.get("subject") or "Предмет")
        table = item.get("table") or {}
        payload = item.get("payload")
        teacher = str(item.get("teacher") or "")
        summary = build_criteria_subject_summary(payload if isinstance(payload, dict) else None)

        ws = wb.create_sheet(title=_excel_sheet_title(subj_name, used_titles))
        table_start = _append_subject_sheet_header(
            ws, class_name, subj_name, teacher, summary
        )
        _apply_table_block(ws, table, table_start)

        ncols = max(len(table.get("headers") or []), _META_LAST_COL)
        _autosize_worksheet_columns(ws, ncols, label_width=14)

    if not wb.sheetnames:
        ws = wb.create_sheet("Нет данных")
        ws.append(["Нет данных"])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def collect_subject_tables_for_class(
    reports: list,
    class_name: str,
    period_number: int,
    school_id: int,
) -> list[dict[str, Any]]:
    """Данные по предметам класса за период (отдельный лист на каждый отчёт)."""
    sheets: list[dict[str, Any]] = []
    for entry in list_criteria_subject_entries(
        reports, school_id, period_number, class_name=class_name
    ):
        table = table_for_period_payload(period_number, entry.get("payload"))
        if not table:
            continue
        sheets.append(
            {
                "subject": entry["display_name"],
                "table": table,
                "payload": entry.get("payload"),
                "teacher": entry.get("teacher") or "",
            }
        )
    return sheets


def build_criteria_period_zip(
    org_name: str,
    period_number: int,
    reports: list,
    active_class_names: set[str],
    school_id: int,
) -> BytesIO | None:
    """
    ZIP: {организация}/{период}/{класс}/предметы.xlsx — по листу на предмет в файле.
    Возвращает None, если нет ни одного файла.
    """
    from collections import defaultdict

    org = safe_path_segment(org_name)
    period_slug = criteria_period_path_slug(period_number)

    sheets_by_class: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in list_criteria_subject_entries(reports, school_id, period_number):
        class_name = entry["class_name"]
        if class_name not in active_class_names:
            continue
        table = table_for_period_payload(period_number, entry.get("payload"))
        if not table:
            continue
        sheets_by_class[class_name].append(
            {
                "subject": entry["display_name"],
                "table": table,
                "payload": entry.get("payload"),
                "teacher": entry.get("teacher") or "",
            }
        )

    zip_buf = BytesIO()
    files_added = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for class_name in sorted(sheets_by_class.keys(), key=kazakh_sort_key):
            subject_sheets = sheets_by_class[class_name]
            if not subject_sheets:
                continue
            xlsx_io = build_subjects_workbook(class_name, subject_sheets)
            class_slug = safe_path_segment(class_name)
            arcname = f"{org}/{period_slug}/{class_slug}/предметы.xlsx"
            zf.writestr(arcname, xlsx_io.getvalue())
            files_added += 1

    if files_added == 0:
        return None
    zip_buf.seek(0)
    return zip_buf
