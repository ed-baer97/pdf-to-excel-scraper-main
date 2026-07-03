"""Excel/ZIP-экспорт критериального оценивания."""

from __future__ import annotations

import re
import zipfile
from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ...constants import kazakh_sort_key
from .periods import criteria_period_path_slug, safe_path_segment
from .queries import list_criteria_subject_entries
from .tables import build_criteria_subject_summary


def _excel_sheet_title(subject_name: str, used: set[str]) -> str:
    """Имя листа Excel (≤31 символ, уникальное)."""
    base = re.sub(r"[\[\]:*?/\\]", "_", (subject_name or "Предмет").strip())[:31]
    if not base:
        base = "Предмет"
    title = base
    n = 2
    while title in used:
        suffix = f" ({n})"
        title = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        n += 1
    used.add(title)
    return title


# Симметричная сетка шапки: A — подписи, B..M — 4 блока по 3 столбца (12 col).
_META_LABEL_COL = 1
_META_FIRST_COL = 2
_META_BLOCK_COLS = 3
_META_BLOCKS = 4
_META_LAST_COL = _META_FIRST_COL + _META_BLOCKS * _META_BLOCK_COLS - 1  # 13 (M)


def _criteria_workbook_styles() -> dict[str, Any]:
    thin = Side(style="thin", color="CED4DA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "border": border,
        "label_font": Font(bold=True, size=10, color="343A40"),
        "label_fill": PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"),
        "value_font": Font(size=10, color="212529"),
        "value_bold": Font(bold=True, size=11, color="0D6EFD"),
        "sub_label_font": Font(bold=True, size=9, color="495057"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "meta_value_fill": PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid"),
        "table_header_font": Font(bold=True, color="FFFFFF", size=10),
        "table_header_fill": PatternFill(
            start_color="0D6EFD", end_color="0D6EFD", fill_type="solid"
        ),
        "stripe_fill": PatternFill(start_color="F4F8FF", end_color="F4F8FF", fill_type="solid"),
        "grade_styles": {
            "5": {
                "fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"),
                "font": Font(bold=True, size=11, color="198754"),
            },
            "4": {
                "fill": PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="0A58CA"),
            },
            "3": {
                "fill": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
                "font": Font(bold=True, size=11, color="B45309"),
            },
            "2": {
                "fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
                "font": Font(bold=True, size=11, color="B02A37"),
            },
        },
        "metric_styles": {
            "count": {
                "fill": PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="0D6EFD"),
            },
            "quality": {
                "fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"),
                "font": Font(bold=True, size=11, color="198754"),
            },
            "success": {
                "fill": PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="0A58CA"),
            },
            "total": {
                "fill": PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"),
                "font": Font(bold=True, size=11, color="495057"),
            },
        },
    }


def _block_bounds(block_index: int) -> tuple[int, int]:
    """Индекс блока 0..3 → (first_col, last_col)."""
    c1 = _META_FIRST_COL + block_index * _META_BLOCK_COLS
    return c1, c1 + _META_BLOCK_COLS - 1


def _paint_range(
    ws,
    row: int,
    col_start: int,
    col_end: int,
    *,
    styles: dict,
    fill=None,
    font=None,
) -> None:
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = styles["border"]
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font


def _merge_block(
    ws,
    row_start: int,
    row_end: int,
    block_index: int,
    text: str,
    *,
    styles: dict,
    fill=None,
    font=None,
    align=None,
) -> None:
    c1, c2 = _block_bounds(block_index)
    ws.merge_cells(start_row=row_start, start_column=c1, end_row=row_end, end_column=c2)
    cell = ws.cell(row=row_start, column=c1, value=text)
    cell.font = font or styles["value_font"]
    cell.fill = fill or styles["meta_value_fill"]
    cell.alignment = align or styles["center"]
    cell.border = styles["border"]
    _paint_range(ws, row_start, c1, c2, styles=styles, fill=fill, font=font)
    if row_end > row_start:
        _paint_range(ws, row_end, c1, c2, styles=styles, fill=fill, font=font)


def _merge_row_value(
    ws,
    row: int,
    text: str,
    *,
    styles: dict,
) -> None:
    ws.merge_cells(
        start_row=row,
        start_column=_META_FIRST_COL,
        end_row=row,
        end_column=_META_LAST_COL,
    )
    cell = ws.cell(row=row, column=_META_FIRST_COL, value=text)
    cell.font = styles["value_font"]
    cell.fill = styles["meta_value_fill"]
    cell.alignment = styles["left"]
    cell.border = styles["border"]
    _paint_range(
        ws, row, _META_FIRST_COL, _META_LAST_COL, styles=styles, fill=styles["meta_value_fill"]
    )


def _merge_side_label(ws, row_start: int, row_end: int, text: str, *, styles: dict) -> None:
    ws.merge_cells(
        start_row=row_start,
        start_column=_META_LABEL_COL,
        end_row=row_end,
        end_column=_META_LABEL_COL,
    )
    cell = ws.cell(row=row_start, column=_META_LABEL_COL, value=text)
    cell.font = styles["label_font"]
    cell.fill = styles["label_fill"]
    cell.alignment = styles["center"]
    cell.border = styles["border"]


def _append_subject_sheet_header(
    ws,
    class_name: str,
    subject_name: str,
    teacher_name: str,
    summary: dict[str, Any],
) -> int:
    """
    Симметричная шапка: 4 блока × 3 столбца (B–M), подписи в A.
    Возвращает номер строки заголовка таблицы учеников.
    """
    styles = _criteria_workbook_styles()
    gc = summary.get("grades_count") or {}

    for row, label, value in (
        (1, "Класс", class_name),
        (2, "Предмет", subject_name),
        (3, "ФИО учителя", teacher_name or "—"),
    ):
        c = ws.cell(row=row, column=_META_LABEL_COL, value=label)
        c.font = styles["label_font"]
        c.fill = styles["label_fill"]
        c.alignment = styles["center"]
        c.border = styles["border"]
        _merge_row_value(ws, row, value, styles=styles)
        ws.row_dimensions[row].height = 22

    _merge_side_label(ws, 4, 5, "Распределение", styles=styles)
    for i, grade in enumerate(("5", "4", "3", "2")):
        gs = styles["grade_styles"][grade]
        _merge_block(
            ws,
            4,
            5,
            i,
            f"{grade}\n{gc.get(grade, 0)}",
            styles=styles,
            fill=gs["fill"],
            font=gs["font"],
        )

    _merge_side_label(ws, 6, 7, "Показатели", styles=styles)
    total_students = summary.get("total_students") or summary.get("with_grade", 0)
    metrics = (
        ("С оценкой", f"{summary.get('with_grade', 0)} уч.", "count"),
        ("Качество", f"{summary.get('quality_percent', 0)}%", "quality"),
        ("Успеваемость", f"{summary.get('success_percent', 0)}%", "success"),
        ("Всего в классе", f"{total_students} уч.", "total"),
    )
    for i, (title, val, key) in enumerate(metrics):
        ms = styles["metric_styles"][key]
        _merge_block(
            ws,
            6,
            6,
            i,
            title,
            styles=styles,
            fill=styles["label_fill"],
            font=styles["sub_label_font"],
        )
        _merge_block(
            ws,
            7,
            7,
            i,
            val,
            styles=styles,
            fill=ms["fill"],
            font=ms["font"],
        )

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 26

    return 9


def _apply_table_block(ws, table: dict[str, Any], start_row: int) -> None:
    """Заголовок и строки таблицы учеников со стилем."""
    styles = _criteria_workbook_styles()
    headers = list(table.get("headers") or [])
    if not headers:
        return

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = styles["table_header_font"]
        cell.fill = styles["table_header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    row_idx = start_row + 1
    stripe = False
    for row in table.get("rows") or []:
        if not isinstance(row, dict) or not row.get("cells"):
            continue
        stripe = not stripe
        row_fill = styles["stripe_fill"] if stripe else None
        for col, val in enumerate(row["cells"], start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = styles["value_font"]
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col == 1 else styles["left"]
            if row_fill is not None:
                cell.fill = row_fill
        row_idx += 1


def _autosize_worksheet_columns(ws, max_col: int, *, label_width: float = 16) -> None:
    ws.column_dimensions["A"].width = label_width
    for col in range(2, max_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)) + 2, 36))
        ws.column_dimensions[letter].width = max_len


def build_subjects_workbook(
    class_name: str,
    subject_sheets: list[dict[str, Any]],
) -> BytesIO:
    """Один xlsx: лист на предмет (шапка + таблица учеников)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    for item in subject_sheets:
        subj_name = str(item.get("subject") or "Предмет")
        table = item.get("table") or {}
        payload = item.get("payload")
        teacher = str(item.get("teacher") or "")
        summary = build_criteria_subject_summary(payload if isinstance(payload, dict) else None)

        ws = wb.create_sheet(title=_excel_sheet_title(subj_name, used_titles))
        table_start = _append_subject_sheet_header(
            ws, class_name, subj_name, teacher, summary
        )
        _apply_table_block(ws, table, table_start)

        ncols = max(len(table.get("headers") or []), _META_LAST_COL)
        _autosize_worksheet_columns(ws, ncols, label_width=14)

    if not wb.sheetnames:
        ws = wb.create_sheet("Нет данных")
        ws.append(["Нет данных"])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_criteria_period_zip(
    org_name: str,
    period_number: int,
    reports: list,
    active_class_names: set[str],
    school_id: int,
) -> BytesIO | None:
    """
    ZIP: {организация}/{период}/{класс}/предметы.xlsx — по листу на предмет в файле.
    Возвращает None, если нет ни одного файла.
    """
    from .periods import table_for_period_payload

    org = safe_path_segment(org_name)
    period_slug = criteria_period_path_slug(period_number)

    sheets_by_class: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in list_criteria_subject_entries(reports, school_id, period_number):
        class_name = entry["class_name"]
        if class_name not in active_class_names:
            continue
        table = table_for_period_payload(period_number, entry.get("payload"))
        if not table:
            continue
        sheets_by_class[class_name].append(
            {
                "subject": entry["display_name"],
                "table": table,
                "payload": entry.get("payload"),
                "teacher": entry.get("teacher") or "",
            }
        )

    zip_buf = BytesIO()
    files_added = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for class_name in sorted(sheets_by_class.keys(), key=kazakh_sort_key):
            subject_sheets = sheets_by_class[class_name]
            if not subject_sheets:
                continue
            xlsx_io = build_subjects_workbook(class_name, subject_sheets)
            class_slug = safe_path_segment(class_name)
            arcname = f"{org}/{period_slug}/{class_slug}/предметы.xlsx"
            zf.writestr(arcname, xlsx_io.getvalue())
            files_added += 1

    if files_added == 0:
        return None
    zip_buf.seek(0)
    return zip_buf
