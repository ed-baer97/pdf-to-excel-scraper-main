"""Лист «Качество по ступеням»: таблица показателей 2–11 классов и две диаграммы."""

from __future__ import annotations

from typing import Callable

from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.drawing.colors import ColorChoice
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..excel.charts import apply_rect_table_borders, chart_title_large
from .data import QUALITY_PERIODS, school_grade_distribution_2_11
from .styles import DYNAMICS_HEADER_FILL, style_chart_axes

GRADE_ROW_KEYS = (
    "final_report_grade_row_5",
    "final_report_grade_row_4",
    "final_report_grade_row_3",
    "final_report_grade_row_2",
)
GRADE_CHART_COLORS = ("70AD47", "4472C4", "ED7D31", "C00000")


def _period_header_labels(tr: Callable[[str], str]) -> list[str]:
    return [
        tr("final_report_hdr_q1"),
        tr("final_report_hdr_q2"),
        tr("final_report_hdr_q3"),
        tr("final_report_hdr_q4"),
        tr("final_report_hdr_total"),
    ]


def write_quality_stages_sheet(
    ws,
    school_id: int,
    academic_year: int,
    active_names: set[str],
    tr: Callable[[str], str],
) -> None:
    """Вкладка: показатели качества знаний (таблица + диаграммы 2–11 классы)."""
    period_labels = _period_header_labels(tr)
    last_col = 1 + len(period_labels)
    distributions = [
        school_grade_distribution_2_11(school_id, academic_year, pn, active_names)
        for pn in QUALITY_PERIODS
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=tr("final_report_quality_title"))
    title_cell.font = Font(bold=True, underline="single", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    headers = [tr("final_report_col_indicator")] + period_labels
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.fill = DYNAMICS_HEADER_FILL
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    metric_rows: list[tuple[str, str]] = [
        (key, field)
        for key, field in zip(
            GRADE_ROW_KEYS,
            ("count_5", "count_4", "count_3", "count_2"),
            strict=True,
        )
    ]
    metric_rows += [
        ("metrics_row_success", "success_percent"),
        ("metrics_row_quality", "quality_percent"),
    ]

    row = 3
    for label_key, field in metric_rows:
        label_cell = ws.cell(row=row, column=1, value=tr(label_key))
        label_cell.font = Font(bold=field.endswith("percent"), size=11)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, dist in enumerate(distributions, start=2):
            # Пустой период (нет данных) -> ячейка остаётся пустой, чтобы график
            # не проваливался в ноль и на столбцах не висели нулевые подписи.
            val = dist.get(field) if dist.get("total") else None
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(val, float):
                cell.number_format = "0.0"
        row += 1

    last_data_row = row - 1
    apply_rect_table_borders(ws, 1, last_data_row, 1, last_col)

    ws.column_dimensions["A"].width = 22
    for col_i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    _add_quality_stages_charts(ws, tr, last_col=last_col, last_data_row=last_data_row)


def _add_quality_stages_charts(
    ws,
    tr: Callable[[str], str],
    *,
    last_col: int,
    last_data_row: int,
) -> None:
    """Столбчатая (накопительная) и линейная диаграммы под таблицей."""
    hdr_row = 2
    grade_first_row = 3
    grade_last_row = 6
    pct_first_row = 7
    pct_last_row = 8
    chart_anchor_row = last_data_row + 3

    cats = Reference(ws, min_col=2, min_row=hdr_row, max_col=last_col, max_row=hdr_row)

    y_max = 0
    for col in range(2, last_col + 1):
        col_sum = sum(
            int(ws.cell(row=r, column=col).value or 0)
            for r in range(grade_first_row, grade_last_row + 1)
            if isinstance(ws.cell(row=r, column=col).value, (int, float))
        )
        y_max = max(y_max, col_sum)
    bar_y_max = max(100, int((y_max * 1.15) // 100 + 1) * 100)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.varyColors = False
    bar.style = 2
    bar.title = chart_title_large(tr("final_report_chart_grades_2_11"))
    bar.title.overlay = False
    bar.gapWidth = 60
    style_chart_axes(
        bar,
        y_title=tr("final_report_chart_y_students"),
        y_min=0,
        y_max=bar_y_max,
        y_unit=200 if bar_y_max > 1000 else 100,
    )
    bar.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        showLegendKey=False,
        dLblPos="ctr",
        numFmt="0;-0;;",
    )
    bar.dLbls.numFmt = "0;-0;;"
    bar.dLbls.sourceLinked = False
    bar.display_blanks = "gap"
    bar.width = 24
    bar.height = 14

    for row_idx, color in zip(
        range(grade_first_row, grade_last_row + 1),
        GRADE_CHART_COLORS,
        strict=True,
    ):
        data = Reference(
            ws, min_col=2, min_row=row_idx, max_col=last_col, max_row=row_idx
        )
        ser = Series(data, title=str(ws.cell(row=row_idx, column=1).value))
        ser.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
        ser.graphicalProperties.line.noFill = True
        bar.series.append(ser)
    bar.set_categories(cats)
    bar.legend.position = "r"
    bar.legend.overlay = False
    ws.add_chart(bar, f"A{chart_anchor_row}")

    line_anchor_row = chart_anchor_row + 30
    line = LineChart()
    line.style = 2
    line.title = chart_title_large(tr("final_report_chart_quality_2_11"))
    line.title.overlay = False
    style_chart_axes(
        line,
        y_title=tr("final_report_chart_y_percent"),
        y_min=50,
        y_max=100,
        y_unit=10,
    )
    line.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        showLegendKey=False,
        dLblPos="t",
    )
    line.display_blanks = "span"
    line.width = 24
    line.height = 14

    line_colors = ("7030A0", "00B0B9")
    markers = ("circle", "square")
    for row_idx, color, symbol in zip(
        range(pct_first_row, pct_last_row + 1),
        line_colors,
        markers,
        strict=True,
    ):
        data = Reference(
            ws, min_col=2, min_row=row_idx, max_col=last_col, max_row=row_idx
        )
        ser = Series(data, title=str(ws.cell(row=row_idx, column=1).value))
        ser.smooth = False
        ser.graphicalProperties.line.solidFill = ColorChoice(srgbClr=color)
        ser.graphicalProperties.line.width = 28575
        ser.marker = Marker(symbol=symbol, size=7)
        ser.marker.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
        ser.marker.graphicalProperties.line.solidFill = ColorChoice(srgbClr=color)
        line.series.append(ser)
    line.set_categories(cats)
    line.legend.position = "r"
    line.legend.overlay = False
    ws.add_chart(line, f"A{line_anchor_row}")
