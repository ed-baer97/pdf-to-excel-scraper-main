"""Листы ручного ввода итогового отчёта."""

from __future__ import annotations

import json
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Font

from .styles import add_bar_chart, write_data_row, write_header_row


def write_json_sheet(wb: Workbook, title_key: str, data: Any, tr: Callable[[str], str]) -> None:
    """Сохранить существующий JSON-раздел ГИА в отдельном листе."""
    ws = wb.create_sheet(title=tr(title_key)[:31])
    ws.cell(row=1, column=1, value=tr(title_key)).font = Font(bold=True, size=14)
    if isinstance(data, dict):
        row_i = 3
        for key, value in data.items():
            ws.cell(row=row_i, column=1, value=str(key))
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False, indent=2)
            ws.cell(row=row_i, column=2, value=value)
            row_i += 2 if isinstance(value, str) and "\n" in value else 1
    else:
        ws.cell(row=3, column=1, value=json.dumps(data, ensure_ascii=False, indent=2))


def write_ent_sheet(wb: Workbook, ent: dict, tr: Callable[[str], str]) -> None:
    """Экспортировать сохранённые данные ЕНТ и диаграмму динамики."""
    ws_ent = wb.create_sheet(title=tr("final_report_sheet_ent")[:31])
    ws_ent.cell(row=1, column=1, value=tr("final_report_sheet_ent")).font = Font(bold=True, size=14)
    write_header_row(
        ws_ent,
        3,
        [
            tr("final_report_ent_col_period"),
            tr("final_report_ent_col_count"),
            tr("final_report_ent_col_avg"),
            tr("final_report_ent_col_max"),
        ],
    )
    ent_row = 4
    periods = ent.get("periods") or []
    for period in periods:
        write_data_row(
            ws_ent,
            ent_row,
            [
                period.get("month") or period.get("period") or "",
                period.get("count"),
                period.get("avg_score"),
                period.get("max_score"),
            ],
            number_cols={3},
        )
        ent_row += 1
    if len(periods) >= 2:
        # Keep the chart source below the visible summary fields so it does not
        # overwrite the period table, forecast, or recommendations.
        chart_hdr = ent_row + 6
        for column, period in enumerate(periods, start=2):
            ws_ent.cell(
                row=chart_hdr,
                column=column,
                value=period.get("month") or period.get("period") or "",
            )
        ws_ent.cell(row=chart_hdr + 1, column=1, value=tr("final_report_ent_col_avg"))
        for column, period in enumerate(periods, start=2):
            ws_ent.cell(row=chart_hdr + 1, column=column, value=period.get("avg_score"))
        add_bar_chart(
            wb,
            ws_ent,
            title=tr("final_report_chart_ent"),
            hdr_row=chart_hdr,
            data_row=chart_hdr + 1,
            last_col=1 + len(periods),
            chart_idx=1,
            y_max=140,
        )
    if ent.get("forecast_avg") is not None:
        ws_ent.cell(row=ent_row + 1, column=1, value=tr("final_report_ent_forecast"))
        ws_ent.cell(row=ent_row + 1, column=2, value=ent.get("forecast_avg"))
    if ent.get("recommendations"):
        ws_ent.cell(row=ent_row + 3, column=1, value=tr("final_report_recommendations"))
        ws_ent.cell(row=ent_row + 4, column=1, value=ent.get("recommendations"))


def write_awards_sheet(wb: Workbook, awards: dict, tr: Callable[[str], str]) -> None:
    """Аттестаты: счётчики «Алтын белгі»/с отличием и список учеников."""
    ws_aw = wb.create_sheet(title=tr("final_report_sheet_awards")[:31])
    ws_aw.cell(row=1, column=1, value=tr("final_report_sheet_awards")).font = Font(bold=True, size=14)
    write_header_row(
        ws_aw,
        3,
        [
            tr("final_report_awards_altyn"),
            tr("final_report_awards_excellent_11"),
            tr("final_report_awards_excellent_9"),
        ],
    )
    write_data_row(
        ws_aw,
        4,
        [
            awards.get("altyn_belgi", 0),
            awards.get("excellent_11", 0),
            awards.get("excellent_9", 0),
        ],
    )
    students_aw = awards.get("students") or []
    if students_aw:
        write_header_row(ws_aw, 6, [tr("final_report_col_student_name"), tr("final_report_col_award_type")])
        ar = 7
        for st in students_aw:
            if isinstance(st, dict):
                write_data_row(ws_aw, ar, [st.get("name", ""), st.get("award", "")])
            else:
                write_data_row(ws_aw, ar, [str(st), ""])
            ar += 1
