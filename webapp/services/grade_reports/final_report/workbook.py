"""Оркестратор: собирает многолистовой Excel итогового отчёта из модулей листов."""

from __future__ import annotations

from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ....extensions import db
from ....models import School
from ...academic_year import resolve_academic_year
from ..final_report_data import load_all_sections, load_sections_for_years
from ..periods import class_name_sort_key
from .data import (
    active_class_names,
    class_grade_summary,
    dynamics_year_columns,
    resolve_years,
)
from .dynamics_sheet import write_contingent_dynamics_sheet
from .manual_sheets import write_awards_sheet, write_ent_sheet, write_json_sheet
from .quality_sheet import write_quality_stages_sheet
from .summary_sheets import (
    write_classes_sheet,
    write_excellent_sheet,
    write_parallels_sheet,
    write_quarters_sheet,
    write_subjects_and_problems_sheets,
    write_summary_sheet,
)


def build_final_report_workbook(
    school_id: int,
    *,
    academic_year: int | None = None,
    years_back: int = 3,
    tr: Callable[[str], str],
) -> tuple[BytesIO, str]:
    """Собрать многолистовой Excel итогового отчёта."""
    school = db.session.get(School, school_id)
    school_name = school.name if school else f"School {school_id}"
    anchor_year = resolve_academic_year(academic_year)
    years = resolve_years(school_id, anchor_year, years_back)
    active_names = active_class_names(school_id)
    manual_by_year = load_sections_for_years(school_id, years)

    wb = Workbook()
    year_cols = dynamics_year_columns(anchor_year, years_back)

    # --- Вкладка 1: динамика численности за 3 года ---
    ws_dyn = wb.active
    ws_dyn.title = tr("final_report_sheet_dynamics")[:31]
    ws_dyn.sheet_view.showGridLines = True
    write_contingent_dynamics_sheet(ws_dyn, school_id, anchor_year, years_back, tr)

    # --- Вкладка 2: показатели качества знаний (2–11 классы) ---
    ws_qual = wb.create_sheet(title=tr("final_report_sheet_quality")[:31], index=1)
    ws_qual.sheet_view.showGridLines = True
    write_quality_stages_sheet(ws_qual, school_id, anchor_year, active_names, tr)

    # --- Листы по данным оценок ---
    write_summary_sheet(wb, school_id, school_name, anchor_year, active_names, tr)

    class_summaries = [
        class_grade_summary(school_id, cn, anchor_year)
        for cn in sorted(active_names, key=class_name_sort_key)
    ]
    write_classes_sheet(wb, class_summaries, tr)
    write_parallels_sheet(wb, class_summaries, tr)
    write_quarters_sheet(wb, school_id, anchor_year, active_names, tr)
    write_subjects_and_problems_sheets(wb, school_id, anchor_year, active_names, tr)
    write_excellent_sheet(wb, school_id, anchor_year, active_names, tr)

    # --- ГИА / ЕНТ / Аттестаты (ручной ввод) ---
    manual = manual_by_year.get(anchor_year, load_all_sections(school_id, anchor_year))
    write_json_sheet(wb, "final_report_sheet_gia9", manual.get("gia9", {}), tr)
    write_json_sheet(wb, "final_report_sheet_gia11", manual.get("gia11", {}), tr)
    write_ent_sheet(wb, manual.get("ent", {}), tr)
    write_awards_sheet(wb, manual.get("awards", {}), tr)

    # Auto column widths
    for ws in wb.worksheets:
        for col in range(1, min(ws.max_column + 1, 12)):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 18

    years_label = "_".join(str(y) for y in year_cols)
    filename = f"Итоговый_отчёт_{years_label}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename
