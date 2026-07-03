"""Листы по данным оценок: сводка KPI, классы, ступени, четверти, предметы, проблемы, отличники."""

from __future__ import annotations

from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font

from ....constants import kazakh_sort_key
from ...admin_dashboard import aggregate_class_metrics
from ...academic_year import format_academic_year
from ...year_grades import YEAR_UI_PERIOD
from ..periods import class_name_sort_key
from .data import class_students_map, lowest_quality, parallel_summary, subject_quality_matrix
from .styles import add_bar_chart, write_data_row, write_header_row


def write_summary_sheet(
    wb: Workbook,
    school_id: int,
    school_name: str,
    anchor_year: int,
    active_names: set[str],
    tr: Callable[[str], str],
) -> None:
    """Сводка KPI: качество/успеваемость по школе и ступеням за год."""
    ws_sum = wb.create_sheet(title=tr("final_report_sheet_summary")[:31])
    ws_sum.cell(row=1, column=1, value=tr("final_report_title")).font = Font(bold=True, size=16)
    ws_sum.cell(row=2, column=1, value=school_name)
    ws_sum.cell(row=3, column=1, value=format_academic_year(anchor_year))
    row = 5
    year_agg = aggregate_class_metrics(
        school_id, YEAR_UI_PERIOD, active_names, academic_year=anchor_year
    )
    write_header_row(ws_sum, row, [tr("metrics_col_indicator"), tr("metrics_row_quality"), tr("metrics_row_success")])
    row += 1
    write_data_row(
        ws_sum,
        row,
        [
            tr("metrics_excel_school_weighted"),
            year_agg.get("school_quality"),
            year_agg.get("school_success"),
        ],
        number_cols={2, 3},
    )
    row += 2
    for key, label_key in (
        ("1-4", "metrics_excel_parallel_1_4"),
        ("5-9", "metrics_excel_parallel_5_9"),
        ("10-11", "metrics_excel_parallel_10_11"),
    ):
        pr = (year_agg.get("parallel") or {}).get(key) or {}
        write_data_row(
            ws_sum,
            row,
            [tr(label_key), pr.get("quality"), pr.get("success")],
            number_cols={2, 3},
        )
        row += 1


def write_classes_sheet(
    wb: Workbook,
    class_summaries: list[dict],
    tr: Callable[[str], str],
) -> None:
    """Успеваемость по классам: распределение 5/4/3/2 и проценты."""
    ws_cls = wb.create_sheet(title=tr("final_report_sheet_classes")[:31])
    cls_headers = [
        tr("class"),
        tr("final_report_col_students"),
        tr("final_report_col_on_5"),
        tr("final_report_col_on_4"),
        tr("final_report_col_one_3"),
        tr("final_report_col_two_3"),
        tr("final_report_col_on_3"),
        tr("final_report_col_on_2"),
        tr("metrics_row_success"),
        tr("metrics_row_quality"),
    ]
    write_header_row(ws_cls, 1, cls_headers)
    cr = 2
    for s in class_summaries:
        if s["total"] <= 0:
            continue
        write_data_row(
            ws_cls,
            cr,
            [
                s["class_name"],
                s["total"],
                s["count_5"],
                s["count_4"],
                s["one_3"],
                s["two_plus_3"],
                s["count_3"],
                s["count_2"],
                s["success_percent"],
                s["quality_percent"],
            ],
            number_cols={9, 10},
        )
        cr += 1


def write_parallels_sheet(
    wb: Workbook,
    class_summaries: list[dict],
    tr: Callable[[str], str],
) -> None:
    """По ступеням: агрегаты 1–4 / 5–9 / 10–11."""
    ws_par = wb.create_sheet(title=tr("final_report_sheet_parallels")[:31])
    write_header_row(
        ws_par,
        1,
        [
            tr("final_report_col_stage"),
            tr("final_report_col_students"),
            tr("metrics_row_quality"),
            tr("metrics_row_success"),
        ],
    )
    pr = 2
    for bucket, label_key in (
        ("1-4", "metrics_excel_sec_1_4"),
        ("5-9", "metrics_excel_sec_5_9"),
        ("10-11", "metrics_excel_parallel_10_11"),
    ):
        ps = parallel_summary(class_summaries, bucket)
        write_data_row(
            ws_par,
            pr,
            [tr(label_key), ps["total"], ps["quality_percent"], ps["success_percent"]],
            number_cols={3, 4},
        )
        pr += 1


def write_quarters_sheet(
    wb: Workbook,
    school_id: int,
    anchor_year: int,
    active_names: set[str],
    tr: Callable[[str], str],
) -> None:
    """По четвертям: таблица + скрытые ряды для диаграммы динамики."""
    ws_q = wb.create_sheet(title=tr("final_report_sheet_quarters")[:31])
    write_header_row(
        ws_q,
        1,
        [
            tr("quarter_label"),
            tr("metrics_row_quality"),
            tr("metrics_row_success"),
            tr("final_report_col_classes_with_data"),
        ],
    )
    qr = 2
    quarter_chart_hdr = 20
    ws_q.cell(row=quarter_chart_hdr, column=1, value=tr("quarter_label"))
    for qi, qn in enumerate([1, 2, 3, 4, YEAR_UI_PERIOD], start=2):
        label = (
            tr(f"metrics_excel_period_q{qn}")
            if qn <= 4
            else tr("metrics_excel_period_year")
        )
        ws_q.cell(row=quarter_chart_hdr, column=qi, value=label)
    ws_q.cell(row=quarter_chart_hdr + 1, column=1, value=tr("metrics_row_quality"))
    ws_q.cell(row=quarter_chart_hdr + 2, column=1, value=tr("metrics_row_success"))
    for qi, qn in enumerate([1, 2, 3, 4, YEAR_UI_PERIOD], start=2):
        agg = aggregate_class_metrics(
            school_id, qn, active_names, academic_year=anchor_year
        )
        label = (
            tr(f"metrics_excel_period_q{qn}")
            if qn <= 4
            else tr("metrics_excel_period_year")
        )
        write_data_row(
            ws_q,
            qr,
            [label, agg.get("school_quality"), agg.get("school_success"), agg.get("classes_with_data")],
            number_cols={2, 3},
        )
        ws_q.cell(row=quarter_chart_hdr + 1, column=qi, value=agg.get("school_quality"))
        ws_q.cell(row=quarter_chart_hdr + 2, column=qi, value=agg.get("school_success"))
        qr += 1
    add_bar_chart(
        wb,
        ws_q,
        title=tr("final_report_chart_quarters"),
        hdr_row=quarter_chart_hdr,
        data_row=quarter_chart_hdr + 1,
        last_col=6,
        chart_idx=0,
    )


def write_subjects_and_problems_sheets(
    wb: Workbook,
    school_id: int,
    anchor_year: int,
    active_names: set[str],
    tr: Callable[[str], str],
) -> None:
    """Качество по предметам и «проблемные зоны» (нижние 10 по качеству)."""
    subj_rows = subject_quality_matrix(school_id, anchor_year, active_names)
    ws_subj = wb.create_sheet(title=tr("final_report_sheet_subjects")[:31])
    write_header_row(
        ws_subj,
        1,
        [tr("subject"), tr("class"), tr("metrics_row_quality"), tr("metrics_row_success")],
    )
    sr = 2
    for r in subj_rows:
        write_data_row(
            ws_subj,
            sr,
            [r["subject"], r["class_name"], r["quality_percent"], r["success_percent"]],
            number_cols={3, 4},
        )
        sr += 1

    ws_prob = wb.create_sheet(title=tr("final_report_sheet_problems")[:31])
    write_header_row(
        ws_prob,
        1,
        [tr("class"), tr("subject"), tr("metrics_row_quality")],
    )
    prb = 2
    for r in lowest_quality(subj_rows):
        write_data_row(
            ws_prob,
            prb,
            [r["class_name"], r["subject"], r["quality_percent"]],
            number_cols={3},
        )
        prb += 1
    ws_prob.cell(row=prb + 1, column=1, value=tr("final_report_recommendations"))
    ws_prob.cell(row=prb + 2, column=1, value=tr("final_report_rec_weak"))
    ws_prob.cell(row=prb + 3, column=1, value=tr("final_report_rec_math"))


def write_excellent_sheet(
    wb: Workbook,
    school_id: int,
    anchor_year: int,
    active_names: set[str],
    tr: Callable[[str], str],
) -> None:
    """Отличники: ученики со всеми годовыми оценками «5»."""
    ws_exc = wb.create_sheet(title=tr("final_report_sheet_excellent")[:31])
    write_header_row(ws_exc, 1, [tr("class"), tr("final_report_col_student_name")])
    exr = 2
    for cn in sorted(active_names, key=class_name_sort_key):
        students = class_students_map(school_id, cn, anchor_year)
        for name, grades in sorted(students.items(), key=lambda x: kazakh_sort_key(x[0])):
            vals = [g.get("grade") for g in grades.values() if g.get("grade") is not None]
            if vals and all(g >= 5 for g in vals):
                write_data_row(ws_exc, exr, [cn, name])
                exr += 1
