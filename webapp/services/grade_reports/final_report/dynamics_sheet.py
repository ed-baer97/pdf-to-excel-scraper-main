"""Лист «Динамика численности»: таблица контингента за N лет по ступеням."""

from __future__ import annotations

from typing import Callable

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ...academic_year import format_academic_year
from ..excel.charts import apply_rect_table_borders
from .data import (
    class_counts_by_period,
    dynamics_year_columns,
    section_total,
    stage_breakdown,
    year_has_grade_data,
)
from .styles import DYNAMICS_HEADER_FILL, cell_display


def write_contingent_dynamics_sheet(
    ws,
    school_id: int,
    anchor_year: int,
    years_back: int,
    tr: Callable[[str], str],
) -> None:
    """Первая вкладка: таблица динамики численности за 3 года (как в Word-шаблоне)."""
    year_cols = dynamics_year_columns(anchor_year, years_back)
    last_col = 1 + len(year_cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=tr("final_report_dynamics_title"))
    title_cell.font = Font(bold=True, underline="single", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = [tr("final_report_col_parameters")] + [
        format_academic_year(y) for y in year_cols
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.fill = DYNAMICS_HEADER_FILL
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    sections: list[tuple[str, str, int]] = [
        (tr("final_report_students_begin"), "students", 1),
        (tr("final_report_students_end"), "students", 4),
        (tr("final_report_class_sets_begin"), "classes", 1),
        (tr("final_report_avg_fill"), "avg_fill", 1),
    ]
    stage_keys = [
        ("primary", tr("final_report_stage_primary")),
        ("basic", tr("final_report_stage_basic")),
        ("secondary", tr("final_report_stage_secondary")),
    ]

    row = 3
    for section_label, metric, period_number in sections:
        year_breakdowns: list[dict | None] = []
        for yr in year_cols:
            if not year_has_grade_data(school_id, yr):
                year_breakdowns.append(None)
                continue
            counts = class_counts_by_period(school_id, yr, period_number)
            year_breakdowns.append(stage_breakdown(counts) if counts else None)

        sec_cell = ws.cell(row=row, column=1, value=section_label)
        sec_cell.font = Font(bold=True, size=11)
        sec_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col_idx, bd in enumerate(year_breakdowns, start=2):
            total = section_total(bd, metric) if bd else None
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=cell_display(total, tr=tr, is_avg=(metric == "avg_fill")),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(total, float) and metric == "avg_fill":
                cell.number_format = "0.0"
        row += 1

        for stage_key, stage_label in stage_keys:
            label_cell = ws.cell(row=row, column=1, value=f"■ {stage_label}")
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            for col_idx, bd in enumerate(year_breakdowns, start=2):
                stage = (bd or {}).get(stage_key) if bd else None
                val = stage.get(metric) if stage else None
                cell = ws.cell(
                    row=row,
                    column=col_idx,
                    value=cell_display(val, tr=tr, is_avg=(metric == "avg_fill")),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, float) and metric == "avg_fill":
                    cell.number_format = "0.0"
            row += 1

    last_data_row = row - 1
    apply_rect_table_borders(ws, 1, last_data_row, 1, last_col)

    ws.column_dimensions["A"].width = 42
    for col_i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 16
