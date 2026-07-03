"""Стили ячеек и общие хелперы рендеринга листов/диаграмм итогового отчёта."""

from __future__ import annotations

from typing import Callable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines, Scaling
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..excel.charts import chart_title_large, excel_chart_sheet_name

HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
DYNAMICS_HEADER_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="CED4DA")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def cell_display(val: int | float | None, *, tr: Callable[[str], str], is_avg: bool = False) -> str | int | float:
    if val is None:
        return tr("final_report_no_data")
    return val


def write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_row(ws, row: int, values: list, *, number_cols: set[int] | None = None) -> None:
    number_cols = number_cols or set()
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col in number_cols and isinstance(val, (int, float)):
            cell.number_format = "0.0"


def axis_line_sppr(color: str = "495057", w: int = 19050) -> GraphicalProperties:
    """Жирная контрастная линия оси (как в качественных диаграммах charts.py)."""
    return GraphicalProperties(ln=LineProperties(w=w, solidFill=ColorChoice(srgbClr=color)))


def major_gridlines(color: str = "DEE2E6", w: int = 6350) -> ChartLines:
    return ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(w=w, solidFill=ColorChoice(srgbClr=color)))
    )


def style_chart_axes(chart, *, y_title: str, y_min: float, y_max: float, y_unit: float) -> None:
    """Единое качественное оформление осей: видимые оси, сетка, прозрачная область."""
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.title = y_title
    chart.y_axis.scaling = Scaling(min=y_min, max=y_max)
    chart.y_axis.majorUnit = y_unit
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.spPr = axis_line_sppr()
    chart.y_axis.majorGridlines = major_gridlines()
    chart.x_axis.lblAlgn = "ctr"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.spPr = axis_line_sppr()
    chart.plot_area.spPr = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )


def add_bar_chart(
    wb: Workbook,
    ws_data,
    *,
    title: str,
    hdr_row: int,
    data_row: int,
    last_col: int,
    chart_idx: int,
    y_max: float = 100,
) -> None:
    """Отдельный лист с простой столбчатой диаграммой по одной строке данных."""
    sheet_name = excel_chart_sheet_name(wb, title, chart_idx)
    wch = wb.create_sheet(title=sheet_name)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = chart_title_large(title)
    chart.y_axis.scaling = Scaling(min=0, max=y_max)
    chart.y_axis.majorUnit = 10 if y_max >= 50 else None
    chart.dLbls = DataLabelList(showVal=True, showSerName=False, showCatName=False)
    cats = Reference(ws_data, min_col=2, min_row=hdr_row, max_col=last_col)
    ser = Series(
        Reference(ws_data, min_col=2, min_row=data_row, max_col=last_col),
        title_from_data=True,
    )
    chart.append(ser)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = ColorChoice(srgbClr="0D6EFD")
    wch.add_chart(chart, "A1")
