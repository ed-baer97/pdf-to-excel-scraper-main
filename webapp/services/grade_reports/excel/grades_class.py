"""Excel: сводная таблица оценок класса."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..periods import YEAR_UI_PERIOD
from .styles import create_excel_styles, is_border_percent


def build_grades_class_workbook(
    class_name: str,
    period_name: str,
    period_number: int,
    subjects_list: list,
    students_list: list,
    subject_stats: dict,
) -> tuple[BytesIO, str]:
    """Возвращает (BytesIO xlsx, имя файла для скачивания)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Оценки {class_name}"

    styles = create_excel_styles()
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    total_cols = len(subjects_list) + 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = f"Сводная таблица оценок: {class_name} ({period_name})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    header_row = 3
    headers = ["№", "ФИО ученика"] + subjects_list + [
        "Кол-во 5",
        "Кол-во 4",
        "Кол-во 3",
        "Кол-во 2",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_alignment"]
        cell.border = styles["border"]

    col_5_idx = len(subjects_list) + 3
    col_4_idx = len(subjects_list) + 4
    col_3_idx = len(subjects_list) + 5
    col_2_idx = len(subjects_list) + 6
    for col_idx, fill_key in (
        (col_5_idx, "count_5_fill"),
        (col_4_idx, "count_4_fill"),
        (col_3_idx, "count_3_fill"),
        (col_2_idx, "count_2_fill"),
    ):
        ws.cell(row=header_row, column=col_idx).fill = styles[fill_key]
        ws.cell(row=header_row, column=col_idx).font = Font(bold=True)

    for row_idx, student in enumerate(students_list, header_row + 1):
        ws.cell(row=row_idx, column=1, value=row_idx - header_row).border = styles["border"]
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=2, value=student["name"]).border = styles["border"]

        for col_idx, subject in enumerate(subjects_list, 3):
            grade_info = student["grades"].get(subject, {})
            grade = grade_info.get("grade")
            percent = grade_info.get("percent")
            if grade:
                cell_value = f"{grade} ({percent}%)" if percent else str(grade)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if is_border_percent(percent):
                    cell.fill = styles["border_highlight_fill"]
                    cell.font = Font(bold=True, color="B45309")
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value="—")
            cell.border = styles["border"]
            cell.alignment = center_align

        for col_idx, key, fill_key in (
            (col_5_idx, "count_5", "count_5_fill"),
            (col_4_idx, "count_4", "count_4_fill"),
            (col_3_idx, "count_3", "count_3_fill"),
            (col_2_idx, "count_2", "count_2_fill"),
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=student[key])
            cell.border = styles["border"]
            cell.alignment = center_align
            cell.fill = styles[fill_key]
            cell.font = bold_font

    footer_start = header_row + len(students_list) + 1

    row = footer_start
    ws.cell(row=row, column=2, value='Кол-во «5»').font = bold_font
    for col_idx, subj in enumerate(subjects_list, 3):
        cell = ws.cell(row=row, column=col_idx, value=subject_stats[subj]["count_5"])
        cell.border = styles["border"]
        cell.alignment = center_align
        cell.fill = styles["count_5_fill"]
        cell.font = bold_font
    cell = ws.cell(
        row=row,
        column=col_5_idx,
        value=sum(1 for s in students_list if s["count_4"] == 0),
    )
    cell.border = styles["border"]
    cell.alignment = center_align
    cell.fill = styles["count_5_fill"]
    cell.font = bold_font

    row = footer_start + 1
    ws.cell(row=row, column=2, value='Кол-во «4»').font = bold_font
    for col_idx, subj in enumerate(subjects_list, 3):
        cell = ws.cell(row=row, column=col_idx, value=subject_stats[subj]["count_4"])
        cell.border = styles["border"]
        cell.alignment = center_align
        cell.fill = styles["count_4_fill"]
        cell.font = bold_font
    cell = ws.cell(
        row=row,
        column=col_4_idx,
        value=sum(1 for s in students_list if s["count_3"] == 0),
    )
    cell.border = styles["border"]
    cell.alignment = center_align
    cell.fill = styles["count_4_fill"]
    cell.font = bold_font

    row = footer_start + 2
    ws.cell(row=row, column=2, value='Кол-во «3»').font = bold_font
    for col_idx, subj in enumerate(subjects_list, 3):
        cell = ws.cell(row=row, column=col_idx, value=subject_stats[subj]["count_3"])
        cell.border = styles["border"]
        cell.alignment = center_align
        cell.fill = styles["count_3_fill"]
        cell.font = bold_font
    cell = ws.cell(
        row=row,
        column=col_3_idx,
        value=sum(1 for s in students_list if s["count_2"] == 0),
    )
    cell.border = styles["border"]
    cell.alignment = center_align
    cell.fill = styles["count_3_fill"]
    cell.font = bold_font

    row = footer_start + 3
    ws.cell(row=row, column=2, value='Кол-во «2»').font = bold_font
    for col_idx, subj in enumerate(subjects_list, 3):
        cell = ws.cell(row=row, column=col_idx, value=subject_stats[subj]["count_2"])
        cell.border = styles["border"]
        cell.alignment = center_align
        cell.fill = styles["count_2_fill"]
        cell.font = bold_font
    cell = ws.cell(
        row=row,
        column=col_2_idx,
        value=sum(1 for s in students_list if s["count_2"] > 0),
    )
    cell.border = styles["border"]
    cell.alignment = center_align
    cell.fill = styles["count_2_fill"]
    cell.font = bold_font

    row = footer_start + 4
    ws.cell(row=row, column=2, value="Качество %").font = bold_font
    for col_idx, subj in enumerate(subjects_list, 3):
        cell = ws.cell(
            row=row, column=col_idx, value=f"{subject_stats[subj]['quality_percent']}%"
        )
        cell.border = styles["border"]
        cell.alignment = center_align
        cell.fill = styles["quality_fill"]
        cell.font = bold_font

    row = footer_start + 5
    ws.cell(row=row, column=2, value="Успеваемость %").font = bold_font
    for col_idx, subj in enumerate(subjects_list, 3):
        cell = ws.cell(
            row=row, column=col_idx, value=f"{subject_stats[subj]['success_percent']}%"
        )
        cell.border = styles["border"]
        cell.alignment = center_align
        cell.fill = styles["success_fill"]
        cell.font = bold_font

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    for col in range(3, len(subjects_list) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for col_idx in (col_5_idx, col_4_idx, col_3_idx, col_2_idx):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    period_slug = (
        "учебный_год" if period_number == YEAR_UI_PERIOD else f"{period_number}_четверть"
    )
    filename = f"Оценки_{class_name}_{period_slug}.xlsx"
    return output, filename
