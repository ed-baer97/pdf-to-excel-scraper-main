"""Excel: диаграммы метрик классов."""

from __future__ import annotations

from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.axis import ChartLines, Scaling
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable as ChartDataTable
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..aggregation import chart_series_from_class_totals
from ..periods import YEAR_UI_PERIOD, class_name_sort_key, parse_class_grade


def avg_pct_metrics(values: list) -> float | None:
    """Среднее по числам в строке метрик (как «Итого» на странице диаграмм)."""
    nums: list[float] = []
    for v in values:
        try:
            if v is not None:
                nums.append(float(v))
        except (TypeError, ValueError):
            pass
    if not nums:
        return None
    return round(sum(nums) / len(nums), 1)


def excel_chart_sheet_name(wb: Workbook, title: str, index: int) -> str:
    """Имя листа для гистограммы: ≤31 символа, без \\/*?:[], уникально в книге."""
    bad = "\\/*?:[]"
    raw = "".join("_" if c in bad else c for c in (title or "").strip()) or f"График_{index + 1}"
    for n in range(50):
        suffix = "" if n == 0 else f" ({n})"
        base = raw if n == 0 else raw[: max(1, 31 - len(suffix))].rstrip() + suffix
        candidate = base[:31]
        if candidate and candidate not in wb.sheetnames:
            return candidate
    return f"Гр_{index + 1}"[:31]


def apply_rect_table_borders(
    ws,
    top_row: int,
    bottom_row: int,
    left_col: int,
    right_col: int,
) -> None:
    """Рамка таблицы: снаружи жирнее, внутри «перегородки» между строками/столбцами."""
    v_in = Side(style="thin", color="9CA3AF")
    h_mid = Side(style="medium", color="6C757D")
    out = Side(style="medium", color="495057")
    for rr in range(top_row, bottom_row + 1):
        for cc in range(left_col, right_col + 1):
            left = out if cc == left_col else v_in
            right = out if cc == right_col else v_in
            top = out if rr == top_row else h_mid
            bottom = out if rr == bottom_row else h_mid
            ws.cell(row=rr, column=cc).border = Border(left=left, right=right, top=top, bottom=bottom)


def chart_title_large(text: str, sz: int = 2200):
    """Заголовок диаграммы Excel с увеличенным шрифтом (sz — размер в 1/100 pt, OOXML)."""
    from openpyxl.chart.text import Text
    from openpyxl.chart.title import Title
    from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RegularTextRun

    rpr = CharacterProperties(sz=sz, b=True)
    paras = []
    for line in text.split("\n"):
        run = RegularTextRun(t=line)
        run.rPr = rpr
        pp = ParagraphProperties()
        pp.defRPr = CharacterProperties(sz=sz, b=True)
        paras.append(Paragraph(r=[run], pPr=pp))
    t = Title()
    t.tx = Text()
    t.tx.rich.paragraphs = paras
    return t


def build_class_metrics_charts_workbook(
    *,
    scope: str,
    period_number: int,
    agg: dict,
    tr: Callable[[str], str],
) -> tuple[BytesIO, str, str]:
    """Return (xlsx BytesIO, ascii filename, utf-8 filename)."""
    labels, quality_values, success_values = chart_series_from_class_totals(agg["class_totals"])
    class_totals = agg["class_totals"]

    PAGE_WIDE_COL = 8
    thin_grid = Side(style="thin", color="CED4DA")
    grid_border = Border(left=thin_grid, right=thin_grid, top=thin_grid, bottom=thin_grid)
    table_header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    card_header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    card_header_font = Font(bold=True, color="FFFFFF", size=13)
    section_label_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    section_label_font = Font(bold=True, size=11, color="0D6EFD")
    sheet_band_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    fill_q = PatternFill(start_color="ECF4FF", end_color="ECF4FF", fill_type="solid")
    fill_s = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    label_font = Font(bold=True, size=10, color="212529")

    rows_data: list[dict] = []
    for i, name in enumerate(labels):
        rows_data.append(
            {
                "name": name,
                "grade": parse_class_grade(name),
                "q": quality_values[i],
                "s": success_values[i],
                "w": int(class_totals.get(name, {}).get("weight_total") or 0),
            }
        )
    rows_data.sort(key=lambda r: class_name_sort_key(r["name"]))

    def filt(pred):
        return [r for r in rows_data if r["grade"] is not None and pred(r["grade"])]

    overall_sections = [
        (tr("metrics_excel_sec_1_4"), lambda g: 1 <= g <= 4, tr("metrics_excel_empty_1_4")),
        (tr("metrics_excel_sec_5_9"), lambda g: 5 <= g <= 9, tr("metrics_excel_empty_5_9")),
        (tr("metrics_excel_sec_9_11"), lambda g: 9 <= g <= 11, tr("metrics_excel_empty_9_11")),
        (tr("metrics_excel_sec_9"), lambda g: g == 9, tr("metrics_excel_empty_9")),
        (tr("metrics_excel_sec_11"), lambda g: g == 11, tr("metrics_excel_empty_11")),
    ]

    grades_parallel = sorted({r["grade"] for r in rows_data if r["grade"] is not None})

    if scope == "overall":
        note_scope = tr("metrics_excel_note_overall")
    else:
        note_scope = tr("metrics_excel_note_parallel")

    wb = Workbook()
    ws = wb.active
    ws.title = tr("metrics_excel_sheet_tables")
    ws.sheet_view.showGridLines = False
    chart_specs: list[dict] = []

    if period_number == YEAR_UI_PERIOD:
        period_label = tr("metrics_excel_period_year")
    elif 1 <= period_number <= 4:
        period_label = tr(f"metrics_excel_period_q{period_number}")
    else:
        period_label = tr("metrics_excel_period_fallback").format(n=period_number)
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=PAGE_WIDE_COL)
    t1 = ws.cell(row=r, column=1, value=tr("metrics_excel_title"))
    t1.font = Font(bold=True, size=17, color="212529")
    t1.fill = sheet_band_fill
    t1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    tper = ws.cell(row=r, column=6, value=period_label)
    tper.font = Font(size=11, color="495057")
    tper.fill = sheet_band_fill
    tper.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[r].height = 32
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=PAGE_WIDE_COL)
    st = ws.cell(row=r, column=1, value=tr("metrics_excel_subtitle"))
    st.font = Font(size=11, color="6C757D")
    st.fill = sheet_band_fill
    st.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[r].height = 22
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=PAGE_WIDE_COL)
    tnote = ws.cell(row=r, column=1, value=note_scope)
    tnote.font = Font(size=9, color="6C757D")
    tnote.fill = sheet_band_fill
    tnote.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
    ws.row_dimensions[r].height = 34
    r += 1
    r += 1
    apply_rect_table_borders(ws, 1, 3, 1, PAGE_WIDE_COL)

    def write_table_block(title: str, subset: list[dict], empty_msg: str) -> None:
        nonlocal r
        n = len(subset)
        if not subset:
            empty_w = 4
            r_head = r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=empty_w)
            h_empty = ws.cell(row=r, column=1, value=title)
            h_empty.font = card_header_font
            h_empty.fill = card_header_fill
            h_empty.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[r].height = 22
            r += 1
            r_msg = r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=empty_w)
            em = ws.cell(row=r, column=1, value=empty_msg)
            em.font = Font(italic=True, size=10, color="6C757D")
            em.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            em.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
            ws.row_dimensions[r].height = 22
            r += 1
            apply_rect_table_borders(ws, r_head, r_msg, 1, empty_w)
            r += 1
            r += 1
            return

        last_tot_col = 2 + n
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_tot_col)
        h = ws.cell(row=r, column=1, value=title)
        h.font = card_header_font
        h.fill = card_header_fill
        h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[r].height = 24
        r += 1

        qvals = [rec["q"] for rec in subset]
        svals = [rec["s"] for rec in subset]
        q_tot = avg_pct_metrics(qvals)
        s_tot = avg_pct_metrics(svals)

        hdr_row = r
        c0 = ws.cell(row=hdr_row, column=1, value=tr("metrics_col_indicator"))
        c0.font = header_font_white
        c0.fill = table_header_fill
        c0.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, rec in enumerate(subset, start=2):
            chc = ws.cell(row=hdr_row, column=i, value=rec["name"])
            chc.font = header_font_white
            chc.fill = table_header_fill
            chc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ct = ws.cell(row=hdr_row, column=last_tot_col, value=tr("metrics_col_total"))
        ct.font = header_font_white
        ct.fill = table_header_fill
        ct.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[hdr_row].height = 20

        row_q = hdr_row + 1
        ws.cell(row=row_q, column=1, value=tr("metrics_row_quality")).font = label_font
        ws.cell(row=row_q, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.cell(row=row_q, column=1).fill = fill_q
        for ci, v in enumerate(qvals, start=2):
            cell = ws.cell(row=row_q, column=ci, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill_q
            if isinstance(v, (int, float)):
                cell.number_format = "0.0"
        cqt = ws.cell(
            row=row_q,
            column=last_tot_col,
            value=q_tot if q_tot is not None else "—",
        )
        cqt.alignment = Alignment(horizontal="center", vertical="center")
        cqt.font = Font(bold=True, size=10)
        cqt.fill = fill_q
        if q_tot is not None:
            cqt.number_format = "0.0"

        row_s = hdr_row + 2
        ws.cell(row=row_s, column=1, value=tr("metrics_row_success")).font = label_font
        ws.cell(row=row_s, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.cell(row=row_s, column=1).fill = fill_s
        for ci, v in enumerate(svals, start=2):
            cell = ws.cell(row=row_s, column=ci, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill_s
            if isinstance(v, (int, float)):
                cell.number_format = "0.0"
        cst = ws.cell(
            row=row_s,
            column=last_tot_col,
            value=s_tot if s_tot is not None else "—",
        )
        cst.alignment = Alignment(horizontal="center", vertical="center")
        cst.font = Font(bold=True, size=10)
        cst.fill = fill_s
        if s_tot is not None:
            cst.number_format = "0.0"
        ws.row_dimensions[row_q].height = 19
        ws.row_dimensions[row_s].height = 19

        apply_rect_table_borders(ws, hdr_row, row_s, 1, last_tot_col)

        chart_specs.append(
            {
                "title": title,
                "hdr_row": hdr_row,
                "row_q": row_q,
                "row_s": row_s,
                "last_tot_col": last_tot_col,
                "n": n,
            }
        )
        r = row_s + 1
        r += 2

    def append_histogram_sheets() -> None:
        for idx, spec in enumerate(chart_specs):
            title = spec["title"]
            hdr_row = spec["hdr_row"]
            row_q = spec["row_q"]
            row_s = spec["row_s"]
            last_tot_col = spec["last_tot_col"]
            n = spec["n"]
            wsheet = excel_chart_sheet_name(wb, title, idx)
            wch = wb.create_sheet(title=wsheet)
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.varyColors = False
            chart.style = 2
            chart.title = chart_title_large(title)
            chart.title.overlay = False
            # У openpyxl по умолчанию у catAx и valAx axPos="l" — для столбчатой диаграммы нужно X снизу, Y слева.
            chart.x_axis.axPos = "b"
            chart.y_axis.axPos = "l"
            # Явно «не удалять» оси — в Excel галочка «Оси» в элементах диаграммы (delete val="0" в XML).
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            chart.y_axis.title = "%"
            chart.y_axis.scaling = Scaling(min=0, max=100)
            chart.y_axis.majorUnit = 10
            chart.y_axis.tickLblPos = "nextTo"
            chart.y_axis.spPr = GraphicalProperties(
                ln=LineProperties(
                    w=19050,
                    solidFill=ColorChoice(srgbClr="495057"),
                )
            )
            chart.y_axis.majorGridlines = ChartLines(
                spPr=GraphicalProperties(
                    ln=LineProperties(
                        w=6350,
                        solidFill=ColorChoice(srgbClr="DEE2E6"),
                    )
                )
            )
            chart.plot_area.spPr = GraphicalProperties(
                noFill=True,
                ln=LineProperties(noFill=True),
            )
            chart.x_axis.lblAlgn = "ctr"
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.spPr = GraphicalProperties(
                ln=LineProperties(
                    w=19050,
                    solidFill=ColorChoice(srgbClr="495057"),
                )
            )
            chart.gapWidth = 78
            # Явно только значение — иначе Excel показывает «ряд; категория; значение» и всё слипается.
            chart.dLbls = DataLabelList(
                showVal=True,
                showSerName=False,
                showCatName=False,
                showLegendKey=False,
                dLblPos="inEnd",
            )
            chart.width = min(36.0, max(15.0, 5.5 + (n + 1) * 1.28))
            if n <= 2:
                chart.height = 12.0
            elif n <= 6:
                chart.height = 13.0
            elif n <= 12:
                chart.height = 14.5
            else:
                chart.height = 16.0

            cats = Reference(ws, min_col=2, min_row=hdr_row, max_col=last_tot_col)
            ser_q = Series(
                Reference(ws, min_col=2, min_row=row_q, max_col=last_tot_col),
                title=tr("metrics_row_quality"),
            )
            ser_s = Series(
                Reference(ws, min_col=2, min_row=row_s, max_col=last_tot_col),
                title=tr("metrics_row_success"),
            )
            chart.append(ser_q)
            chart.append(ser_s)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.solidFill = ColorChoice(srgbClr="0D6EFD")
            chart.series[1].graphicalProperties.solidFill = ColorChoice(srgbClr="198754")
            chart.plot_area.dTable = ChartDataTable(
                showHorzBorder=True,
                showVertBorder=True,
                showOutline=True,
                showKeys=True,
            )
            chart.legend = None
            wch.add_chart(chart, "A1")

    if scope == "overall":
        row_ob = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=PAGE_WIDE_COL)
        tab_o = ws.cell(row=r, column=1, value=tr("metrics_excel_section_overall"))
        tab_o.font = section_label_font
        tab_o.fill = section_label_fill
        tab_o.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[r].height = 24
        r += 1
        r += 1
        apply_rect_table_borders(ws, row_ob, row_ob, 1, PAGE_WIDE_COL)

        first_block_row = r

        for title, pred, empty_msg in overall_sections:
            write_table_block(title, filt(pred), empty_msg)

        row_sv = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=PAGE_WIDE_COL)
        sv = ws.cell(row=r, column=1, value=tr("metrics_excel_section_summary"))
        sv.font = section_label_font
        sv.fill = section_label_fill
        sv.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[r].height = 24
        r += 1
        apply_rect_table_borders(ws, row_sv, row_sv, 1, PAGE_WIDE_COL)

        sum_top = r
        sq, ss = agg.get("school_quality"), agg.get("school_success")
        for col, val in enumerate(
            (tr("metrics_col_indicator"), tr("metrics_row_quality"), tr("metrics_row_success")),
            start=1,
        ):
            c = ws.cell(row=r, column=col, value=val)
            c.font = header_font_white
            c.fill = table_header_fill
            c.border = grid_border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 19
        r += 1

        c_school = ws.cell(row=r, column=1, value=tr("metrics_excel_school_weighted"))
        c_school.border = grid_border
        c_school.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c_school.font = Font(size=10, color="212529")
        for col, val in ((2, sq), (3, ss)):
            cell = ws.cell(row=r, column=col, value=val if val is not None else "—")
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)
            if isinstance(val, (int, float)):
                cell.number_format = "0.0"
        ws.row_dimensions[r].height = 18
        r += 1

        parallel = agg.get("parallel") or {}
        for pkey, title in (
            ("1-4", tr("metrics_excel_parallel_1_4")),
            ("5-9", tr("metrics_excel_parallel_5_9")),
            ("10-11", tr("metrics_excel_parallel_10_11")),
        ):
            pr = parallel.get(pkey) or {}
            pq, ps = pr.get("quality"), pr.get("success")
            c1 = ws.cell(row=r, column=1, value=title)
            c1.border = grid_border
            c1.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            c1.font = Font(size=10, color="212529")
            for col, val in ((2, pq), (3, ps)):
                cell = ws.cell(row=r, column=col, value=val if val is not None else "—")
                cell.border = grid_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=10)
                if isinstance(val, (int, float)):
                    cell.number_format = "0.0"
            ws.row_dimensions[r].height = 18
            r += 1
        apply_rect_table_borders(ws, sum_top, r - 1, 1, 3)
    else:
        row_tp = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=PAGE_WIDE_COL)
        tab_p = ws.cell(row=r, column=1, value=tr("metrics_excel_section_parallel"))
        tab_p.font = section_label_font
        tab_p.fill = section_label_fill
        tab_p.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[r].height = 24
        r += 1
        r += 1
        apply_rect_table_borders(ws, row_tp, row_tp, 1, PAGE_WIDE_COL)

        first_block_row = r

        for g in grades_parallel:
            sub = [x for x in rows_data if x["grade"] == g]
            write_table_block(
                f"{g} {tr('metrics_class_word')}",
                sub,
                tr("metrics_parallel_empty").format(grade=g),
            )

    for spec in chart_specs:
        for col_i in range(2, spec["last_tot_col"] + 1):
            letter = get_column_letter(col_i)
            prev = ws.column_dimensions[letter].width
            ws.column_dimensions[letter].width = max(float(prev or 10), 11.0)

    append_histogram_sheets()

    ws.freeze_panes = f"A{first_block_row}"
    ws.column_dimensions["A"].width = 26
    for col_i in range(2, 15):
        letter = get_column_letter(col_i)
        prev = ws.column_dimensions[letter].width
        ws.column_dimensions[letter].width = max(float(prev or 10), 11.5)

    suffix_key = "metrics_excel_fn_suf_o" if scope == "overall" else "metrics_excel_fn_suf_p"
    filename_local = tr("metrics_excel_fn_pattern").format(period=period_number, suffix=tr(suffix_key))
    suffix_ascii = "overall" if scope == "overall" else "parallel"
    filename_ascii = f"stat_classes_{period_number}ch_{suffix_ascii}.xlsx"
    wb.properties.title = filename_local.replace(".xlsx", "")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename_ascii, filename_local
