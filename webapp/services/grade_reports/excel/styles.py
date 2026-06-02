"""Общие стили openpyxl для админских выгрузок."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_excel_styles() -> dict:
    """Создание стилей для Excel."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    grade_fills = {
        5: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        4: PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        3: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        2: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "header_alignment": header_alignment,
        "border": border,
        "grade_fills": grade_fills,
        "border_highlight_fill": PatternFill(
            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
        ),
        "count_5_fill": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "count_4_fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "count_3_fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "count_2_fill": PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid"),
        "quality_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "success_fill": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    }


def is_border_percent(pct) -> bool:
    """Проверка: пограничный процент (37-39%, 61-64%, 82-84%)."""
    if pct is None:
        return False
    return (37 <= pct <= 39) or (61 <= pct <= 64) or (82 <= pct <= 84)
