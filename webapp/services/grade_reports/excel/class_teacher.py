"""Excel: отчёт классных руководителей по категориям."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .styles import create_excel_styles


def build_class_teacher_workbook(
    categories_data: dict,
    period_name: str,
) -> tuple[BytesIO, str]:
    """Возвращает (BytesIO xlsx, имя файла для скачивания)."""
    wb = Workbook()
    styles = create_excel_styles()

    cat_meta = [
        ("excellent", "на 5", "C6EFCE", ["Класс", "№", "ФИО", "Классный руководитель"]),
        ("good", "на 4", "BDD7EE", ["Класс", "№", "ФИО", "Классный руководитель"]),
        (
            "one_4",
            "С одной 4",
            "D9EAD3",
            ["Класс", "№", "ФИО", "Предмет", "Учитель", "Классный руководитель"],
        ),
        (
            "satisfactory",
            "на 3",
            "FFEB9C",
            [
                "Класс",
                "ФИО",
                "Предмет 1",
                "Предмет 2",
                "Предмет 3",
                "Предмет 4",
                "Предмет 5+",
                "Классный руководитель",
            ],
        ),
        (
            "one_3",
            "С одной 3",
            "FBE5D6",
            ["Класс", "№", "ФИО", "Предмет", "Учитель", "Классный руководитель"],
        ),
        (
            "poor",
            "Неуспевающие",
            "FFC7CE",
            ["Класс", "№", "ФИО", "Предмет", "Учитель", "Классный руководитель"],
        ),
    ]

    first_sheet = True
    for cat_key, cat_label, cat_color, headers in cat_meta:
        blocks = categories_data[cat_key]
        total_count = sum(len(b["students"]) for b in blocks)
        sheet_name = f"{cat_label} ({total_count})"[:31]

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(
            row=1,
            column=1,
            value=f"Отчёт классных руководителей — {cat_label} ({period_name})",
        )
        c.font = Font(bold=True, size=13)
        c.alignment = Alignment(horizontal="center")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = styles["header_font"]
            cell.fill = PatternFill(
                start_color=cat_color, end_color=cat_color, fill_type="solid"
            )
            cell.alignment = styles["header_alignment"]
            cell.border = styles["border"]

        row = 4
        if not blocks:
            ws.cell(row=row, column=1, value="Нет данных")
            continue

        for block in blocks:
            cls = block["class_name"]
            ct = block["class_teacher"]

            if cat_key == "satisfactory":
                details = block.get("troechniki_detailed", [])
                n = len(details)
                if n == 0:
                    continue
                ws.merge_cells(start_row=row, start_column=1, end_row=row + n - 1, end_column=1)
                ws.cell(row=row, column=1, value=cls).font = Font(bold=True)
                ws.cell(row=row, column=1).border = styles["border"]
                ws.cell(row=row, column=1).alignment = Alignment(
                    vertical="center", horizontal="center"
                )
                ws.merge_cells(start_row=row, start_column=8, end_row=row + n - 1, end_column=8)
                ws.cell(row=row, column=8, value=ct).border = styles["border"]
                ws.cell(row=row, column=8).alignment = Alignment(vertical="center")

                for item in details:
                    ws.cell(row=row, column=2, value=item["student"]).border = styles["border"]
                    for i in range(4):
                        val = ""
                        if i < len(item["subjects_1_4"]):
                            s = item["subjects_1_4"][i]
                            val = f"{s['subject_name']} ({s['grade']})"
                        ws.cell(row=row, column=3 + i, value=val or "—").border = styles["border"]
                    val5 = ", ".join(
                        f"{s['subject_name']} ({s['grade']})"
                        for s in item.get("subjects_5", [])
                    )
                    ws.cell(row=row, column=7, value=val5 or "—").border = styles["border"]
                    row += 1

            elif cat_key in ("one_4", "one_3", "poor"):
                students = block["students"]
                n = len(students)
                ws.merge_cells(start_row=row, start_column=1, end_row=row + n - 1, end_column=1)
                ws.cell(row=row, column=1, value=cls).font = Font(bold=True)
                ws.cell(row=row, column=1).border = styles["border"]
                ws.cell(row=row, column=1).alignment = Alignment(
                    vertical="center", horizontal="center"
                )
                ws.merge_cells(start_row=row, start_column=6, end_row=row + n - 1, end_column=6)
                ws.cell(row=row, column=6, value=ct).border = styles["border"]
                ws.cell(row=row, column=6).alignment = Alignment(vertical="center")

                for idx, item in enumerate(students, 1):
                    ws.cell(row=row, column=2, value=idx).border = styles["border"]
                    ws.cell(row=row, column=3, value=item["student"]).border = styles["border"]
                    ws.cell(row=row, column=4, value=item["subject"]).border = styles["border"]
                    ws.cell(row=row, column=5, value=item["teacher"]).border = styles["border"]
                    row += 1

            else:
                students = block["students"]
                n = len(students)
                ws.merge_cells(start_row=row, start_column=1, end_row=row + n - 1, end_column=1)
                ws.cell(row=row, column=1, value=cls).font = Font(bold=True)
                ws.cell(row=row, column=1).border = styles["border"]
                ws.cell(row=row, column=1).alignment = Alignment(
                    vertical="center", horizontal="center"
                )
                ws.merge_cells(start_row=row, start_column=4, end_row=row + n - 1, end_column=4)
                ws.cell(row=row, column=4, value=ct).border = styles["border"]
                ws.cell(row=row, column=4).alignment = Alignment(vertical="center")

                for idx, student in enumerate(students, 1):
                    ws.cell(row=row, column=2, value=idx).border = styles["border"]
                    ws.cell(row=row, column=3, value=student).border = styles["border"]
                    row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                12, len(headers[col_idx - 1]) + 5
            )
        ws.column_dimensions["C"].width = 35

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Отчёт_классных_руководителей_{period_name.replace(' ', '_')}.xlsx"
    return output, filename
