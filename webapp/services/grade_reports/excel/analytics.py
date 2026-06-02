"""Excel: аналитика СОР / СОЧ / оценки."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ....constants import kazakh_sort_key
from .styles import create_excel_styles


def build_analytics_workbook(
    period_name: str,
    subjects_data_sor: dict,
    subjects_data_soch: dict,
    subjects_data_grades: dict,
) -> BytesIO:
    """Собирает xlsx с тремя листами: СОР, СОЧ, Оценки."""
    styles = create_excel_styles()
    wb = Workbook()

    def _write_sor_sheet() -> None:
        ws = wb.active
        ws.title = "СОР"[:31]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws["A1"] = f"Аналитика СОР ({period_name})"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        row = 3
        for subj, data_list in sorted(
            subjects_data_sor.items(), key=lambda item: kazakh_sort_key(item[0])
        ):
            ws.cell(row=row, column=1, value=subj).font = Font(bold=True, size=12)
            row += 1
            headers = [
                "Класс",
                "СОР",
                "5",
                "4",
                "3",
                "2",
                "Всего",
                "Качество %",
                "Успеваемость %",
                "Учитель",
            ]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = styles["header_font"]
                c.fill = styles["header_fill"]
                c.border = styles["border"]
            row += 1
            for item in data_list:
                if item["sor_list"]:
                    for sor in item["sor_list"]:
                        ws.cell(row=row, column=1, value=item["class_name"]).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=2, value=sor.get("name", "-")).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=3, value=sor.get("count_5", 0)).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=4, value=sor.get("count_4", 0)).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=5, value=sor.get("count_3", 0)).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=6, value=sor.get("count_2", 0)).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=7, value=sor.get("total", 0)).border = (
                            styles["border"]
                        )
                        ws.cell(row=row, column=8, value=sor.get("quality") or "-").border = (
                            styles["border"]
                        )
                        ws.cell(
                            row=row, column=9, value=sor.get("success_rate") or "-"
                        ).border = styles["border"]
                        ws.cell(row=row, column=10, value=item["teacher"] or "-").border = (
                            styles["border"]
                        )
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=item["class_name"]).border = (
                        styles["border"]
                    )
                    for col in range(2, 10):
                        ws.cell(row=row, column=col, value="-").border = styles["border"]
                    ws.cell(row=row, column=10, value=item["teacher"] or "-").border = (
                        styles["border"]
                    )
                    row += 1
            row += 2

    def _write_soch_sheet() -> None:
        ws = wb.create_sheet(title="СОЧ"[:31])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws["A1"] = f"Аналитика СОЧ ({period_name})"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        row = 3
        for subj, data_list in sorted(
            subjects_data_soch.items(), key=lambda item: kazakh_sort_key(item[0])
        ):
            ws.cell(row=row, column=1, value=subj).font = Font(bold=True, size=12)
            row += 1
            headers = [
                "Класс",
                "5",
                "4",
                "3",
                "2",
                "Всего",
                "Качество %",
                "Успеваемость %",
                "Учитель",
            ]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = styles["header_font"]
                c.fill = styles["header_fill"]
                c.border = styles["border"]
            row += 1
            for item in data_list:
                ws.cell(row=row, column=1, value=item["class_name"]).border = styles["border"]
                ws.cell(row=row, column=2, value=item["count_5"]).border = styles["border"]
                ws.cell(row=row, column=3, value=item["count_4"]).border = styles["border"]
                ws.cell(row=row, column=4, value=item["count_3"]).border = styles["border"]
                ws.cell(row=row, column=5, value=item["count_2"]).border = styles["border"]
                ws.cell(row=row, column=6, value=item["total"]).border = styles["border"]
                ws.cell(row=row, column=7, value=item["quality"] or "-").border = styles["border"]
                ws.cell(row=row, column=8, value=item["success_rate"] or "-").border = (
                    styles["border"]
                )
                ws.cell(row=row, column=9, value=item["teacher"] or "-").border = styles["border"]
                row += 1
            row += 2

    def _write_grades_sheet() -> None:
        ws = wb.create_sheet(title="Оценки"[:31])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws["A1"] = f"Аналитика оценок ({period_name})"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        row = 3
        for subj, data_list in sorted(
            subjects_data_grades.items(), key=lambda item: kazakh_sort_key(item[0])
        ):
            ws.cell(row=row, column=1, value=subj).font = Font(bold=True, size=12)
            row += 1
            headers = [
                "Класс",
                "5",
                "4",
                "3",
                "2",
                "Всего",
                "Качество %",
                "Успеваемость %",
                "Учитель",
            ]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = styles["header_font"]
                c.fill = styles["header_fill"]
                c.border = styles["border"]
            row += 1
            for item in data_list:
                ws.cell(row=row, column=1, value=item["class_name"]).border = styles["border"]
                ws.cell(row=row, column=2, value=item["count_5"]).border = styles["border"]
                ws.cell(row=row, column=3, value=item["count_4"]).border = styles["border"]
                ws.cell(row=row, column=4, value=item["count_3"]).border = styles["border"]
                ws.cell(row=row, column=5, value=item["count_2"]).border = styles["border"]
                ws.cell(row=row, column=6, value=item["total"]).border = styles["border"]
                ws.cell(row=row, column=7, value=item["quality"] or "-").border = styles["border"]
                ws.cell(row=row, column=8, value=item["success_rate"] or "-").border = (
                    styles["border"]
                )
                ws.cell(row=row, column=9, value=item["teacher"] or "-").border = styles["border"]
                row += 1
            row += 2

    _write_sor_sheet()
    _write_soch_sheet()
    _write_grades_sheet()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
