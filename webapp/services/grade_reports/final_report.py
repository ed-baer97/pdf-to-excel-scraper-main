"""Сборка школьного итогового отчёта в Excel из БД оценок и ручных данных."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines, Scaling
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ...constants import kazakh_sort_key, normalize_subject_name
from ...extensions import db
from ...models import Class, GradeReport, School
from ..academic_year import available_academic_years, format_academic_year, resolve_academic_year
from ..admin_dashboard import aggregate_class_metrics
from ..year_grades import YEAR_UI_PERIOD
from .excel.charts import apply_rect_table_borders, chart_title_large, excel_chart_sheet_name
from .final_report_data import load_all_sections, load_sections_for_years
from .payload import report_grades_payload
from .periods import class_accordion_group, class_name_sort_key, parse_class_grade
from .queries import get_period_reports

_HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
_DYNAMICS_HEADER_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="CED4DA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _resolve_years(school_id: int, academic_year: int | None, years_back: int = 3) -> list[int]:
    """Годы с данными (для имени файла и прочих листов)."""
    all_years = available_academic_years(school_id)
    anchor = resolve_academic_year(academic_year)
    if anchor in all_years:
        idx = all_years.index(anchor)
        picked = all_years[idx : idx + years_back]
    else:
        picked = [anchor]
    return sorted(picked)


def _dynamics_year_columns(anchor_year: int, years_back: int = 3) -> list[int]:
    """Ровно N колонок учебных лет подряд, заканчивая anchor_year (для таблицы динамики)."""
    n = max(1, int(years_back))
    start = anchor_year - n + 1
    return list(range(start, anchor_year + 1))


def _year_has_grade_data(school_id: int, academic_year: int) -> bool:
    return (
        GradeReport.query.filter_by(school_id=school_id, academic_year=academic_year)
        .limit(1)
        .first()
        is not None
    )


def _class_counts_by_period(
    school_id: int,
    academic_year: int,
    period_number: int,
) -> dict[str, int]:
    """Численность по классам за период (макс. total_students по предметам класса)."""
    reports = get_period_reports(
        school_id, period_number, academic_year=academic_year
    )
    by_class: dict[str, int] = {}
    for report in reports:
        payload = report_grades_payload(report)
        if not payload:
            continue
        students = payload.get("students") or []
        total = int(payload.get("total_students") or len(students) or 0)
        if total <= 0:
            continue
        by_class[report.class_name] = max(by_class.get(report.class_name, 0), total)
    return by_class


def _stage_breakdown(class_counts: dict[str, int]) -> dict[str, dict[str, int | float] | None]:
    """Разбивка по ступеням: начальная (1–4), основная (5–9), средняя (10–11)."""
    buckets: dict[str, list[int]] = {"primary": [], "basic": [], "secondary": []}
    bucket_map = {"1-4": "primary", "5-9": "basic", "10-11": "secondary"}
    for class_name, count in class_counts.items():
        key = bucket_map.get(class_accordion_group(class_name))
        if key:
            buckets[key].append(count)

    def _stage(items: list[int]) -> dict[str, int | float] | None:
        if not items:
            return None
        return {
            "students": sum(items),
            "classes": len(items),
            "avg_fill": round(sum(items) / len(items), 1),
        }

    return {
        "primary": _stage(buckets["primary"]),
        "basic": _stage(buckets["basic"]),
        "secondary": _stage(buckets["secondary"]),
    }


def _section_total(
    breakdown: dict[str, dict[str, int | float] | None],
    metric: str,
) -> int | float | None:
    stages = [s for s in breakdown.values() if s]
    if not stages:
        return None
    if metric == "students":
        return sum(int(s["students"]) for s in stages)
    if metric == "classes":
        return sum(int(s["classes"]) for s in stages)
    if metric == "avg_fill":
        total_students = sum(int(s["students"]) for s in stages)
        total_classes = sum(int(s["classes"]) for s in stages)
        return round(total_students / total_classes, 1) if total_classes else None
    return None


def _cell_display(val: int | float | None, *, tr: Callable[[str], str], is_avg: bool = False) -> str | int | float:
    if val is None:
        return tr("final_report_no_data")
    return val


def _write_contingent_dynamics_sheet(
    ws,
    school_id: int,
    anchor_year: int,
    years_back: int,
    tr: Callable[[str], str],
) -> None:
    """Первая вкладка: таблица динамики численности за 3 года (как в Word-шаблоне)."""
    year_cols = _dynamics_year_columns(anchor_year, years_back)
    last_col = 1 + len(year_cols)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=tr("final_report_dynamics_title"))
    title_cell.font = Font(bold=True, underline="single", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = [tr("final_report_col_parameters")] + [
        format_academic_year(y) for y in year_cols
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.fill = _DYNAMICS_HEADER_FILL
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    sections: list[tuple[str, str, int]] = [
        (tr("final_report_students_begin"), "students", 1),
        (tr("final_report_students_end"), "students", 4),
        (tr("final_report_class_sets_begin"), "classes", 1),
        (tr("final_report_avg_fill"), "avg_fill", 1),
    ]
    stage_keys = [
        ("primary", tr("final_report_stage_primary")),
        ("basic", tr("final_report_stage_basic")),
        ("secondary", tr("final_report_stage_secondary")),
    ]

    row = 3
    for section_label, metric, period_number in sections:
        year_breakdowns: list[dict | None] = []
        for yr in year_cols:
            if not _year_has_grade_data(school_id, yr):
                year_breakdowns.append(None)
                continue
            counts = _class_counts_by_period(school_id, yr, period_number)
            year_breakdowns.append(_stage_breakdown(counts) if counts else None)

        sec_cell = ws.cell(row=row, column=1, value=section_label)
        sec_cell.font = Font(bold=True, size=11)
        sec_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col_idx, bd in enumerate(year_breakdowns, start=2):
            total = _section_total(bd, metric) if bd else None
            cell = ws.cell(
                row=row,
                column=col_idx,
                value=_cell_display(total, tr=tr, is_avg=(metric == "avg_fill")),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(total, float) and metric == "avg_fill":
                cell.number_format = "0.0"
        row += 1

        for stage_key, stage_label in stage_keys:
            label_cell = ws.cell(row=row, column=1, value=f"■ {stage_label}")
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            for col_idx, bd in enumerate(year_breakdowns, start=2):
                stage = (bd or {}).get(stage_key) if bd else None
                val = stage.get(metric) if stage else None
                cell = ws.cell(
                    row=row,
                    column=col_idx,
                    value=_cell_display(val, tr=tr, is_avg=(metric == "avg_fill")),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, float) and metric == "avg_fill":
                    cell.number_format = "0.0"
            row += 1

    last_data_row = row - 1
    apply_rect_table_borders(ws, 1, last_data_row, 1, last_col)

    ws.column_dimensions["A"].width = 42
    for col_i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 16


def _active_class_names(school_id: int) -> set[str]:
    return {row.name for row in Class.query.filter_by(school_id=school_id).all()}


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_data_row(ws, row: int, values: list, *, number_cols: set[int] | None = None) -> None:
    number_cols = number_cols or set()
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col in number_cols and isinstance(val, (int, float)):
            cell.number_format = "0.0"


def _class_students_map(
    school_id: int,
    class_name: str,
    academic_year: int,
    period_number: int = YEAR_UI_PERIOD,
) -> dict[str, dict[str, dict]]:
    """Ученики класса: предмет → оценка за выбранный период."""
    reports = get_period_reports(
        school_id,
        period_number,
        class_name=class_name,
        academic_year=academic_year,
    )
    students: dict[str, dict[str, dict]] = {}
    for report in reports:
        subj = normalize_subject_name(report.subject_name, school_id)
        grades_data = report_grades_payload(report)
        if not grades_data:
            continue
        for student in grades_data.get("students", []) or []:
            name = (student.get("name") or "").strip()
            if not name:
                continue
            if name not in students:
                students[name] = {}
            existing = students[name].get(subj)
            new_grade = {
                "percent": student.get("percent"),
                "grade": student.get("grade"),
            }
            if existing is None or existing.get("grade") is None:
                students[name][subj] = new_grade
            elif new_grade.get("grade") is not None and new_grade["grade"] > (
                existing.get("grade") or 0
            ):
                students[name][subj] = new_grade
    return students


def _class_grade_summary(
    school_id: int,
    class_name: str,
    academic_year: int,
    period_number: int = YEAR_UI_PERIOD,
) -> dict[str, Any]:
    """Сводка по классу: кол-во на 5/4/3/2, успеваемость, качество."""
    students = _class_students_map(school_id, class_name, academic_year, period_number)
    total = len(students)
    if total == 0:
        return {
            "class_name": class_name,
            "total": 0,
            "passing": 0,
            "count_5": 0,
            "count_4": 0,
            "one_3": 0,
            "two_plus_3": 0,
            "count_3": 0,
            "count_2": 0,
            "success_percent": None,
            "quality_percent": None,
        }

    count_5 = count_4 = one_3 = two_plus_3 = count_3 = count_2 = 0
    for grades in students.values():
        vals = [g.get("grade") for g in grades.values() if g.get("grade") is not None]
        if not vals:
            continue
        c3 = sum(1 for g in vals if g == 3)
        c2 = sum(1 for g in vals if g is not None and g <= 2)
        if c2 > 0:
            count_2 += 1
        elif all(g >= 5 for g in vals):
            count_5 += 1
        elif c3 == 0 and 4 in vals:
            count_4 += 1
        elif c3 == 1:
            one_3 += 1
        elif c3 >= 2:
            two_plus_3 += 1
            count_3 += 1
        else:
            count_4 += 1

    passing = total - count_2
    quality = round((count_5 + count_4) / total * 100, 1) if total else None
    success = round(passing / total * 100, 1) if total else None
    return {
        "class_name": class_name,
        "total": total,
        "passing": passing,
        "count_5": count_5,
        "count_4": count_4,
        "one_3": one_3,
        "two_plus_3": two_plus_3,
        "count_3": count_3,
        "count_2": count_2,
        "success_percent": success,
        "quality_percent": quality,
    }


def _parallel_summary(class_summaries: list[dict], bucket: str) -> dict[str, Any]:
    subset = [
        s
        for s in class_summaries
        if class_accordion_group(s["class_name"]) == bucket and s["total"] > 0
    ]
    if not subset:
        return {"total": 0, "quality_percent": None, "success_percent": None}
    total = sum(s["total"] for s in subset)
    q_vals = [s["quality_percent"] for s in subset if s["quality_percent"] is not None]
    s_vals = [s["success_percent"] for s in subset if s["success_percent"] is not None]
    return {
        "total": total,
        "quality_percent": round(sum(q_vals) / len(q_vals), 1) if q_vals else None,
        "success_percent": round(sum(s_vals) / len(s_vals), 1) if s_vals else None,
    }


_GRADE_ROW_KEYS = (
    "final_report_grade_row_5",
    "final_report_grade_row_4",
    "final_report_grade_row_3",
    "final_report_grade_row_2",
)
_GRADE_CHART_COLORS = ("70AD47", "4472C4", "ED7D31", "C00000")
_QUALITY_PERIODS = (1, 2, 3, 4, YEAR_UI_PERIOD)


def _school_grade_distribution_2_11(
    school_id: int,
    academic_year: int,
    period_number: int,
    active_names: set[str],
) -> dict[str, Any]:
    """Сводка по школе (классы 2–11): распределение на 5/4/3/2 и проценты."""
    count_5 = count_4 = one_3 = count_3 = count_2 = 0
    for class_name in active_names:
        grade_num = parse_class_grade(class_name)
        if grade_num is None or grade_num < 2 or grade_num > 11:
            continue
        summary = _class_grade_summary(
            school_id, class_name, academic_year, period_number
        )
        if summary["total"] <= 0:
            continue
        count_5 += summary["count_5"]
        count_4 += summary["count_4"]
        one_3 += summary["one_3"]
        count_3 += summary["count_3"]
        count_2 += summary["count_2"]
    total = count_5 + count_4 + one_3 + count_3 + count_2
    on_3 = one_3 + count_3
    return {
        "count_5": count_5,
        "count_4": count_4,
        "count_3": on_3,
        "count_2": count_2,
        "total": total,
        "quality_percent": round((count_5 + count_4) / total * 100, 1) if total else None,
        "success_percent": round((total - count_2) / total * 100, 1) if total else None,
    }


def _period_header_labels(tr: Callable[[str], str]) -> list[str]:
    return [
        tr("final_report_hdr_q1"),
        tr("final_report_hdr_q2"),
        tr("final_report_hdr_q3"),
        tr("final_report_hdr_q4"),
        tr("final_report_hdr_total"),
    ]


def _write_quality_stages_sheet(
    ws,
    school_id: int,
    academic_year: int,
    active_names: set[str],
    tr: Callable[[str], str],
) -> None:
    """Вкладка: показатели качества знаний (таблица + диаграммы 2–11 классы)."""
    period_labels = _period_header_labels(tr)
    last_col = 1 + len(period_labels)
    distributions = [
        _school_grade_distribution_2_11(school_id, academic_year, pn, active_names)
        for pn in _QUALITY_PERIODS
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=tr("final_report_quality_title"))
    title_cell.font = Font(bold=True, underline="single", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    headers = [tr("final_report_col_indicator")] + period_labels
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.fill = _DYNAMICS_HEADER_FILL
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    metric_rows: list[tuple[str, str]] = [
        (key, field)
        for key, field in zip(
            _GRADE_ROW_KEYS,
            ("count_5", "count_4", "count_3", "count_2"),
            strict=True,
        )
    ]
    metric_rows += [
        ("metrics_row_success", "success_percent"),
        ("metrics_row_quality", "quality_percent"),
    ]

    row = 3
    for label_key, field in metric_rows:
        label_cell = ws.cell(row=row, column=1, value=tr(label_key))
        label_cell.font = Font(bold=field.endswith("percent"), size=11)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, dist in enumerate(distributions, start=2):
            # Пустой период (нет данных) -> ячейка остаётся пустой, чтобы график
            # не проваливался в ноль и на столбцах не висели нулевые подписи.
            val = dist.get(field) if dist.get("total") else None
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(val, float):
                cell.number_format = "0.0"
        row += 1

    last_data_row = row - 1
    apply_rect_table_borders(ws, 1, last_data_row, 1, last_col)

    ws.column_dimensions["A"].width = 22
    for col_i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    _add_quality_stages_charts(ws, tr, last_col=last_col, last_data_row=last_data_row)


def _axis_line_sppr(color: str = "495057", w: int = 19050) -> GraphicalProperties:
    """Жирная контрастная линия оси (как в качественных диаграммах charts.py)."""
    return GraphicalProperties(ln=LineProperties(w=w, solidFill=ColorChoice(srgbClr=color)))


def _major_gridlines(color: str = "DEE2E6", w: int = 6350) -> ChartLines:
    return ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(w=w, solidFill=ColorChoice(srgbClr=color)))
    )


def _style_chart_axes(chart, *, y_title: str, y_min: float, y_max: float, y_unit: float) -> None:
    """Единое качественное оформление осей: видимые оси, сетка, прозрачная область."""
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.title = y_title
    chart.y_axis.scaling = Scaling(min=y_min, max=y_max)
    chart.y_axis.majorUnit = y_unit
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.spPr = _axis_line_sppr()
    chart.y_axis.majorGridlines = _major_gridlines()
    chart.x_axis.lblAlgn = "ctr"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.spPr = _axis_line_sppr()
    chart.plot_area.spPr = GraphicalProperties(
        noFill=True, ln=LineProperties(noFill=True)
    )


def _add_quality_stages_charts(
    ws,
    tr: Callable[[str], str],
    *,
    last_col: int,
    last_data_row: int,
) -> None:
    """Столбчатая (накопительная) и линейная диаграммы под таблицей."""
    hdr_row = 2
    grade_first_row = 3
    grade_last_row = 6
    pct_first_row = 7
    pct_last_row = 8
    chart_anchor_row = last_data_row + 3

    cats = Reference(ws, min_col=2, min_row=hdr_row, max_col=last_col, max_row=hdr_row)

    y_max = 0
    for col in range(2, last_col + 1):
        col_sum = sum(
            int(ws.cell(row=r, column=col).value or 0)
            for r in range(grade_first_row, grade_last_row + 1)
            if isinstance(ws.cell(row=r, column=col).value, (int, float))
        )
        y_max = max(y_max, col_sum)
    bar_y_max = max(100, int((y_max * 1.15) // 100 + 1) * 100)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.varyColors = False
    bar.style = 2
    bar.title = chart_title_large(tr("final_report_chart_grades_2_11"))
    bar.title.overlay = False
    bar.gapWidth = 60
    _style_chart_axes(
        bar,
        y_title=tr("final_report_chart_y_students"),
        y_min=0,
        y_max=bar_y_max,
        y_unit=200 if bar_y_max > 1000 else 100,
    )
    bar.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        showLegendKey=False,
        dLblPos="ctr",
        numFmt="0;-0;;",
    )
    bar.dLbls.numFmt = "0;-0;;"
    bar.dLbls.sourceLinked = False
    bar.display_blanks = "gap"
    bar.width = 24
    bar.height = 14

    for row_idx, color in zip(
        range(grade_first_row, grade_last_row + 1),
        _GRADE_CHART_COLORS,
        strict=True,
    ):
        data = Reference(
            ws, min_col=2, min_row=row_idx, max_col=last_col, max_row=row_idx
        )
        ser = Series(data, title=str(ws.cell(row=row_idx, column=1).value))
        ser.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
        ser.graphicalProperties.line.noFill = True
        bar.series.append(ser)
    bar.set_categories(cats)
    bar.legend.position = "r"
    bar.legend.overlay = False
    ws.add_chart(bar, f"A{chart_anchor_row}")

    line_anchor_row = chart_anchor_row + 30
    line = LineChart()
    line.style = 2
    line.title = chart_title_large(tr("final_report_chart_quality_2_11"))
    line.title.overlay = False
    _style_chart_axes(
        line,
        y_title=tr("final_report_chart_y_percent"),
        y_min=50,
        y_max=100,
        y_unit=10,
    )
    line.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        showLegendKey=False,
        dLblPos="t",
    )
    line.display_blanks = "span"
    line.width = 24
    line.height = 14

    line_colors = ("7030A0", "00B0B9")
    markers = ("circle", "square")
    for row_idx, color, symbol in zip(
        range(pct_first_row, pct_last_row + 1),
        line_colors,
        markers,
        strict=True,
    ):
        data = Reference(
            ws, min_col=2, min_row=row_idx, max_col=last_col, max_row=row_idx
        )
        ser = Series(data, title=str(ws.cell(row=row_idx, column=1).value))
        ser.smooth = False
        ser.graphicalProperties.line.solidFill = ColorChoice(srgbClr=color)
        ser.graphicalProperties.line.width = 28575
        ser.marker = Marker(symbol=symbol, size=7)
        ser.marker.graphicalProperties.solidFill = ColorChoice(srgbClr=color)
        ser.marker.graphicalProperties.line.solidFill = ColorChoice(srgbClr=color)
        line.series.append(ser)
    line.set_categories(cats)
    line.legend.position = "r"
    line.legend.overlay = False
    ws.add_chart(line, f"A{line_anchor_row}")


def _subject_quality_matrix(
    school_id: int,
    academic_year: int,
    active_names: set[str],
) -> list[dict[str, Any]]:
    """Качество по предметам и классам."""
    reports = get_period_reports(school_id, YEAR_UI_PERIOD, academic_year=academic_year)
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for report in reports:
        if report.class_name not in active_names:
            continue
        subj = normalize_subject_name(report.subject_name, school_id)
        key = (report.class_name, subj)
        if key in seen:
            continue
        seen.add(key)
        grades_data = report_grades_payload(report)
        if not grades_data:
            continue
        qp = grades_data.get("quality_percent")
        sp = grades_data.get("success_percent")
        if qp is None:
            students = grades_data.get("students", []) or []
            s5 = s4 = s3 = s2 = 0
            for st in students:
                g = st.get("grade")
                if g == 5:
                    s5 += 1
                elif g == 4:
                    s4 += 1
                elif g == 3:
                    s3 += 1
                elif g is not None and g <= 2:
                    s2 += 1
            denom = s5 + s4 + s3 + s2
            qp = round((s5 + s4) / denom * 100, 1) if denom else None
            sp = round((s5 + s4 + s3) / denom * 100, 1) if denom else None
        rows.append(
            {
                "class_name": report.class_name,
                "subject": subj,
                "grade": parse_class_grade(report.class_name),
                "quality_percent": qp,
                "success_percent": sp,
            }
        )
    rows.sort(key=lambda r: (r["subject"], class_name_sort_key(r["class_name"])))
    return rows


def _lowest_quality(rows: list[dict], limit: int = 10) -> list[dict]:
    valid = [r for r in rows if r.get("quality_percent") is not None]
    valid.sort(key=lambda r: r["quality_percent"])
    return valid[:limit]


def _add_bar_chart(
    wb: Workbook,
    ws_data,
    *,
    title: str,
    hdr_row: int,
    data_row: int,
    last_col: int,
    chart_idx: int,
    y_max: float = 100,
) -> None:
    sheet_name = excel_chart_sheet_name(wb, title, chart_idx)
    wch = wb.create_sheet(title=sheet_name)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = chart_title_large(title)
    chart.y_axis.scaling = Scaling(min=0, max=y_max)
    chart.y_axis.majorUnit = 10 if y_max >= 50 else None
    chart.dLbls = DataLabelList(showVal=True, showSerName=False, showCatName=False)
    cats = Reference(ws_data, min_col=2, min_row=hdr_row, max_col=last_col)
    ser = Series(
        Reference(ws_data, min_col=2, min_row=data_row, max_col=last_col),
        title_from_data=True,
    )
    chart.append(ser)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = ColorChoice(srgbClr="0D6EFD")
    wch.add_chart(chart, "A1")


def build_final_report_workbook(
    school_id: int,
    *,
    academic_year: int | None = None,
    years_back: int = 3,
    tr: Callable[[str], str],
) -> tuple[BytesIO, str]:
    """Собрать многолистовой Excel итогового отчёта."""
    school = db.session.get(School, school_id)
    school_name = school.name if school else f"School {school_id}"
    anchor_year = resolve_academic_year(academic_year)
    years = _resolve_years(school_id, anchor_year, years_back)
    active_names = _active_class_names(school_id)
    manual_by_year = load_sections_for_years(school_id, years)

    wb = Workbook()
    year_cols = _dynamics_year_columns(anchor_year, years_back)

    # --- Вкладка 1: динамика численности за 3 года ---
    ws_dyn = wb.active
    ws_dyn.title = tr("final_report_sheet_dynamics")[:31]
    ws_dyn.sheet_view.showGridLines = True
    _write_contingent_dynamics_sheet(ws_dyn, school_id, anchor_year, years_back, tr)

    # --- Вкладка 2: показатели качества знаний (2–11 классы) ---
    ws_qual = wb.create_sheet(title=tr("final_report_sheet_quality")[:31], index=1)
    ws_qual.sheet_view.showGridLines = True
    _write_quality_stages_sheet(ws_qual, school_id, anchor_year, active_names, tr)

    # --- Сводка KPI ---
    ws_sum = wb.create_sheet(title=tr("final_report_sheet_summary")[:31])
    ws_sum.cell(row=1, column=1, value=tr("final_report_title")).font = Font(bold=True, size=16)
    ws_sum.cell(row=2, column=1, value=school_name)
    ws_sum.cell(row=3, column=1, value=format_academic_year(anchor_year))
    row = 5
    year_agg = aggregate_class_metrics(
        school_id, YEAR_UI_PERIOD, active_names, academic_year=anchor_year
    )
    _write_header_row(ws_sum, row, [tr("metrics_col_indicator"), tr("metrics_row_quality"), tr("metrics_row_success")])
    row += 1
    _write_data_row(
        ws_sum,
        row,
        [
            tr("metrics_excel_school_weighted"),
            year_agg.get("school_quality"),
            year_agg.get("school_success"),
        ],
        number_cols={2, 3},
    )
    row += 2
    for key, label_key in (
        ("1-4", "metrics_excel_parallel_1_4"),
        ("5-9", "metrics_excel_parallel_5_9"),
        ("10-11", "metrics_excel_parallel_10_11"),
    ):
        pr = (year_agg.get("parallel") or {}).get(key) or {}
        _write_data_row(
            ws_sum,
            row,
            [tr(label_key), pr.get("quality"), pr.get("success")],
            number_cols={2, 3},
        )
        row += 1

    # --- Успеваемость по классам ---
    class_summaries = [
        _class_grade_summary(school_id, cn, anchor_year)
        for cn in sorted(active_names, key=class_name_sort_key)
    ]
    ws_cls = wb.create_sheet(title=tr("final_report_sheet_classes")[:31])
    cls_headers = [
        tr("class"),
        tr("final_report_col_students"),
        tr("final_report_col_on_5"),
        tr("final_report_col_on_4"),
        tr("final_report_col_one_3"),
        tr("final_report_col_two_3"),
        tr("final_report_col_on_3"),
        tr("final_report_col_on_2"),
        tr("metrics_row_success"),
        tr("metrics_row_quality"),
    ]
    _write_header_row(ws_cls, 1, cls_headers)
    cr = 2
    for s in class_summaries:
        if s["total"] <= 0:
            continue
        _write_data_row(
            ws_cls,
            cr,
            [
                s["class_name"],
                s["total"],
                s["count_5"],
                s["count_4"],
                s["one_3"],
                s["two_plus_3"],
                s["count_3"],
                s["count_2"],
                s["success_percent"],
                s["quality_percent"],
            ],
            number_cols={9, 10},
        )
        cr += 1

    # --- По ступеням ---
    ws_par = wb.create_sheet(title=tr("final_report_sheet_parallels")[:31])
    _write_header_row(
        ws_par,
        1,
        [
            tr("final_report_col_stage"),
            tr("final_report_col_students"),
            tr("metrics_row_quality"),
            tr("metrics_row_success"),
        ],
    )
    pr = 2
    for bucket, label_key in (
        ("1-4", "metrics_excel_sec_1_4"),
        ("5-9", "metrics_excel_sec_5_9"),
        ("10-11", "metrics_excel_parallel_10_11"),
    ):
        ps = _parallel_summary(class_summaries, bucket)
        _write_data_row(
            ws_par,
            pr,
            [tr(label_key), ps["total"], ps["quality_percent"], ps["success_percent"]],
            number_cols={3, 4},
        )
        pr += 1

    # --- По четвертям ---
    ws_q = wb.create_sheet(title=tr("final_report_sheet_quarters")[:31])
    _write_header_row(
        ws_q,
        1,
        [
            tr("quarter_label"),
            tr("metrics_row_quality"),
            tr("metrics_row_success"),
            tr("final_report_col_classes_with_data"),
        ],
    )
    qr = 2
    quarter_chart_hdr = 20
    ws_q.cell(row=quarter_chart_hdr, column=1, value=tr("quarter_label"))
    for qi, qn in enumerate([1, 2, 3, 4, YEAR_UI_PERIOD], start=2):
        label = (
            tr(f"metrics_excel_period_q{qn}")
            if qn <= 4
            else tr("metrics_excel_period_year")
        )
        ws_q.cell(row=quarter_chart_hdr, column=qi, value=label)
    ws_q.cell(row=quarter_chart_hdr + 1, column=1, value=tr("metrics_row_quality"))
    ws_q.cell(row=quarter_chart_hdr + 2, column=1, value=tr("metrics_row_success"))
    for qi, qn in enumerate([1, 2, 3, 4, YEAR_UI_PERIOD], start=2):
        agg = aggregate_class_metrics(
            school_id, qn, active_names, academic_year=anchor_year
        )
        label = (
            tr(f"metrics_excel_period_q{qn}")
            if qn <= 4
            else tr("metrics_excel_period_year")
        )
        _write_data_row(
            ws_q,
            qr,
            [label, agg.get("school_quality"), agg.get("school_success"), agg.get("classes_with_data")],
            number_cols={2, 3},
        )
        ws_q.cell(row=quarter_chart_hdr + 1, column=qi, value=agg.get("school_quality"))
        ws_q.cell(row=quarter_chart_hdr + 2, column=qi, value=agg.get("school_success"))
        qr += 1
    _add_bar_chart(
        wb,
        ws_q,
        title=tr("final_report_chart_quarters"),
        hdr_row=quarter_chart_hdr,
        data_row=quarter_chart_hdr + 1,
        last_col=6,
        chart_idx=0,
    )

    # --- Качество по предметам ---
    subj_rows = _subject_quality_matrix(school_id, anchor_year, active_names)
    ws_subj = wb.create_sheet(title=tr("final_report_sheet_subjects")[:31])
    _write_header_row(
        ws_subj,
        1,
        [tr("subject"), tr("class"), tr("metrics_row_quality"), tr("metrics_row_success")],
    )
    sr = 2
    for r in subj_rows:
        _write_data_row(
            ws_subj,
            sr,
            [r["subject"], r["class_name"], r["quality_percent"], r["success_percent"]],
            number_cols={3, 4},
        )
        sr += 1

    # --- Проблемные зоны ---
    ws_prob = wb.create_sheet(title=tr("final_report_sheet_problems")[:31])
    _write_header_row(
        ws_prob,
        1,
        [tr("class"), tr("subject"), tr("metrics_row_quality")],
    )
    prb = 2
    for r in _lowest_quality(subj_rows):
        _write_data_row(
            ws_prob,
            prb,
            [r["class_name"], r["subject"], r["quality_percent"]],
            number_cols={3},
        )
        prb += 1
    ws_prob.cell(row=prb + 1, column=1, value=tr("final_report_recommendations"))
    ws_prob.cell(row=prb + 2, column=1, value=tr("final_report_rec_weak"))
    ws_prob.cell(row=prb + 3, column=1, value=tr("final_report_rec_math"))

    # --- Отличники ---
    ws_exc = wb.create_sheet(title=tr("final_report_sheet_excellent")[:31])
    _write_header_row(ws_exc, 1, [tr("class"), tr("final_report_col_student_name")])
    exr = 2
    for cn in sorted(active_names, key=class_name_sort_key):
        students = _class_students_map(school_id, cn, anchor_year)
        for name, grades in sorted(students.items(), key=lambda x: kazakh_sort_key(x[0])):
            vals = [g.get("grade") for g in grades.values() if g.get("grade") is not None]
            if vals and all(g >= 5 for g in vals):
                _write_data_row(ws_exc, exr, [cn, name])
                exr += 1

    # --- ГИА / ЕНТ / Аттестаты (ручной ввод) ---
    manual = manual_by_year.get(anchor_year, load_all_sections(school_id, anchor_year))

    def _write_json_sheet(title_key: str, data: Any) -> None:
        ws = wb.create_sheet(title=tr(title_key)[:31])
        ws.cell(row=1, column=1, value=tr(title_key)).font = Font(bold=True, size=14)
        if isinstance(data, dict):
            row_i = 3
            for k, v in data.items():
                if isinstance(v, (list, dict)):
                    ws.cell(row=row_i, column=1, value=str(k))
                    ws.cell(row=row_i, column=2, value=json.dumps(v, ensure_ascii=False, indent=2))
                    row_i += 2
                else:
                    ws.cell(row=row_i, column=1, value=str(k))
                    ws.cell(row=row_i, column=2, value=v)
                    row_i += 1
        else:
            ws.cell(row=3, column=1, value=json.dumps(data, ensure_ascii=False, indent=2))

    _write_json_sheet("final_report_sheet_gia9", manual.get("gia9", {}))
    _write_json_sheet("final_report_sheet_gia11", manual.get("gia11", {}))

    # ЕНТ с диаграммой
    ent = manual.get("ent", {})
    ws_ent = wb.create_sheet(title=tr("final_report_sheet_ent")[:31])
    ws_ent.cell(row=1, column=1, value=tr("final_report_sheet_ent")).font = Font(bold=True, size=14)
    _write_header_row(
        ws_ent,
        3,
        [
            tr("final_report_ent_col_period"),
            tr("final_report_ent_col_count"),
            tr("final_report_ent_col_avg"),
            tr("final_report_ent_col_max"),
        ],
    )
    ent_row = 4
    periods = ent.get("periods") or []
    chart_hdr = 3
    for i, p in enumerate(periods):
        _write_data_row(
            ws_ent,
            ent_row,
            [
                p.get("month") or p.get("period") or "",
                p.get("count"),
                p.get("avg_score"),
                p.get("max_score"),
            ],
            number_cols={3},
        )
        ent_row += 1
    if len(periods) >= 2:
        for i, p in enumerate(periods, start=2):
            ws_ent.cell(row=chart_hdr, column=i, value=p.get("month") or p.get("period") or "")
        ws_ent.cell(row=chart_hdr + 1, column=1, value=tr("final_report_ent_col_avg"))
        for i, p in enumerate(periods, start=2):
            ws_ent.cell(row=chart_hdr + 1, column=i, value=p.get("avg_score"))
        _add_bar_chart(
            wb,
            ws_ent,
            title=tr("final_report_chart_ent"),
            hdr_row=chart_hdr,
            data_row=chart_hdr + 1,
            last_col=1 + len(periods),
            chart_idx=1,
            y_max=140,
        )
    if ent.get("forecast_avg") is not None:
        ws_ent.cell(row=ent_row + 1, column=1, value=tr("final_report_ent_forecast"))
        ws_ent.cell(row=ent_row + 1, column=2, value=ent.get("forecast_avg"))
    if ent.get("recommendations"):
        ws_ent.cell(row=ent_row + 3, column=1, value=tr("final_report_recommendations"))
        ws_ent.cell(row=ent_row + 4, column=1, value=ent.get("recommendations"))

    # Аттестаты
    awards = manual.get("awards", {})
    ws_aw = wb.create_sheet(title=tr("final_report_sheet_awards")[:31])
    ws_aw.cell(row=1, column=1, value=tr("final_report_sheet_awards")).font = Font(bold=True, size=14)
    _write_header_row(
        ws_aw,
        3,
        [
            tr("final_report_awards_altyn"),
            tr("final_report_awards_excellent_11"),
            tr("final_report_awards_excellent_9"),
        ],
    )
    _write_data_row(
        ws_aw,
        4,
        [
            awards.get("altyn_belgi", 0),
            awards.get("excellent_11", 0),
            awards.get("excellent_9", 0),
        ],
    )
    students_aw = awards.get("students") or []
    if students_aw:
        _write_header_row(ws_aw, 6, [tr("final_report_col_student_name"), tr("final_report_col_award_type")])
        ar = 7
        for st in students_aw:
            if isinstance(st, dict):
                _write_data_row(ws_aw, ar, [st.get("name", ""), st.get("award", "")])
            else:
                _write_data_row(ws_aw, ar, [str(st), ""])
            ar += 1

    # Auto column widths
    for ws in wb.worksheets:
        for col in range(1, min(ws.max_column + 1, 12)):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 18

    years_label = "_".join(str(y) for y in year_cols)
    filename = f"Итоговый_отчёт_{years_label}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename
