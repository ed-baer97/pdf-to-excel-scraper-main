"""Smoke tests for final school report Excel builder."""

from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

from webapp import create_app
from webapp.config import TestingConfig
from webapp.extensions import db
from webapp.models import Class, FinalReportData, GradeReport, Role, School, User
from webapp.services.grade_reports.final_report import build_final_report_workbook
from webapp.services.grade_reports.final_report_data import (
    load_all_sections,
    save_section_data,
)
from webapp.translator import gettext


@pytest.fixture
def app():
    application = create_app(TestingConfig)
    with application.app_context():
        db.create_all()
        yield application
        db.session.remove()
        db.drop_all()


def _seed_school(app) -> dict:
    with app.app_context():
        school = School(name="Final Report School")
        db.session.add(school)
        db.session.flush()
        teacher = User(
            username="final_teacher",
            full_name="Final Teacher",
            role=Role.TEACHER.value,
            school_id=school.id,
        )
        teacher.set_password("pass")
        db.session.add(teacher)
        db.session.flush()
        db.session.add(Class(school_id=school.id, name="5А"))
        grades = {
            "students": [
                {"name": "Алиев А.", "grade": 5, "percent": 100},
                {"name": "Беков Б.", "grade": 4, "percent": 85},
            ],
            "total_students": 2,
            "quality_percent": 100.0,
            "success_percent": 100.0,
        }
        db.session.add(
            GradeReport(
                teacher_id=teacher.id,
                school_id=school.id,
                class_name="5А",
                subject_name="Математика",
                period_type="quarter",
                period_number=1,
                academic_year=2025,
                grades_json=json.dumps(grades, ensure_ascii=False),
            )
        )
        db.session.add(
            GradeReport(
                teacher_id=teacher.id,
                school_id=school.id,
                class_name="5А",
                subject_name="Математика",
                period_type="quarter",
                period_number=2,
                academic_year=2025,
                grades_json=json.dumps(grades, ensure_ascii=False),
            )
        )
        db.session.add(
            GradeReport(
                teacher_id=teacher.id,
                school_id=school.id,
                class_name="5А",
                subject_name="Математика",
                period_type="quarter",
                period_number=4,
                academic_year=2025,
                grades_json=json.dumps(grades, ensure_ascii=False),
            )
        )
        db.session.commit()
        save_section_data(
            school.id,
            2025,
            "awards",
            {
                "altyn_belgi": 1,
                "excellent_11": 2,
                "excellent_9": 3,
                "students": [{"name": "Алиев А.", "award": "altyn_belgi"}],
            },
        )
        return {"school_id": school.id}


def test_build_final_report_workbook_smoke(app):
    ctx = _seed_school(app)
    with app.app_context():
        buf, filename = build_final_report_workbook(
            ctx["school_id"],
            academic_year=2025,
            years_back=3,
            tr=lambda k: gettext(k, "ru"),
        )
        assert filename.endswith(".xlsx")
        wb = load_workbook(buf)
        sheet_names = wb.sheetnames
        assert sheet_names[0] == "Динамика численности"
        ws = wb["Динамика численности"]
        assert "ДИНАМИКА ЧИСЛЕННОСТИ" in str(ws.cell(1, 1).value)
        assert ws.cell(2, 2).value == "2023–2024"
        assert ws.cell(2, 4).value == "2025–2026"
        assert ws.cell(3, 2).value == "-"
        assert ws.cell(3, 4).value == 2
        assert sheet_names[1] == "Качество по ступеням"
        ws_q = wb["Качество по ступеням"]
        assert "ПОКАЗАТЕЛИ КАЧЕСТВА" in str(ws_q.cell(1, 1).value)
        assert ws_q.cell(2, 2).value == "1 четверть"
        assert ws_q.cell(3, 2).value == 1
        assert ws_q.cell(3, 5).value == 1
        assert len(ws_q._charts) == 2
        assert any("Сводка" in n for n in sheet_names)
        assert any("Аттестаты" in n for n in sheet_names)
        wb.close()


def test_final_report_data_roundtrip(app):
    ctx = _seed_school(app)
    with app.app_context():
        row = FinalReportData.query.filter_by(
            school_id=ctx["school_id"], section="awards"
        ).first()
        assert row is not None
        data = json.loads(row.data_json)
        assert data.get("altyn_belgi") == 1


def test_legacy_manual_sections_remain_readable_and_exported(app):
    ctx = _seed_school(app)
    legacy_data = {
        "gia9": {"classes": [{"name": "9А", "students": 20}], "notes": "Архив"},
        "gia11": {"classes": [{"name": "11А", "students": 15}], "notes": ""},
        "ent": {
            "periods": [
                {"month": "Январь", "count": 10, "avg_score": 80.0, "max_score": 120},
                {"month": "Май", "count": 10, "avg_score": 93.0, "max_score": 130},
            ],
            "forecast_avg": 96.7,
            "recommendations": "Продолжить подготовку",
        },
    }

    with app.app_context():
        for section, data in legacy_data.items():
            db.session.add(
                FinalReportData(
                    school_id=ctx["school_id"],
                    academic_year=2025,
                    section=section,
                    data_json=json.dumps(data, ensure_ascii=False),
                )
            )
        db.session.commit()

        loaded = load_all_sections(ctx["school_id"], 2025)
        assert loaded["gia9"]["notes"] == "Архив"
        assert loaded["gia11"]["classes"][0]["name"] == "11А"
        assert loaded["ent"]["forecast_avg"] == 96.7

        # Legacy data is preserved for compatibility, not re-enabled for editing.
        with pytest.raises(ValueError, match="invalid section"):
            save_section_data(ctx["school_id"], 2025, "ent", {"forecast_avg": 100})

        buf, _ = build_final_report_workbook(
            ctx["school_id"],
            academic_year=2025,
            years_back=3,
            tr=lambda key: gettext(key, "ru"),
        )
        wb = load_workbook(buf)
        assert "ГИА-9" in wb.sheetnames
        assert "ГИА-11" in wb.sheetnames
        assert "ЕНТ" in wb.sheetnames
        assert wb["ЕНТ"].cell(row=4, column=3).value == 80.0
        wb.close()
