"""Характеризующие тесты criteria_grades: разбор payload, таблицы, Excel/ZIP."""

from __future__ import annotations

import zipfile

from openpyxl import load_workbook

from webapp.services.criteria_grades import (
    build_criteria_period_zip,
    build_criteria_subject_summary,
    build_criteria_table,
    build_final_table,
    build_simple_grades_table,
    build_subjects_workbook,
    criteria_period_path_slug,
    format_score_with_max,
    grade_distribution,
    has_criteria_data,
    has_final_data,
    is_final_period,
    is_year_period,
    ordered_criteria_sections,
    parse_points_by_section,
    safe_path_segment,
    table_for_period_payload,
)


def _criteria_block() -> dict:
    return {
        "quarter_num": 2,
        "sections": [1, 2],
        "max_points": {"1": 20, "2": 15, "0": 30},
        "students": [
            {
                "num": 2,
                "fio": "Беков Б.",
                "average": "8",
                "points": {
                    "chetvert_2_razdel_1_0": "15",
                    "chetvert_2_razdel_2_0": "10",
                    "chetvert_2_razdel_0_0": "25",
                },
                "total_pct": "83%",
                "grade": "4",
            },
            {
                "num": 1,
                "fio": "Алиев А.",
                "average": "9",
                "points": {
                    "chetvert_2_razdel_1_0": "18",
                    "chetvert_2_razdel_0_0": "28",
                },
                "total_pct": "92%",
                "grade": "5",
            },
        ],
    }


class TestParsePointsBySection:
    def test_groups_by_section(self):
        pts = {
            "chetvert_2_razdel_1_0": "15",
            "chetvert_2_razdel_2_1": "10",
            "chetvert_2_razdel_0_0": "25",
        }
        assert parse_points_by_section(pts, 2) == {1: "15", 2: "10", 0: "25"}

    def test_ignores_other_quarters_and_garbage(self):
        pts = {
            "chetvert_1_razdel_1_0": "5",
            "chetvert_2_razdel_x_0": "7",
            "not_a_key": "1",
        }
        assert parse_points_by_section(pts, 2) == {}

    def test_empty_input(self):
        assert parse_points_by_section({}, 2) == {}
        assert parse_points_by_section(None, 2) == {}


class TestOrderedCriteriaSections:
    def test_sor_then_soch(self):
        sections = ordered_criteria_sections(_criteria_block())
        assert sections == [(1, "СОр 1"), (2, "СОр 2"), (0, "СОЧ")]

    def test_sections_inferred_from_student_points(self):
        criteria = {
            "quarter_num": 1,
            "students": [
                {"num": 1, "fio": "X", "points": {"chetvert_1_razdel_3_0": "5"}},
            ],
        }
        assert ordered_criteria_sections(criteria) == [(3, "СОр 3")]


class TestFormatScoreWithMax:
    def test_appends_max(self):
        assert format_score_with_max("17", 1, {"1": 20}) == "17/20"

    def test_percent_passthrough(self):
        assert format_score_with_max("85%", 1, {"1": 20}) == "85%"

    def test_no_max_known(self):
        assert format_score_with_max("17", 5, {"1": 20}) == "17"

    def test_empty(self):
        assert format_score_with_max("", 1, {"1": 20}) == ""
        assert format_score_with_max(None, 1, {"1": 20}) == ""


class TestBuildCriteriaTable:
    def test_headers_and_sorted_rows(self):
        table = build_criteria_table(_criteria_block())
        assert table["headers"] == [
            "№", "ФИО", "ФО",
            "СОр 1 (макс. 20)", "СОр 2 (макс. 15)", "СОЧ (макс. 30)",
            "Сумма", "Оценка",
        ]
        rows = table["rows"]
        assert [r["cells"][1] for r in rows] == ["Алиев А.", "Беков Б."]
        assert rows[0]["cells"] == ["1", "Алиев А.", "9", "18/20", "", "28/30", "92%", "5"]
        assert rows[1]["cells"] == ["2", "Беков Б.", "8", "15/20", "10/15", "25/30", "83%", "4"]

    def test_skips_students_without_fio(self):
        criteria = {"quarter_num": 1, "students": [{"num": 1, "fio": " "}]}
        assert build_criteria_table(criteria)["rows"] == []


class TestGradeDistribution:
    def test_counts(self):
        payload = {
            "students": [
                {"grade": 5}, {"grade": "4"}, {"grade": 4.0},
                {"grade": 3}, {"grade": 2}, {"grade": None}, {"grade": ""},
            ]
        }
        assert grade_distribution(payload) == {"5": 1, "4": 2, "3": 1, "2": 1}

    def test_non_dict_payload(self):
        assert grade_distribution(None) == {"5": 0, "4": 0, "3": 0, "2": 0}


class TestBuildCriteriaSubjectSummary:
    def test_uses_payload_percentages(self):
        payload = {
            "students": [{"grade": 5}, {"grade": 3}],
            "total_students": 2,
            "quality_percent": 50.0,
            "success_percent": 100.0,
        }
        summary = build_criteria_subject_summary(payload)
        assert summary["total_students"] == 2
        assert summary["with_grade"] == 2
        assert summary["quality_percent"] == 50.0
        assert summary["success_percent"] == 100.0

    def test_computes_percentages_when_missing(self):
        payload = {"students": [{"grade": 5}, {"grade": 4}, {"grade": 2}]}
        summary = build_criteria_subject_summary(payload)
        assert summary["quality_percent"] == round(2 / 3 * 100, 1)
        assert summary["success_percent"] == round(2 / 3 * 100, 1)


class TestFinalAndSimpleTables:
    def test_build_final_table(self):
        final = {
            "columns": [{"key": "q1", "label": "1 четв."}, {"key": "year", "label": "Год"}],
            "students": [
                {"num": 2, "fio": "Беков Б.", "q1": "4", "year": "4"},
                {"num": 1, "fio": "Алиев А.", "q1": "5", "year": "5"},
            ],
        }
        table = build_final_table(final)
        assert table["headers"] == ["№", "ФИО", "1 четв.", "Год"]
        assert table["rows"][0]["cells"] == ["1", "Алиев А.", "5", "5"]

    def test_build_simple_grades_table(self):
        payload = {"students": [{"name": "Алиев А.", "grade": 5}, {"name": "", "grade": 3}]}
        table = build_simple_grades_table(payload)
        assert table["headers"] == ["№", "ФИО", "Оценка"]
        assert table["rows"] == [{"cells": ["1", "Алиев А.", "5"]}]


class TestPeriodsAndPaths:
    def test_period_predicates(self):
        assert is_final_period(6)
        assert not is_final_period(5)
        assert is_year_period(5)
        assert not is_year_period(6)

    def test_criteria_period_path_slug(self):
        assert criteria_period_path_slug(1) == "1_четверть"
        assert criteria_period_path_slug(5) == "учебный_год"
        assert criteria_period_path_slug(6) == "итог"

    def test_safe_path_segment(self):
        assert safe_path_segment('7"А"/Б:класс') == "7_А_Б_класс"
        assert safe_path_segment("") == "unknown"
        assert len(safe_path_segment("х" * 200)) == 80


class TestTableForPeriodPayload:
    def test_quarter_uses_criteria(self):
        payload = {"criteria": _criteria_block()}
        assert has_criteria_data(payload)
        table = table_for_period_payload(2, payload)
        assert table is not None
        assert table["headers"][0] == "№"

    def test_year_uses_simple_table(self):
        payload = {"students": [{"name": "Алиев А.", "grade": 5}]}
        table = table_for_period_payload(5, payload)
        assert table["headers"] == ["№", "ФИО", "Оценка"]

    def test_final_uses_final_block(self):
        payload = {
            "final": {
                "columns": [{"key": "year", "label": "Год"}],
                "students": [{"num": 1, "fio": "Алиев А.", "year": "5"}],
            }
        }
        assert has_final_data(payload)
        table = table_for_period_payload(6, payload)
        assert table["headers"] == ["№", "ФИО", "Год"]

    def test_no_data(self):
        assert table_for_period_payload(2, None) is None
        assert table_for_period_payload(2, {"students": []}) is None


class TestExcelBuilders:
    def _subject_sheets(self):
        return [
            {
                "subject": "Математика",
                "table": build_criteria_table(_criteria_block()),
                "payload": {
                    "students": [{"grade": 5}, {"grade": 4}],
                    "total_students": 2,
                    "quality_percent": 100.0,
                    "success_percent": 100.0,
                    "criteria": _criteria_block(),
                },
                "teacher": "Учитель У.",
            }
        ]

    def test_build_subjects_workbook(self):
        buf = build_subjects_workbook("7А", self._subject_sheets())
        wb = load_workbook(buf)
        assert wb.sheetnames == ["Математика"]
        ws = wb["Математика"]
        assert ws.cell(1, 1).value == "Класс"
        assert ws.cell(1, 2).value == "7А"
        assert ws.cell(9, 1).value == "№"  # заголовок таблицы учеников
        wb.close()

    def test_build_subjects_workbook_empty(self):
        buf = build_subjects_workbook("7А", [])
        wb = load_workbook(buf)
        assert wb.sheetnames == ["Нет данных"]
        wb.close()


class TestCriteriaPeriodZip:
    class _FakeReport:
        def __init__(self, report_id, class_name, subject_name, grades_json):
            self.id = report_id
            self.class_name = class_name
            self.subject_name = subject_name
            self.grades_json = grades_json
            self.teacher = None
            self.teacher_id = None

    def test_zip_structure(self):
        import json

        payload = {
            "students": [{"name": "Алиев А.", "grade": 5}],
            "total_students": 1,
            "criteria": _criteria_block(),
        }
        report = self._FakeReport(1, "7А", "Математика", json.dumps(payload, ensure_ascii=False))
        zip_buf = build_criteria_period_zip(
            "Школа №1", 2, [report], {"7А"}, school_id=None
        )
        assert zip_buf is not None
        with zipfile.ZipFile(zip_buf) as zf:
            names = zf.namelist()
        assert names == ["Школа №1/2_четверть/7А/предметы.xlsx"]

    def test_zip_none_when_no_data(self):
        assert build_criteria_period_zip("Школа", 2, [], set(), school_id=None) is None
