"""
Subject Report Widget — отчёт предметника

По каждому предмету учителя — таблица с классами:
Класс | «5» | «4» | «3» | «2» | Всего | Качество % | Успеваемость %

Плюс данные аналитики СОР/СОЧ (если есть).
"""
from typing import Optional, TYPE_CHECKING
from pathlib import Path
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QComboBox, QLabel, QHeaderView, QScrollArea, QFrame,
    QGroupBox, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont

from .loading_overlay import LoadingOverlay, ApiWorker
from .translator import get_translator

if TYPE_CHECKING:
    from .api_client import MektepAPIClient


class SubjectReportWidget(QWidget):
    """Отчёт предметника"""

    PERIOD_ITEMS = [
        ("1 четверть", "quarter", 1),
        ("2 четверть", "quarter", 2),
        ("3 четверть", "quarter", 3),
        ("4 четверть", "quarter", 4),
    ]

    def __init__(self, api_client: Optional["MektepAPIClient"] = None):
        super().__init__()
        self.api_client = api_client
        self._worker = None
        self.tr = get_translator()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(0, 0, 0, 0)

        # --- Фильтры ---
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel(self.tr.tr('period')))
        self.period_combo = QComboBox()
        for label, *_ in self.PERIOD_ITEMS:
            self.period_combo.addItem(label)
        self.period_combo.setCurrentIndex(1)
        filter_bar.addWidget(self.period_combo)

        filter_bar.addSpacing(15)
        self.refresh_btn = QPushButton(self.tr.tr('refresh'))
        self.refresh_btn.clicked.connect(self.load_data)
        filter_bar.addWidget(self.refresh_btn)

        filter_bar.addSpacing(15)
        self.export_btn = QPushButton(self.tr.tr('export_excel'))
        self.export_btn.setEnabled(False)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #22c55e;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #16a34a;
            }
            QPushButton:disabled {
                background-color: #e5e7eb;
                color: #9ca3af;
            }
        """)
        self.export_btn.clicked.connect(self._export_excel)
        filter_bar.addWidget(self.export_btn)

        filter_bar.addStretch()
        layout.addLayout(filter_bar)

        # --- Прокручиваемый контент ---
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setSpacing(12)
        self.content_layout.setContentsMargins(0, 0, 0, 0)

        self.placeholder = QLabel(self.tr.tr('press_refresh_subject'))
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setStyleSheet("color: #6c757d; font-size: 13px; padding: 40px;")
        self.content_layout.addWidget(self.placeholder)
        self.content_layout.addStretch()

        self.scroll.setWidget(self.content)
        layout.addWidget(self.scroll)

        # --- Оверлей загрузки ---
        self.loading_overlay = LoadingOverlay(self)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setGeometry(self.rect())

    # ------------------------------------------------------------------
    def load_data(self):
        if not self.api_client or not self.api_client.is_authenticated():
            return

        _, period_type, period_number = self.PERIOD_ITEMS[self.period_combo.currentIndex()]

        self.loading_overlay.show_overlay(self.tr.tr('loading_subject_report'))
        self.refresh_btn.setEnabled(False)

        self._worker = ApiWorker(
            self.api_client.get_subject_report,
            period_type=period_type,
            period_number=period_number
        )
        self._worker.finished.connect(self._on_data_loaded)
        self._worker.start()

    def _on_data_loaded(self, result: dict):
        """Callback после загрузки данных"""
        self.loading_overlay.hide_overlay()
        self.refresh_btn.setEnabled(True)

        if not result.get("success"):
            self.export_btn.setEnabled(False)
            return

        subjects = result.get("subjects", [])
        self._last_subjects_data = subjects
        self._render(subjects)
        self.export_btn.setEnabled(bool(subjects))

    # ------------------------------------------------------------------
    def _render(self, subjects):
        # Очищаем
        while self.content_layout.count():
            child = self.content_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if not subjects:
            lbl = QLabel(self.tr.tr('no_data_for_period'))
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("color: #6c757d; padding: 40px;")
            self.content_layout.addWidget(lbl)
            self.content_layout.addStretch()
            return

        bold = QFont()
        bold.setBold(True)

        for subj_data in subjects:
            subj_name = subj_data.get("subject_name", "")
            classes = subj_data.get("classes", [])

            # Карточка предмета
            group = QGroupBox(subj_name)
            group.setStyleSheet("""
                QGroupBox {
                    font-weight: bold; font-size: 13px;
                    border: 1px solid #e5e7eb; border-radius: 8px;
                    margin-top: 1.2em; padding-top: 10px;
                    background: white;
                }
                QGroupBox::title {
                    subcontrol-origin: margin; left: 12px;
                    padding: 0 6px; background: white;
                }
            """)
            g_layout = QVBoxLayout(group)
            g_layout.setContentsMargins(10, 16, 10, 10)

            # --- Таблица оценок ---
            headers = [self.tr.tr('class_col'), "«5»", "«4»", "«3»", "«2»", self.tr.tr('total'), self.tr.tr('quality_percent'), self.tr.tr('success_percent')]
            table = QTableWidget(len(classes), len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.verticalHeader().setVisible(False)
            table.setFixedHeight(40 + len(classes) * 32)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            table.setAlternatingRowColors(True)
            table.setShowGrid(True)
            table.setStyleSheet("""
                QTableWidget {
                    background: white; alternate-background-color: #f9fafb;
                    border: 1px solid #e5e7eb; gridline-color: #e5e7eb; font-size: 12px;
                }
                QHeaderView::section {
                    background: #f3f4f6; border: 1px solid #e5e7eb;
                    padding: 4px; font-weight: 600; font-size: 11px;
                }
            """)

            grade_colors = {
                "count_5": "#d1fae5", "count_4": "#dbeafe",
                "count_3": "#fef3c7", "count_2": "#fecaca",
            }

            for row, cls_data in enumerate(classes):
                # Класс
                table.setItem(row, 0, QTableWidgetItem(cls_data.get("class_name", "")))

                # 5, 4, 3, 2
                for col, key in enumerate(["count_5", "count_4", "count_3", "count_2"]):
                    val = cls_data.get(key, 0)
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setBackground(QColor(grade_colors[key]))
                    if val > 0:
                        item.setFont(bold)
                    table.setItem(row, 1 + col, item)

                # Всего
                total_item = QTableWidgetItem(str(cls_data.get("total", 0)))
                total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                total_item.setFont(bold)
                table.setItem(row, 5, total_item)

                # Качество
                q_item = QTableWidgetItem(f"{cls_data.get('quality_percent', 0)}%")
                q_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                q_item.setBackground(QColor("#d1fae5"))
                table.setItem(row, 6, q_item)

                # Успеваемость
                s_item = QTableWidgetItem(f"{cls_data.get('success_percent', 0)}%")
                s_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                s_item.setBackground(QColor("#dbeafe"))
                table.setItem(row, 7, s_item)

            header = table.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            for c in range(1, len(headers)):
                header.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)

            g_layout.addWidget(table)

            # --- Аналитика СОР/СОЧ (если есть) ---
            for cls_data in classes:
                analytics = cls_data.get("analytics")
                if not analytics:
                    continue

                cls_name = cls_data.get("class_name", "")
                sor_list = analytics.get("sor", [])
                soch = analytics.get("soch")

                if not sor_list and not soch:
                    continue

                a_label = QLabel(f"  Аналитика СОР/СОЧ — {cls_name}")
                a_label.setStyleSheet("color: #4b5563; font-size: 11px; font-weight: bold; margin-top: 4px;")
                g_layout.addWidget(a_label)

                items = []
                for sor in sor_list:
                    items.append(sor)
                if soch:
                    soch_copy = dict(soch)
                    soch_copy["name"] = "СОЧ"
                    items.append(soch_copy)

                a_headers = ["", "«5»", "«4»", "«3»", "«2»"]
                a_table = QTableWidget(len(items), len(a_headers))
                a_table.setHorizontalHeaderLabels(a_headers)
                a_table.verticalHeader().setVisible(False)
                a_table.setFixedHeight(36 + len(items) * 28)
                a_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                a_table.setShowGrid(True)
                a_table.setStyleSheet("""
                    QTableWidget {
                        background: #fafafa; border: 1px solid #e5e7eb;
                        gridline-color: #e5e7eb; font-size: 11px;
                    }
                    QHeaderView::section {
                        background: #f3f4f6; border: 1px solid #e5e7eb;
                        padding: 3px; font-weight: 600; font-size: 10px;
                    }
                """)

                for r, item_data in enumerate(items):
                    a_table.setItem(r, 0, QTableWidgetItem(item_data.get("name", "")))
                    for c, key in enumerate(["count_5", "count_4", "count_3", "count_2"]):
                        val = item_data.get(key, 0)
                        cell = QTableWidgetItem(str(val))
                        cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        cell.setBackground(QColor(grade_colors[key]))
                        a_table.setItem(r, 1 + c, cell)

                a_h = a_table.horizontalHeader()
                a_h.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
                for c in range(1, len(a_headers)):
                    a_h.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)

                g_layout.addWidget(a_table)

            self.content_layout.addWidget(group)

        self.content_layout.addStretch()

    # ------------------------------------------------------------------
    # Экспорт в Excel
    # ------------------------------------------------------------------

    def _export_excel(self):
        """Экспорт отчёта предметника в Excel"""
        if not hasattr(self, '_last_subjects_data') or not self._last_subjects_data:
            return

        period_label = self.period_combo.currentText()
        default_name = f"Предметник_{period_label}.xlsx".replace(" ", "_")

        file_path, _ = QFileDialog.getSaveFileName(
            self, self.tr.tr('save_excel'), str(Path.home() / "Documents" / default_name),
            "Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            wb.remove(wb.active)

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            header_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

            grade_fills = {
                "count_5": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                "count_4": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                "count_3": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                "count_2": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
            }

            for subj_data in self._last_subjects_data:
                subj_name = subj_data.get("subject_name", "Предмет")
                classes = subj_data.get("classes", [])

                # Название листа (макс 31 символ)
                sheet_name = subj_name[:31]
                ws = wb.create_sheet(title=sheet_name)

                headers = [self.tr.tr('class_col'), "5", "4", "3", "2",
                           self.tr.tr('total'), self.tr.tr('quality_percent'),
                           self.tr.tr('success_percent')]
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

                for row_idx, cls_data in enumerate(classes):
                    r = row_idx + 2
                    ws.cell(row=r, column=1, value=cls_data.get("class_name", "")).border = thin_border

                    for col, key in enumerate(["count_5", "count_4", "count_3", "count_2"]):
                        val = cls_data.get(key, 0)
                        cell = ws.cell(row=r, column=2 + col, value=val)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = grade_fills[key]
                        if val > 0:
                            cell.font = Font(bold=True)

                    ws.cell(row=r, column=6, value=cls_data.get("total", 0)).border = thin_border
                    ws.cell(row=r, column=6).alignment = Alignment(horizontal='center')
                    ws.cell(row=r, column=6).font = Font(bold=True)

                    q_cell = ws.cell(row=r, column=7, value=f"{cls_data.get('quality_percent', 0)}%")
                    q_cell.border = thin_border
                    q_cell.alignment = Alignment(horizontal='center')
                    q_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

                    s_cell = ws.cell(row=r, column=8, value=f"{cls_data.get('success_percent', 0)}%")
                    s_cell.border = thin_border
                    s_cell.alignment = Alignment(horizontal='center')
                    s_cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

                # Авто-ширина
                ws.column_dimensions['A'].width = 12
                for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws.column_dimensions[col_letter].width = 14

            wb.save(file_path)
            QMessageBox.information(self, self.tr.tr('success'), self.tr.tr('excel_saved'))
        except Exception as e:
            QMessageBox.critical(self, self.tr.tr('error'), f"{self.tr.tr('excel_save_error')}\n{str(e)}")
