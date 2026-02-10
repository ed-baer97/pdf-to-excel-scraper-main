"""
Grades Widget — сводная таблица оценок по классам

Учитель видит:
1. Список классов, в которых преподаёт
2. При нажатии — таблица ученик × предмет с оценками
"""
from typing import Optional, TYPE_CHECKING
from pathlib import Path
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QComboBox, QLabel, QHeaderView, QStackedWidget, QFrame,
    QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont

from .loading_overlay import LoadingOverlay, ApiWorker
from .translator import get_translator

if TYPE_CHECKING:
    from .api_client import MektepAPIClient


class GradesWidget(QWidget):
    """Сводная таблица оценок по классам"""

    PERIOD_ITEMS = [
        ("1 четверть", "quarter", 1),
        ("2 четверть", "quarter", 2),
        ("3 четверть", "quarter", 3),
        ("4 четверть", "quarter", 4),
    ]

    # Цвета оценок
    GRADE_COLORS = {
        5: "#d1fae5",  # зелёный
        4: "#dbeafe",  # голубой
        3: "#fef3c7",  # жёлтый
        2: "#fecaca",  # красный
    }

    @staticmethod
    def _is_border_percent(pct) -> bool:
        """Пограничный процент: 37-39% (2/3), 61-64% (3/4), 82-84% (4/5)"""
        if pct is None:
            return False
        return (37 <= pct <= 39) or (61 <= pct <= 64) or (82 <= pct <= 84)

    def __init__(self, api_client: Optional["MektepAPIClient"] = None):
        super().__init__()
        self.api_client = api_client
        self._current_class = None
        self._worker = None
        self.tr = get_translator()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(0, 0, 0, 0)

        # --- Панель фильтров ---
        filter_bar = QHBoxLayout()

        filter_bar.addWidget(QLabel(self.tr.tr('period')))
        self.period_combo = QComboBox()
        for label, *_ in self.PERIOD_ITEMS:
            self.period_combo.addItem(label)
        self.period_combo.setCurrentIndex(1)
        self.period_combo.currentIndexChanged.connect(self._on_period_changed)
        filter_bar.addWidget(self.period_combo)

        filter_bar.addSpacing(15)
        filter_bar.addWidget(QLabel(self.tr.tr('class_label')))
        self.class_combo = QComboBox()
        self.class_combo.setMinimumWidth(120)
        self.class_combo.currentIndexChanged.connect(self._on_class_changed)
        filter_bar.addWidget(self.class_combo)

        filter_bar.addSpacing(15)
        self.refresh_btn = QPushButton(self.tr.tr('refresh'))
        self.refresh_btn.clicked.connect(self.load_data)
        filter_bar.addWidget(self.refresh_btn)

        filter_bar.addSpacing(15)
        self.export_btn = QPushButton(self.tr.tr('export_excel'))
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self._export_excel)
        filter_bar.addWidget(self.export_btn)

        filter_bar.addStretch()
        layout.addLayout(filter_bar)

        # --- Стек: заглушка / таблица ---
        self.stack = QStackedWidget()

        # Заглушка
        placeholder = QLabel(self.tr.tr('select_class_for_grades'))
        placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        placeholder.setStyleSheet("color: #6c757d; font-size: 13px; padding: 40px;")
        self.stack.addWidget(placeholder)  # index 0

        # Контейнер таблицы
        table_container = QWidget()
        tc_layout = QVBoxLayout(table_container)
        tc_layout.setContentsMargins(0, 0, 0, 0)

        # Карточки сводки
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet(
            "background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 6px; "
            "padding: 8px 12px; color: #0369a1; font-size: 12px;"
        )
        tc_layout.addWidget(self.summary_label)

        # Таблица
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9fafb;
                border: 1px solid #e5e7eb;
                gridline-color: #e5e7eb;
                font-size: 12px;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                border: 1px solid #e5e7eb;
                padding: 6px 4px;
                font-weight: 600;
                font-size: 11px;
            }
        """)
        tc_layout.addWidget(self.table)

        self.stack.addWidget(table_container)  # index 1
        layout.addWidget(self.stack)

        # --- Оверлей загрузки ---
        self.loading_overlay = LoadingOverlay(self)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setGeometry(self.rect())

    # ------------------------------------------------------------------
    # Загрузка данных
    # ------------------------------------------------------------------

    def load_data(self):
        """Загрузить список классов и обновить комбобокс"""
        if not self.api_client or not self.api_client.is_authenticated():
            return

        self.loading_overlay.show_overlay(self.tr.tr('loading_classes'))
        self.refresh_btn.setEnabled(False)

        self._worker = ApiWorker(self.api_client.get_my_classes)
        self._worker.finished.connect(self._on_classes_loaded)
        self._worker.start()

    def _on_classes_loaded(self, result: dict):
        """Callback после загрузки списка классов"""
        if not result.get("success"):
            self.loading_overlay.hide_overlay()
            self.refresh_btn.setEnabled(True)
            return

        # Собираем уникальные классы
        classes = set()
        for subj in result.get("subjects", []):
            for cls in subj.get("classes", []):
                classes.add(cls["class_name"])

        # Добавляем управляемые классы
        for cls_name in result.get("managed_classes", []):
            classes.add(cls_name)

        sorted_classes = sorted(classes)

        self.class_combo.blockSignals(True)
        prev = self.class_combo.currentText()
        self.class_combo.clear()
        for c in sorted_classes:
            self.class_combo.addItem(c)
        # Восстанавливаем выбор
        idx = self.class_combo.findText(prev)
        if idx >= 0:
            self.class_combo.setCurrentIndex(idx)
        self.class_combo.blockSignals(False)

        if self.class_combo.count() > 0:
            self._load_class_grades()
        else:
            self.loading_overlay.hide_overlay()
            self.refresh_btn.setEnabled(True)

    def _on_period_changed(self, _idx):
        self._load_class_grades()

    def _on_class_changed(self, _idx):
        self._load_class_grades()

    def _load_class_grades(self):
        class_name = self.class_combo.currentText()
        if not class_name or not self.api_client:
            self.stack.setCurrentIndex(0)
            return

        _, period_type, period_number = self.PERIOD_ITEMS[self.period_combo.currentIndex()]

        self.loading_overlay.show_overlay(self.tr.tr('loading_grades', class_name))
        self.refresh_btn.setEnabled(False)

        self._worker = ApiWorker(
            self.api_client.get_class_grades,
            class_name=class_name,
            period_type=period_type,
            period_number=period_number
        )
        self._worker.finished.connect(self._on_grades_loaded)
        self._worker.start()

    def _on_grades_loaded(self, result: dict):
        """Callback после загрузки оценок класса"""
        self.loading_overlay.hide_overlay()
        self.refresh_btn.setEnabled(True)

        if not result.get("success"):
            self.stack.setCurrentIndex(0)
            self.export_btn.setEnabled(False)
            return

        subjects = result.get("subjects", [])
        students = result.get("students", [])
        summary = result.get("summary", {})

        if not students:
            self.stack.setCurrentIndex(0)
            self.export_btn.setEnabled(False)
            return

        self._last_subjects = subjects
        self._last_students = students
        self._last_summary = summary

        self._fill_table(subjects, students, summary,
                         self.class_combo.currentText())
        self.stack.setCurrentIndex(1)
        self.export_btn.setEnabled(True)

    # ------------------------------------------------------------------
    # Заполнение таблицы
    # ------------------------------------------------------------------

    def _fill_table(self, subjects, students, summary, class_name):
        # Сводка
        total = summary.get("total_students", 0)
        quality = summary.get("quality_percent", 0)
        success = summary.get("success_percent", 0)
        period_label = self.period_combo.currentText()
        self.summary_label.setText(
            f"{self.tr.tr('class_col')}: {class_name}  |  {period_label}  |  "
            f"{self.tr.tr('students_count')}: {total}  |  {self.tr.tr('quality_pct')}: {quality}%  |  {self.tr.tr('success_pct')}: {success}%"
        )

        # Колонки: №, ФИО, предмет1, предмет2, ..., Кол-во 5, Кол-во 4, Кол-во 3
        # Переносим длинные названия предметов на несколько строк
        wrapped_subjects = [s.replace(" ", "\n") for s in subjects]
        col_headers = ["№", self.tr.tr('fio')] + wrapped_subjects + ["5", "4", "3"]
        self.table.setColumnCount(len(col_headers))
        self.table.setHorizontalHeaderLabels(col_headers)

        # +5 строк для футера: Кол-во 5, Кол-во 4, Кол-во 3, Качество %, Успеваемость %
        footer_rows = 5
        self.table.setRowCount(len(students) + footer_rows)

        bold_font = QFont()
        bold_font.setBold(True)

        # Статистика по предметам (столбцам)
        subj_stats = {s: {"c5": 0, "c4": 0, "c3": 0, "c2": 0, "total": 0}
                      for s in subjects}

        for row, student in enumerate(students):
            name = student.get("name", "")
            grades = student.get("grades", {})

            # №
            num_item = QTableWidgetItem(str(row + 1))
            num_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 0, num_item)

            # ФИО
            self.table.setItem(row, 1, QTableWidgetItem(name))

            # Оценки по предметам
            count_5 = count_4 = count_3 = 0
            for col_idx, subj in enumerate(subjects):
                grade_info = grades.get(subj, {})
                grade = grade_info.get("grade")
                percent = grade_info.get("percent")

                if grade:
                    text = f"{grade}"
                    if percent is not None:
                        text += f" ({percent}%)"
                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    # Подсветка пограничных процентов: 37-39, 61-64, 82-84
                    if self._is_border_percent(percent):
                        item.setBackground(QColor("#fff3cd"))
                        item.setForeground(QColor("#dc2626"))
                        item.setFont(bold_font)

                    # Статистика по предмету
                    subj_stats[subj]["total"] += 1
                    if grade == 5:
                        count_5 += 1
                        subj_stats[subj]["c5"] += 1
                    elif grade == 4:
                        count_4 += 1
                        subj_stats[subj]["c4"] += 1
                    elif grade == 3:
                        count_3 += 1
                        subj_stats[subj]["c3"] += 1
                    elif grade == 2:
                        subj_stats[subj]["c2"] += 1
                else:
                    item = QTableWidgetItem("—")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setForeground(QColor("#9ca3af"))

                self.table.setItem(row, 2 + col_idx, item)

            # Кол-во 5, 4, 3 по строке
            for offset, count, color in [
                (0, count_5, "#d1fae5"),
                (1, count_4, "#dbeafe"),
                (2, count_3, "#fef3c7"),
            ]:
                item = QTableWidgetItem(str(count))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFont(bold_font)
                item.setBackground(QColor(color))
                self.table.setItem(row, 2 + len(subjects) + offset, item)

        # ==================================================================
        # Футер: итоговые строки по столбцам (предметам)
        # ==================================================================
        footer_defs = [
            (self.tr.tr('count_5'), "c5", "#d1fae5", "#065f46"),
            (self.tr.tr('count_4'), "c4", "#dbeafe", "#1e40af"),
            (self.tr.tr('count_3'), "c3", "#fef3c7", "#92400e"),
        ]

        base_row = len(students)

        for f_idx, (label, key, bg, fg) in enumerate(footer_defs):
            r = base_row + f_idx

            # Пустая ячейка №
            empty = QTableWidgetItem("")
            empty.setBackground(QColor(bg))
            self.table.setItem(r, 0, empty)

            # Название строки
            lbl_item = QTableWidgetItem(label)
            lbl_item.setFont(bold_font)
            lbl_item.setBackground(QColor(bg))
            lbl_item.setForeground(QColor(fg))
            self.table.setItem(r, 1, lbl_item)

            # Значения по предметам
            total_count = 0
            for col_idx, subj in enumerate(subjects):
                val = subj_stats[subj][key]
                total_count += val
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFont(bold_font)
                item.setBackground(QColor(bg))
                item.setForeground(QColor(fg))
                self.table.setItem(r, 2 + col_idx, item)

            # Итого в столбце соответствующей оценки, остальные пустые
            for offset in range(3):
                item = QTableWidgetItem(str(total_count) if offset == f_idx else "")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFont(bold_font)
                item.setBackground(QColor(bg))
                item.setForeground(QColor(fg))
                self.table.setItem(r, 2 + len(subjects) + offset, item)

        # --- Качество % ---
        q_row = base_row + 3
        q_bg = "#d1fae5"
        q_fg = "#065f46"

        empty = QTableWidgetItem("")
        empty.setBackground(QColor(q_bg))
        self.table.setItem(q_row, 0, empty)

        lbl = QTableWidgetItem(self.tr.tr('quality_percent'))
        lbl.setFont(bold_font)
        lbl.setBackground(QColor(q_bg))
        lbl.setForeground(QColor(q_fg))
        self.table.setItem(q_row, 1, lbl)

        for col_idx, subj in enumerate(subjects):
            st = subj_stats[subj]
            t = st["total"]
            pct = round((st["c5"] + st["c4"]) / t * 100, 1) if t > 0 else 0
            item = QTableWidgetItem(f"{pct}%")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFont(bold_font)
            item.setBackground(QColor(q_bg))
            item.setForeground(QColor(q_fg))
            self.table.setItem(q_row, 2 + col_idx, item)

        for offset in range(3):
            item = QTableWidgetItem("")
            item.setBackground(QColor(q_bg))
            self.table.setItem(q_row, 2 + len(subjects) + offset, item)

        # --- Успеваемость % ---
        s_row = base_row + 4
        s_bg = "#dbeafe"
        s_fg = "#1e40af"

        empty = QTableWidgetItem("")
        empty.setBackground(QColor(s_bg))
        self.table.setItem(s_row, 0, empty)

        lbl = QTableWidgetItem(self.tr.tr('success_percent'))
        lbl.setFont(bold_font)
        lbl.setBackground(QColor(s_bg))
        lbl.setForeground(QColor(s_fg))
        self.table.setItem(s_row, 1, lbl)

        for col_idx, subj in enumerate(subjects):
            st = subj_stats[subj]
            t = st["total"]
            pct = round((t - st["c2"]) / t * 100, 1) if t > 0 else 0
            item = QTableWidgetItem(f"{pct}%")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFont(bold_font)
            item.setBackground(QColor(s_bg))
            item.setForeground(QColor(s_fg))
            self.table.setItem(s_row, 2 + col_idx, item)

        for offset in range(3):
            item = QTableWidgetItem("")
            item.setBackground(QColor(s_bg))
            self.table.setItem(s_row, 2 + len(subjects) + offset, item)

        # Авто-ширина
        header = self.table.horizontalHeader()
        # № — фиксированный
        self.table.setColumnWidth(0, 30)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        # ФИО — по содержимому
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        # Предметы — растягиваются
        for c in range(2, 2 + len(subjects)):
            header.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        # "5", "4", "3" — фиксированная узкая ширина
        for c in range(2 + len(subjects), self.table.columnCount()):
            self.table.setColumnWidth(c, 40)
            header.setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)

    # ------------------------------------------------------------------
    # Экспорт в Excel
    # ------------------------------------------------------------------

    def _export_excel(self):
        """Экспорт таблицы оценок в Excel"""
        if not hasattr(self, '_last_students') or not self._last_students:
            return

        class_name = self.class_combo.currentText()
        period_label = self.period_combo.currentText()
        default_name = f"Оценки_{class_name}_{period_label}.xlsx".replace(" ", "_")

        file_path, _ = QFileDialog.getSaveFileName(
            self, self.tr.tr('save_excel'), str(Path.home() / "Documents" / default_name),
            "Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = class_name

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            header_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

            subjects = self._last_subjects
            students = self._last_students
            summary = self._last_summary

            # Заголовок
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(subjects) + 3)
            title_cell = ws.cell(row=1, column=1,
                                 value=f"{class_name} | {period_label} | "
                                       f"Учеников: {summary.get('total_students', 0)} | "
                                       f"Качество: {summary.get('quality_percent', 0)}% | "
                                       f"Успеваемость: {summary.get('success_percent', 0)}%")
            title_cell.font = Font(bold=True, size=11)

            # Заголовки столбцов
            col_headers = ["N", "ФИО"] + subjects + ["5", "4", "3"]
            for col_idx, h in enumerate(col_headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

            # Данные
            grade_fills = {
                5: PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                4: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                3: PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                2: PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
            }
            border_pct_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            red_font = Font(bold=True, color="DC2626")

            # Статистика по предметам (столбцам)
            subj_stats = {s: {"c5": 0, "c4": 0, "c3": 0, "c2": 0, "total": 0}
                          for s in subjects}

            for row_idx, student in enumerate(students):
                r = row_idx + 4
                ws.cell(row=r, column=1, value=row_idx + 1).border = thin_border
                ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=r, column=2, value=student.get("name", "")).border = thin_border

                grades = student.get("grades", {})
                count_5 = count_4 = count_3 = 0

                for col_idx, subj in enumerate(subjects):
                    grade_info = grades.get(subj, {})
                    grade = grade_info.get("grade")
                    percent = grade_info.get("percent")

                    cell = ws.cell(row=r, column=3 + col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

                    if grade:
                        text = str(grade)
                        if percent is not None:
                            text += f" ({percent}%)"
                        cell.value = text

                        if self._is_border_percent(percent):
                            cell.fill = border_pct_fill
                            cell.font = red_font

                        subj_stats[subj]["total"] += 1
                        if grade == 5:
                            count_5 += 1
                            subj_stats[subj]["c5"] += 1
                        elif grade == 4:
                            count_4 += 1
                            subj_stats[subj]["c4"] += 1
                        elif grade == 3:
                            count_3 += 1
                            subj_stats[subj]["c3"] += 1
                        elif grade == 2:
                            subj_stats[subj]["c2"] += 1
                    else:
                        cell.value = "—"

                for offset, count, fill in [
                    (0, count_5, grade_fills[5]),
                    (1, count_4, grade_fills[4]),
                    (2, count_3, grade_fills[3]),
                ]:
                    cell = ws.cell(row=r, column=3 + len(subjects) + offset, value=count)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
                    cell.fill = fill

            # ============================================================
            # Футер: итоговые строки (Кол-во 5/4/3, Качество %, Успеваемость %)
            # ============================================================
            base_row = len(students) + 4  # после данных

            footer_defs = [
                (self.tr.tr('count_5'), "c5", "D1FAE5", "065F46"),
                (self.tr.tr('count_4'), "c4", "DBEAFE", "1E40AF"),
                (self.tr.tr('count_3'), "c3", "FEF3C7", "92400E"),
            ]

            for f_idx, (label, key, bg_hex, fg_hex) in enumerate(footer_defs):
                r = base_row + f_idx
                bg_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                fg_font = Font(bold=True, color=fg_hex)

                # Пустая ячейка N
                c = ws.cell(row=r, column=1, value="")
                c.fill = bg_fill
                c.border = thin_border

                # Название строки
                c = ws.cell(row=r, column=2, value=label)
                c.font = fg_font
                c.fill = bg_fill
                c.border = thin_border

                # Значения по предметам
                total_count = 0
                for col_idx, subj in enumerate(subjects):
                    val = subj_stats[subj][key]
                    total_count += val
                    c = ws.cell(row=r, column=3 + col_idx, value=val)
                    c.alignment = Alignment(horizontal='center')
                    c.font = fg_font
                    c.fill = bg_fill
                    c.border = thin_border

                # Итого в соответствующем столбце оценки
                for offset in range(3):
                    val = total_count if offset == f_idx else ""
                    c = ws.cell(row=r, column=3 + len(subjects) + offset, value=val)
                    c.alignment = Alignment(horizontal='center')
                    c.font = fg_font
                    c.fill = bg_fill
                    c.border = thin_border

            # --- Качество % ---
            q_row = base_row + 3
            q_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            q_font = Font(bold=True, color="065F46")

            ws.cell(row=q_row, column=1, value="").fill = q_fill
            ws.cell(row=q_row, column=1).border = thin_border
            c = ws.cell(row=q_row, column=2, value=self.tr.tr('quality_percent'))
            c.font = q_font
            c.fill = q_fill
            c.border = thin_border

            for col_idx, subj in enumerate(subjects):
                st = subj_stats[subj]
                t = st["total"]
                pct = round((st["c5"] + st["c4"]) / t * 100, 1) if t > 0 else 0
                c = ws.cell(row=q_row, column=3 + col_idx, value=f"{pct}%")
                c.alignment = Alignment(horizontal='center')
                c.font = q_font
                c.fill = q_fill
                c.border = thin_border

            for offset in range(3):
                c = ws.cell(row=q_row, column=3 + len(subjects) + offset, value="")
                c.fill = q_fill
                c.border = thin_border

            # --- Успеваемость % ---
            s_row = base_row + 4
            s_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            s_font = Font(bold=True, color="1E40AF")

            ws.cell(row=s_row, column=1, value="").fill = s_fill
            ws.cell(row=s_row, column=1).border = thin_border
            c = ws.cell(row=s_row, column=2, value=self.tr.tr('success_percent'))
            c.font = s_font
            c.fill = s_fill
            c.border = thin_border

            for col_idx, subj in enumerate(subjects):
                st = subj_stats[subj]
                t = st["total"]
                pct = round((t - st["c2"]) / t * 100, 1) if t > 0 else 0
                c = ws.cell(row=s_row, column=3 + col_idx, value=f"{pct}%")
                c.alignment = Alignment(horizontal='center')
                c.font = s_font
                c.fill = s_fill
                c.border = thin_border

            for offset in range(3):
                c = ws.cell(row=s_row, column=3 + len(subjects) + offset, value="")
                c.fill = s_fill
                c.border = thin_border

            # Авто-ширина
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 25
            for col_idx in range(len(subjects)):
                col_letter = chr(ord('C') + col_idx) if col_idx < 24 else None
                if col_letter:
                    ws.column_dimensions[col_letter].width = 16

            wb.save(file_path)
            QMessageBox.information(self, self.tr.tr('success'), self.tr.tr('excel_saved'))
        except Exception as e:
            QMessageBox.critical(self, self.tr.tr('error'), f"{self.tr.tr('excel_save_error')}\n{str(e)}")
