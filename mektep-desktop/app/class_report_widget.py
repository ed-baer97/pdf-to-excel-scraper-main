"""
Class Report Widget — отчёт классного руководителя

Категоризация учеников класса:
- Отличники (все 5)
- Хорошисты (4-5, нет 3 и 2)
- С одной «4» (близкие к отличникам)
- Троечники (есть 3, нет 2)
- С одной «3» (близкие к хорошистам)
- Неуспевающие (есть 2)
"""
from typing import Optional, TYPE_CHECKING
from pathlib import Path
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QComboBox, QLabel, QHeaderView, QScrollArea, QFrame,
    QGroupBox, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont

from .loading_overlay import LoadingOverlay, ApiWorker
from .translator import get_translator

if TYPE_CHECKING:
    from .api_client import MektepAPIClient


class ClassReportWidget(QWidget):
    """Отчёт классного руководителя"""

    PERIOD_ITEMS = [
        ("1 четверть", "quarter", 1),
        ("2 четверть", "quarter", 2),
        ("3 четверть", "quarter", 3),
        ("4 четверть", "quarter", 4),
    ]

    # Настройки категорий
    CATEGORIES = [
        ("excellent", "Отличники", "#d1fae5", "#065f46"),
        ("good", "Хорошисты", "#dbeafe", "#1e40af"),
        ("one_4", "С одной «4»", "#e0e7ff", "#3730a3"),
        ("satisfactory", "Троечники", "#fef3c7", "#92400e"),
        ("one_3", "С одной «3»", "#fde68a", "#78350f"),
        ("poor", "Неуспевающие", "#fecaca", "#991b1b"),
    ]

    def __init__(self, api_client: Optional["MektepAPIClient"] = None):
        super().__init__()
        self.api_client = api_client
        self._worker = None
        self.tr = get_translator()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(0, 0, 0, 0)

        # --- Фильтры ---
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel(self.tr.tr('period')))
        self.period_combo = QComboBox()
        for label, *_ in self.PERIOD_ITEMS:
            self.period_combo.addItem(label)
        self.period_combo.setCurrentIndex(1)
        filter_bar.addWidget(self.period_combo)

        filter_bar.addSpacing(15)
        self.refresh_btn = QPushButton(self.tr.tr('refresh'))
        self.refresh_btn.clicked.connect(self.load_data)
        filter_bar.addWidget(self.refresh_btn)

        filter_bar.addSpacing(15)
        self.export_btn = QPushButton(self.tr.tr('export_excel'))
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self._export_excel)
        filter_bar.addWidget(self.export_btn)

        filter_bar.addStretch()
        layout.addLayout(filter_bar)

        # --- Прокручиваемый контент ---
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setSpacing(10)
        self.content_layout.setContentsMargins(0, 0, 0, 0)

        self.placeholder = QLabel(self.tr.tr('press_refresh_class_teacher'))
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.placeholder.setStyleSheet("color: #6c757d; font-size: 13px; padding: 40px;")
        self.content_layout.addWidget(self.placeholder)
        self.content_layout.addStretch()

        self.scroll.setWidget(self.content)
        layout.addWidget(self.scroll)

        # --- Оверлей загрузки ---
        self.loading_overlay = LoadingOverlay(self)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.loading_overlay.setGeometry(self.rect())

    # ------------------------------------------------------------------
    def load_data(self):
        if not self.api_client or not self.api_client.is_authenticated():
            return

        _, period_type, period_number = self.PERIOD_ITEMS[self.period_combo.currentIndex()]

        self.loading_overlay.show_overlay(self.tr.tr('loading_class_teacher_report'))
        self.refresh_btn.setEnabled(False)

        self._worker = ApiWorker(
            self.api_client.get_class_teacher_report,
            period_type=period_type,
            period_number=period_number
        )
        self._worker.finished.connect(self._on_data_loaded)
        self._worker.start()

    def _on_data_loaded(self, result: dict):
        """Callback после загрузки данных"""
        self.loading_overlay.hide_overlay()
        self.refresh_btn.setEnabled(True)

        if not result.get("success"):
            self.export_btn.setEnabled(False)
            return

        classes = result.get("classes", [])
        self._last_classes_data = classes
        self._render(classes)
        self.export_btn.setEnabled(bool(classes))

    # ------------------------------------------------------------------
    def _render(self, classes):
        # Очищаем
        while self.content_layout.count():
            child = self.content_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if not classes:
            lbl = QLabel(self.tr.tr('no_data_class_teacher'))
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("color: #6c757d; padding: 40px; font-size: 13px;")
            lbl.setWordWrap(True)
            self.content_layout.addWidget(lbl)
            self.content_layout.addStretch()
            return

        for cls_data in classes:
            cls_name = cls_data.get("class_name", "")
            categories = cls_data.get("categories", {})
            summary = cls_data.get("summary", {})

            # === Карточка класса ===
            class_group = QGroupBox(f"Класс {cls_name}")
            class_group.setStyleSheet("""
                QGroupBox {
                    font-weight: bold; font-size: 14px;
                    border: 1px solid #d1d5db; border-radius: 8px;
                    margin-top: 1.2em; padding-top: 10px;
                    background: white;
                }
                QGroupBox::title {
                    subcontrol-origin: margin; left: 12px;
                    padding: 0 6px; background: white;
                }
            """)
            c_layout = QVBoxLayout(class_group)
            c_layout.setContentsMargins(10, 20, 10, 10)
            c_layout.setSpacing(8)

            # Сводка
            total = summary.get("total_students", 0)
            summary_text = (
                f"Всего учеников: {total}  |  "
                f"Отличники: {summary.get('excellent', 0)}  |  "
                f"Хорошисты: {summary.get('good', 0)}  |  "
                f"С одной «4»: {summary.get('one_4', 0)}  |  "
                f"Троечники: {summary.get('satisfactory', 0)}  |  "
                f"С одной «3»: {summary.get('one_3', 0)}  |  "
                f"Неуспевающие: {summary.get('poor', 0)}"
            )
            sum_label = QLabel(summary_text)
            sum_label.setWordWrap(True)
            sum_label.setStyleSheet(
                "background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 6px; "
                "padding: 8px 12px; color: #0369a1; font-size: 12px; font-weight: normal;"
            )
            c_layout.addWidget(sum_label)

            # Категории
            for cat_key, cat_label, bg_color, text_color in self.CATEGORIES:
                items = categories.get(cat_key, [])
                self._add_category(c_layout, cat_label, items, cat_key, bg_color, text_color)

            self.content_layout.addWidget(class_group)

        self.content_layout.addStretch()

    # ------------------------------------------------------------------
    def _add_category(self, parent_layout, label, items, cat_key, bg_color, text_color):
        """Добавляет секцию категории с таблицей"""
        count = len(items)

        # Заголовок
        hdr = QLabel(f"  {label}: {count}")
        hdr.setStyleSheet(
            f"background: {bg_color}; color: {text_color}; font-weight: bold; "
            f"border-radius: 4px; padding: 4px 8px; font-size: 12px;"
        )
        parent_layout.addWidget(hdr)

        if not items:
            return

        # Таблица содержимого
        if cat_key in ("excellent", "good"):
            # Простой список ФИО
            table = QTableWidget(len(items), 2)
            table.setHorizontalHeaderLabels(["№", "ФИО"])
            table.verticalHeader().setVisible(False)
            for row, item in enumerate(items):
                table.setItem(row, 0, self._centered_item(str(row + 1)))
                table.setItem(row, 1, QTableWidgetItem(item.get("name", "")))
            table.setColumnWidth(0, 35)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        elif cat_key in ("one_4", "one_3"):
            # ФИО + предмет + учитель
            table = QTableWidget(len(items), 4)
            table.setHorizontalHeaderLabels(["№", "ФИО", "Предмет", "Учитель"])
            table.verticalHeader().setVisible(False)
            for row, item in enumerate(items):
                table.setItem(row, 0, self._centered_item(str(row + 1)))
                table.setItem(row, 1, QTableWidgetItem(item.get("name", "")))
                table.setItem(row, 2, QTableWidgetItem(item.get("subject", "")))
                table.setItem(row, 3, QTableWidgetItem(item.get("teacher", "")))
            table.setColumnWidth(0, 35)
            h = table.horizontalHeader()
            h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
            h.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)

        elif cat_key == "satisfactory":
            # ФИО + предметы с тройками
            table = QTableWidget(len(items), 3)
            table.setHorizontalHeaderLabels(["№", "ФИО", "Предметы с «3»"])
            table.verticalHeader().setVisible(False)
            for row, item in enumerate(items):
                table.setItem(row, 0, self._centered_item(str(row + 1)))
                table.setItem(row, 1, QTableWidgetItem(item.get("name", "")))
                subjects = ", ".join(item.get("subjects_with_3", []))
                table.setItem(row, 2, QTableWidgetItem(subjects))
            table.setColumnWidth(0, 35)
            h = table.horizontalHeader()
            h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        elif cat_key == "poor":
            # ФИО + предметы + учителя
            table = QTableWidget(len(items), 3)
            table.setHorizontalHeaderLabels(["№", "ФИО", "Предметы (учитель)"])
            table.verticalHeader().setVisible(False)
            for row, item in enumerate(items):
                table.setItem(row, 0, self._centered_item(str(row + 1)))
                table.setItem(row, 1, QTableWidgetItem(item.get("name", "")))
                subjects_info = "; ".join(
                    f"{s.get('subject', '')} ({s.get('teacher', '')})"
                    for s in item.get("subjects", [])
                )
                table.setItem(row, 2, QTableWidgetItem(subjects_info))
            table.setColumnWidth(0, 35)
            h = table.horizontalHeader()
            h.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            h.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        else:
            return

        table.setMaximumHeight(min(36 + len(items) * 28, 250))
        table.setAlternatingRowColors(True)
        table.setShowGrid(True)
        table.setStyleSheet("""
            QTableWidget {
                background: white; alternate-background-color: #f9fafb;
                border: 1px solid #e5e7eb; gridline-color: #e5e7eb; font-size: 12px;
            }
            QHeaderView::section {
                background: #f3f4f6; border: 1px solid #e5e7eb;
                padding: 4px; font-weight: 600; font-size: 11px;
            }
        """)
        parent_layout.addWidget(table)

    # ------------------------------------------------------------------
    @staticmethod
    def _centered_item(text: str) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        return item

    # ------------------------------------------------------------------
    # Экспорт в Excel
    # ------------------------------------------------------------------

    def _export_excel(self):
        """Экспорт отчёта классного руководителя в Excel"""
        if not hasattr(self, '_last_classes_data') or not self._last_classes_data:
            return

        period_label = self.period_combo.currentText()
        default_name = f"Кл_руководитель_{period_label}.xlsx".replace(" ", "_")

        file_path, _ = QFileDialog.getSaveFileName(
            self, self.tr.tr('save_excel'), str(Path.home() / "Documents" / default_name),
            "Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            wb.remove(wb.active)

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            header_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

            cat_fills = {
                "excellent": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                "good": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                "one_4": PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid"),
                "satisfactory": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                "one_3": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
                "poor": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
            }

            for cls_data in self._last_classes_data:
                cls_name = cls_data.get("class_name", "Класс")
                categories = cls_data.get("categories", {})
                summary = cls_data.get("summary", {})

                sheet_name = f"Класс {cls_name}"[:31]
                ws = wb.create_sheet(title=sheet_name)

                # Сводка
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                sum_cell = ws.cell(row=1, column=1,
                                   value=f"Всего: {summary.get('total_students', 0)} | "
                                         f"Отличники: {summary.get('excellent', 0)} | "
                                         f"Хорошисты: {summary.get('good', 0)} | "
                                         f"Троечники: {summary.get('satisfactory', 0)} | "
                                         f"Неуспевающие: {summary.get('poor', 0)}")
                sum_cell.font = Font(bold=True, size=10)

                current_row = 3

                for cat_key, cat_label, _, _ in self.CATEGORIES:
                    items = categories.get(cat_key, [])
                    fill = cat_fills.get(cat_key, header_fill)

                    # Заголовок категории
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                    cat_cell = ws.cell(row=current_row, column=1, value=f"{cat_label}: {len(items)}")
                    cat_cell.font = Font(bold=True, size=10)
                    cat_cell.fill = fill
                    current_row += 1

                    if not items:
                        current_row += 1
                        continue

                    if cat_key in ("excellent", "good"):
                        headers = ["N", "ФИО"]
                        for col_idx, h in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=h)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        current_row += 1
                        for i, item in enumerate(items):
                            ws.cell(row=current_row, column=1, value=i + 1).border = thin_border
                            ws.cell(row=current_row, column=2, value=item.get("name", "")).border = thin_border
                            current_row += 1

                    elif cat_key in ("one_4", "one_3"):
                        headers = ["N", "ФИО", "Предмет", "Учитель"]
                        for col_idx, h in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=h)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        current_row += 1
                        for i, item in enumerate(items):
                            ws.cell(row=current_row, column=1, value=i + 1).border = thin_border
                            ws.cell(row=current_row, column=2, value=item.get("name", "")).border = thin_border
                            ws.cell(row=current_row, column=3, value=item.get("subject", "")).border = thin_border
                            ws.cell(row=current_row, column=4, value=item.get("teacher", "")).border = thin_border
                            current_row += 1

                    elif cat_key == "satisfactory":
                        headers = ["N", "ФИО", "Предметы с 3"]
                        for col_idx, h in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=h)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        current_row += 1
                        for i, item in enumerate(items):
                            ws.cell(row=current_row, column=1, value=i + 1).border = thin_border
                            ws.cell(row=current_row, column=2, value=item.get("name", "")).border = thin_border
                            subjects = ", ".join(item.get("subjects_with_3", []))
                            ws.cell(row=current_row, column=3, value=subjects).border = thin_border
                            current_row += 1

                    elif cat_key == "poor":
                        headers = ["N", "ФИО", "Предметы (учитель)"]
                        for col_idx, h in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=h)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                        current_row += 1
                        for i, item in enumerate(items):
                            ws.cell(row=current_row, column=1, value=i + 1).border = thin_border
                            ws.cell(row=current_row, column=2, value=item.get("name", "")).border = thin_border
                            subj_info = "; ".join(
                                f"{s.get('subject', '')} ({s.get('teacher', '')})"
                                for s in item.get("subjects", [])
                            )
                            ws.cell(row=current_row, column=3, value=subj_info).border = thin_border
                            current_row += 1

                    current_row += 1

                # Авто-ширина
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 25

            wb.save(file_path)
            QMessageBox.information(self, self.tr.tr('success'), self.tr.tr('excel_saved'))
        except Exception as e:
            QMessageBox.critical(self, self.tr.tr('error'), f"{self.tr.tr('excel_save_error')}\n{str(e)}")
